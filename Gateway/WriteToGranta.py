#--------------------------------------------------------------------------------------------------------------------------------------------------------------------
#
# WriteToGranta.py
#
# PURPOSE: Write Records to Granta MI
#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------
def WriteToGranta(self, window):
    # Import Modules
    from openpyxl import load_workbook
    from powermi import GetParent, GetRecord, WriteRecordData, LinkedFunctional
    import tkinter as tk
    from tkinter import filedialog
    from tkinter import messagebox
    from tkinter import ttk
    import tkinter.font as tkfont 
    import tksheet
    import threading

    # Get Models Table
    self.mod_table = self.db.get_table('Models')

    # Ask the user for the export file
    file = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx')])
    if file is None:
        return
    try:
        wb = load_workbook(file, data_only=True)
    except:
        messagebox.showerror(message="Unable to open the file.")
        return

    # Initialize retrun flag
    self.return_flag = 0    

    # Function to write export files
    def read_data(callback):
        # Get Model Information
        try:
            ws = wb['Model Information']
        except:
            messagebox.showerror(message="Unable to find 'Model Information' sheet in file.")
            callback()
    
        # Extract Model Information
        # -- General Information
        mod_name = ws.cell(row = 3, column = 3).value
        mod_type = ws.cell(row = 4, column = 3).value
        mod_temp = ws.cell(row = 5, column = 3).value
        mod_pd = ws.cell(row = 6, column = 3).value
        mod_notes = ws.cell(row = 7, column = 3).value

        # -- Reversible Information
        rev_model = ws.cell(row = 10, column = 3).value
        rev_mech = ws.cell(row = 11, column = 3).value
        rev_params = []
        row = 14
        while ws.cell(row = row, column=2).value is not None:
            rev_params.append([ws.cell(row = row, column=2).value, float(ws.cell(row = row, column=4).value), ws.cell(row = row, column=3).value])
            row = row + 1
        rev_params = [['' if x is None else x for x in sublist] for sublist in rev_params]

        # -- Irreversible Information
        irrev_model = ws.cell(row = 10, column = 7).value
        irrev_mech = ws.cell(row = 11, column = 7).value
        irrev_params = []
        row = 14
        while ws.cell(row = row, column=6).value is not None:
            irrev_params.append([ws.cell(row = row, column=6).value, float(ws.cell(row = row, column=8).value), ws.cell(row = row, column=7).value])
            row = row + 1
        irrev_params = [['' if x is None else x for x in sublist] for sublist in irrev_params]

        # -- Write the notes
        mod_desc = f"{mod_type} model developed using PyCOMPARE at {mod_temp} °C using the {mod_pd} option. \n"
        if rev_model is not None:
            mod_desc = mod_desc + f"Reversible Model: {rev_model} \n"
        if rev_mech is not None:
            mod_desc = mod_desc + f"Reversible Mechanisms: {rev_mech} \n"
        if rev_model is not None:
            mod_desc = mod_desc + f"Irreversible Model: {irrev_model} \n"
        if rev_mech is not None:
            mod_desc = mod_desc + f"Irreversible Mechanisms: {irrev_mech} \n"
        if mod_notes is not None:
            mod_desc = mod_desc + "/n/n" + mod_notes


        # Extract Test Information
        try:
            ws = wb['Test Information']
        except:
            messagebox.showerror(message="Unable to find 'Test Information' sheet in file.")
            return
        
        test_info = []
        row = 3
        while ws.cell(row = row, column = 2).value is not None:
            test_info.append([ws.cell(row = row, column = 2).value,
                            ws.cell(row = row, column = 3).value,
                            ws.cell(row = row, column = 4).value,
                            ws.cell(row = row, column = 5).value,
                            ws.cell(row = row, column = 6).value,
                            ws.cell(row = row, column = 7).value,
                            ws.cell(row = row, column = 8).value,
                            ws.cell(row = row, column = 9).value,
                            ws.cell(row = row, column = 10).value,
                            ws.cell(row = row, column = 11).value])
            row = row + 1
        if len(test_info) == 0:
            messagebox.showerror(message="No tests found in'Test Information' sheet in file.")
            callback()
        
        # -- Try to get material name and class from test records
        mat_name = None
        mat_class = None
        mat_group = None
        flag = 0
        while flag == 0:
            for test in test_info:
                test_name = test[0]
                attribute = self.table.attributes['Specimen ID']
                search = attribute.search_criterion(contains=test_name)
                records = self.table.search_for_records_where([search])

                if len(records) > 0:
                    record = records[0]
                    if record.attributes['Material Name'].value is not None:
                        mat_name = record.attributes['Material Name'].value
                    if record.attributes['Material Name'].value is not None:
                        mat_class = record.attributes['Material Class'].value
                    mat_group = record.path[1]
                
                if mat_name is not None and mat_class is not None:
                    flag = 1
                    break

            flag = 1
            break

        # Extract Prediction Information
        pred_info = {}
        for test in test_info:
            pred_info[test[0]] = {}
        try:
            ws = wb['Predicted Curves']
        except:
            messagebox.showerror(message="Unable to find 'Predicted Curves' sheet in file.")
            callback()
        
        # -- Write prediction data to dictionary
        for test in pred_info.keys():
            # Find the column where the test exists
            max_col = ws.max_column+1
            for i in range(1,max_col):
                if ws.cell(row = 2, column = i).value == test:
                    start_col = i
                    while ws.cell(row = 2, column = start_col).value is None or ws.cell(row = 2, column = start_col).value == test:
                        array_name = ws.cell(row = 3, column = start_col).value.split('(')[0].strip()
                        array_val = []
                        start_row = 4
                        while ws.cell(row = start_row, column = start_col).value is not None:
                            array_val.append(float(ws.cell(row = start_row, column = start_col).value))
                            start_row = start_row + 1
                        pred_info[test][array_name] = array_val
                        start_col = start_col + 1
                        if start_col == max_col:
                            break

        # Extract Reduced Test Information
        red_info = {}
        for test in test_info:
            red_info[test[0]] = {}
        try:
            ws = wb['Reduced Curves']
        except:
            messagebox.showerror(message="Unable to find 'Reduced Curves' sheet in file.")
            callback()
        
        # -- Write reduced data to dictionary
        for test in red_info.keys():
            # Find the column where the test exists
            max_col = ws.max_column+1
            for i in range(1,max_col):
                if ws.cell(row = 2, column = i).value == test:
                    start_col = i
                    while ws.cell(row = 2, column = start_col).value is None or ws.cell(row = 2, column = start_col).value == test:
                        array_name = ws.cell(row = 3, column = start_col).value.split('(')[0].strip()
                        array_val = []
                        start_row = 4
                        while ws.cell(row = start_row, column = start_col).value is not None:
                            array_val.append(float(ws.cell(row = start_row, column = start_col).value))
                            start_row = start_row + 1
                        red_info[test][array_name] = array_val
                        start_col = start_col + 1
                        if start_col == max_col:
                            break
    
        # Notify when done
        self.return_flag = 1 
        self.read_data = mat_name, mat_group, mat_class, mod_name, mod_name, mod_desc, rev_model, rev_mech, rev_params, irrev_model, irrev_mech, irrev_params, test_info, pred_info, red_info
        callback()
    
    # Function to display progress bar while saving
    def show_read_window():

        # Create the window
        loading = tk.Toplevel(window)
        loading.title("Reading Data")
        loading.geometry("350x100")
        loading.resizable(False, False)
        loading.configure(bg='white')
        loading.grab_set()  

        # Function for progress bar Exit Protocol
        def on_closing_saving(self):

            # Don't allow exit while saving
            return
        
        # Create the window exit protocal
        loading.protocol("WM_DELETE_WINDOW", lambda:on_closing_saving(self))

        # Create the loading label
        ttk.Label(
                loading, 
                text="Reading the Model Data - Please Wait.", 
                style = "Modern2.TLabel"
                ).pack(pady=10)

        # Create the progress bar
        pb = ttk.Progressbar(
                            loading, 
                            mode='indeterminate',
                            style = "Modern.Horizontal.TProgressbar"
                            )
        pb.pack(fill='x', padx=20, pady=10)
        pb.start(10)

        # Function to close window when task is completed
        def on_task_done():

            # Stop Progress bar
            pb.stop()

            # Destroy Window
            loading.destroy()

        # Begin save on background thread
        threading.Thread(target=read_data, args=(on_task_done,), daemon=True).start()

        # Wait until loading window is closed
        window.wait_window(loading)

    # Start Search
    show_read_window()
    
    if self.return_flag == 0:
        return
    
    # Get Data
    mat_name, mat_group, mat_class, mod_name, mod_name, mod_desc, rev_model, rev_mech, rev_params, irrev_model, irrev_mech, irrev_params, test_info, pred_info, red_info = self.read_data

        
    # Create Model Information Sheet
    Cols = ['Attribute', 'Value']
    table_atts = ['Material Name', 'Material Group', 'Material Class', 'Model Name', 'Model ID', 'Model Description', 'Software']
    table_vals = [mat_name, mat_group, mat_class, mod_name, mod_name, mod_desc, 'PyCOMPARE']
    self.exp_sheet = tksheet.Sheet(
                                    window, 
                                    total_rows = len(table_atts), 
                                    total_columns = len(Cols), 
                                    headers = Cols,
                                    show_x_scrollbar = False, 
                                    show_y_scrollbar = True,
                                    font = ("Segoe UI",max([self.min_font, int(12*self.scale)]),"normal"),
                                    header_font = ("Segoe UI",max([self.min_font, int(12*self.scale)]),"bold"),
                                    #table_bg = 'red' # For checking formatting only
                                    )
    self.exp_sheet.place(
                        anchor = 'nw', 
                        relx = self.Placement['Gateway']['ExpSheet'][0], 
                        rely = self.Placement['Gateway']['ExpSheet'][1],
                        relwidth = self.Placement['Gateway']['ExpSheet'][2],
                        relheight = self.Placement['Gateway']['ExpSheet'][3],
                        )
    self.att_list.append('self.exp_sheet')

    # -- Format the table
    self.exp_sheet.change_theme("blue")
    window.update_idletasks()
    total_width = self.exp_sheet.winfo_width()
    self.exp_sheet.column_width(column = 0, width = int(total_width*self.Placement['Gateway']['ExpSheet'][4]), redraw = True)
    self.exp_sheet.column_width(column = 1, width = int(total_width*self.Placement['Gateway']['ExpSheet'][5]), redraw = True)

    # -- Populate the data
    for i, att in enumerate(table_atts):
        self.exp_sheet.set_cell_data(i,0,att)
        self.exp_sheet.set_cell_data(i,1,table_vals[i])


    # -- Resize model description row
    font = ("Segoe UI",max([self.min_font, int(12*self.scale)]),"normal") 
    font_metrics = tkfont.Font(font=font)
    line_height = font_metrics.metrics("linespace")
    line_count = mod_desc.count("\n") + 1
    self.exp_sheet.row_height(row = table_atts.index('Model Description'), height = line_count * line_height + 4)

    # -- Enable Bindings
    self.exp_sheet.enable_bindings('single_select','cell_select','row_select', "edit_cell")


    # Create Test Summary Sheet
    Cols = ['Test', 'Test Type', 'Load Rate', 'Simulation Type', 'Error']
    self.test_sheet = tksheet.Sheet(
                                    window, 
                                    total_rows = len(test_info), 
                                    total_columns = len(Cols), 
                                    headers = Cols,
                                    show_x_scrollbar = False, 
                                    show_y_scrollbar = True,
                                    font = ("Segoe UI",max([self.min_font, int(12*self.scale)]),"normal"),
                                    header_font = ("Segoe UI",max([self.min_font, int(12*self.scale)]),"bold"),
                                    #table_bg = 'red' # For checking formatting only
                                    )
    self.test_sheet.place(
                        anchor = 'nw', 
                        relx = self.Placement['Gateway']['TestSheet'][0], 
                        rely = self.Placement['Gateway']['TestSheet'][1],
                        relwidth = self.Placement['Gateway']['TestSheet'][2],
                        relheight = self.Placement['Gateway']['TestSheet'][3],
                        )
    self.att_list.append('self.test_sheet')

    # -- Format the table
    self.test_sheet.change_theme("blue")
    window.update_idletasks()
    total_width = self.test_sheet.winfo_width()
    self.test_sheet.column_width(column = 0, width = int(total_width*self.Placement['Gateway']['TestSheet'][4]), redraw = True)
    self.test_sheet.column_width(column = 1, width = int(total_width*self.Placement['Gateway']['TestSheet'][5]), redraw = True)
    self.test_sheet.column_width(column = 2, width = int(total_width*self.Placement['Gateway']['TestSheet'][6]), redraw = True)
    self.test_sheet.column_width(column = 3, width = int(total_width*self.Placement['Gateway']['TestSheet'][6]), redraw = True)
    self.test_sheet.column_width(column = 4, width = int(total_width*self.Placement['Gateway']['TestSheet'][7]), redraw = True)


    # -- Populate the data
    for i in range(len(test_info)):
        self.test_sheet.set_cell_data(i,0,test_info[i][0])
        self.test_sheet.set_cell_data(i,1,test_info[i][1])
        self.test_sheet.set_cell_data(i,2,str('{:0.2e}'.format(test_info[i][4])) + ' ' + test_info[i][5])
        self.test_sheet.set_cell_data(i,3,test_info[i][7])
        self.test_sheet.set_cell_data(i,4,test_info[i][9])

    # -- Enable Bindings
    self.test_sheet.enable_bindings('single_select','cell_select','row_select')
        
    
    # Import Function
    def import_to_db():
        def import_data(callback):
            # Set Flag
            self.return_flag = 0

            # Check for empty values
            for i in range(len(self.exp_sheet.data)):
                if self.exp_sheet.data[i][1] == '':
                    self.msg = f"Empty Value for {self.exp_sheet.data[i][0]}."
                    callback()
                
            # Write Table Values to Dictionary
            # -- Write Single Value Attributes
            RecData = {}
            for i in range(len(self.exp_sheet.data)):
                RecData[self.exp_sheet.data[i][0]] = {'Value': self.exp_sheet.data[i][1]}
                
            # -- Write Array Data
            RecData['Simulation Summary'] = {
                                            'Columns':['Notes', 'Linking Value (Model ID)'],
                                            'Value':[],
                                            'Units':['','']
                                            }
            
            for test in pred_info.keys():
                idx = list(pred_info.keys()).index(test)
                note = f"Test Type: {test_info[idx][1]} \n Control Mode: {test_info[idx][3]} \n Load Rate: {test_info[idx][4]} {test_info[idx][5]} \n Weight: {test_info[idx][9]}"
                RecData['Simulation Summary']['Value'].append([note, RecData['Model ID']['Value'] + '_' + test])

            # -- Store Attribute Data
            if len(rev_params) > 0:
                RecData['Viscoelastic Constitutive Model (Define)'] = {
                                                            'Columns':['Parameter', 'Value', 'Unit'],
                                                            'Value':rev_params,
                                                            'Units':['', '', '']
                                                            }
            if len(irrev_params) > 0:
                RecData['Viscoplastic Constitutive Model (Define)'] = {
                                                            'Columns':['Parameter', 'Value', 'Unit'],
                                                            'Value':irrev_params,
                                                            'Units':['', '', '']
                                                            }
                
            # Check for unique ID
            self.loading_label.configure(text = "Checking for unique Model ID")
            self.loading.update_idletasks()

            attribute = self.mod_table.attributes['Model ID']
            search = attribute.search_criterion(contains=RecData['Model ID']['Value'])
            records = self.mod_table.search_for_records_where([search])
            record = None
            if len(records) > 0:
                askyn = messagebox.askyesno( message = f"Model ID '{RecData['Model ID']['Value']}' already exists - do you want to overwrite it?.")
                if askyn:
                    record = records[0]

                else:
                    self.msg = ''
                    callback()
            
            if record is None:
                # Get Record
                tree = [
                        RecData['Material Class']['Value'],
                        RecData['Material Group']['Value'],
                        RecData['Material Name']['Value'],
                        ]
                
                folder, GUIDS, flag, msg = GetParent(self.mi, self.db, self.mod_table, tree)
                record = GetRecord(self.mi, self.db, self.mod_table, RecData['Model Name']['Value'], parent = folder, short_name = RecData['Model Name']['Value'])

            # Write Data
            self.loading_label.configure(text = "Writing Record Data")
            self.loading.update_idletasks()

            att_list = list(RecData.keys())
            att_list.remove('Material Group')
            record = WriteRecordData(self.mi, self.db, record, RecData, attributes=att_list)

            # Write Simulation records
            for i, test in enumerate(pred_info.keys()):
                # Update loading bar
                self.loading_label.configure(text = f"Writing Simulation Record {i+1} of {len(pred_info.keys())}")
                self.loading.update_idletasks()

                # Get the record
                sim_record = GetRecord(self.mi, self.db, self.mod_table, test, parent = record, short_name = test)

                # Write Simulation DAta
                SimRecData = {}

                # -- Create a model ID
                SimRecData['Model ID'] = {'Value':RecData['Model ID']['Value'] + '_' + test}

                # -- Get the simulation type
                SimRecData['Simulation Type'] = {'Value':None}
                if test_info[list(pred_info.keys()).index(test)][7] == 'Characterization':
                    SimRecData['Simulation Type']['Value'] = 'Characterization'
                else:
                    SimRecData['Simulation Type']['Value'] = 'Prediction'
                
                # -- Get the simulation error
                SimRecData['Model Error'] = {'Value':test_info[list(pred_info.keys()).index(test)][9],
                                            'Units':''}
                
                # -- Determine Functional data
                x_arrays = ['Time', 'Strain (11 axis)', 'Strain (22 axis)', 'Strain (33 axis)', 'Strain (12 axis)', 'Strain (13 axis)', 'Strain (23 axis)',]
                y_arrays = ['Strain (11 axis)', 'Strain (22 axis)', 'Strain (33 axis)', 'Strain (12 axis)', 'Strain (13 axis)', 'Strain (23 axis)',
                            'Stress (11 axis)', 'Stress (22 axis)', 'Stress (33 axis)', 'Stress (12 axis)', 'Stress (13 axis)', 'Stress (23 axis)',]
                
                for y in y_arrays:
                    for x in x_arrays:
                        if x != 'Time':
                            xval = x.split('(')[0].strip() + '-' + x.split('(')[1].split(' ')[0].strip()
                        else:
                            xval = x
                        yval = y.split('(')[0].strip() + '-' + y.split('(')[1].split(' ')[0].strip()

                        if f"{y} vs {x}" in sim_record.attributes.keys():

                            # -- Prediction Data
                            if xval in pred_info[test].keys() and yval in pred_info[test].keys():
                                if f"{y} vs {x}" not in SimRecData.keys():
                                    SimRecData[f"{y} vs {x}"] = {'Y':{
                                                                    'Value':[],
                                                                    'Units':[],
                                                                    },
                                                                'X':{
                                                                    'Value':[],
                                                                    'Units':[],
                                                                    },
                                                                'Simulation Data':{
                                                                    'Value':[],
                                                                    },
                                                                }
                                                
                                if 'Strain' in yval:
                                    SimRecData[f"{y} vs {x}"]['Y']['Units'] = '%'
                                    yfactor = 100
                                else:
                                    SimRecData[f"{y} vs {x}"]['Y']['Units'] = 'MPa'
                                    yfactor = 1

                                if 'Strain' in xval:
                                    SimRecData[f"{y} vs {x}"]['X']['Units'] = '%'
                                    xfactor = 100
                                else:
                                    SimRecData[f"{y} vs {x}"]['X']['Units'] = 's'
                                    xfactor = 1
                                    
                                for j in range(len(pred_info[test][yval])):
                                    SimRecData[f"{y} vs {x}"]['Y']['Value'].append(pred_info[test][yval][j]*yfactor)
                                    SimRecData[f"{y} vs {x}"]['X']['Value'].append(pred_info[test][xval][j]*xfactor)
                                    SimRecData[f"{y} vs {x}"]['Simulation Data']['Value'].append('Simulation')

                            # -- Reduced Data
                            if xval in red_info[test].keys() and yval in red_info[test].keys():
                                if f"{y} vs {x}" not in SimRecData.keys():
                                    SimRecData[f"{y} vs {x}"] = {'Y':{
                                                                    'Value':[],
                                                                    'Units':[],
                                                                    },
                                                                'X':{
                                                                    'Value':[],
                                                                    'Units':[],
                                                                    },
                                                                'Simulation Data':{
                                                                    'Value':[],
                                                                    },
                                                                }
                                
                                if 'Strain' in yval:
                                    SimRecData[f"{y} vs {x}"]['Y']['Units'] = '%'
                                    yfactor = 100
                                else:
                                    SimRecData[f"{y} vs {x}"]['Y']['Units'] = 'MPa'
                                    yfactor = 1

                                if 'Strain' in xval:
                                    SimRecData[f"{y} vs {x}"]['X']['Units'] = '%'
                                    xfactor = 100
                                else:
                                    SimRecData[f"{y} vs {x}"]['X']['Units'] = 's'
                                    xfactor = 1
                                    
                                for j in range(len(red_info[test][yval])):
                                    SimRecData[f"{y} vs {x}"]['Y']['Value'].append(red_info[test][yval][j]*yfactor)
                                    SimRecData[f"{y} vs {x}"]['X']['Value'].append(red_info[test][xval][j]*xfactor)
                                    SimRecData[f"{y} vs {x}"]['Simulation Data']['Value'].append('Reduced Test')

                                # -- Add Linked Functional Attribute
                                SimRecData[f"{y} vs {x} Functional Linking Data"] = {
                                                                                    'Columns':['Table', 'Attribute', 'Linking Attribute', 'Linking Value', 'Label'],
                                                                                    'Value':[['Test Data: Thermomechanical', f"{y} vs {x}", 'Specimen ID', test, 'Test']],
                                                                                    'Units':['','','','','']
                                                                                    }


                # Write data
                sim_record = WriteRecordData(self.mi, self.db, sim_record, SimRecData, attributes=list(SimRecData.keys()))

            # Update Linked Functional Attributes
            self.loading_label.configure(text = "Updating Linked Functional Data")
            self.loading.update_idletasks()

            self.msg = LinkedFunctional(self.mi, self.db, self.mod_table)
            
            # Show message
            if "ERROR" in self.msg:
                # Show error message
                callback()

            else:
                # Show success message
                self.return_flag = 1
                self.msg = "Record(s) Added to the Database!"
                callback()
        
        # Function to display progress bar while saving
        def show_import_window():

            # Create the window
            self.loading = tk.Toplevel(window)
            self.loading.title("Write to Database")
            self.loading.geometry("350x100")
            self.loading.resizable(False, False)
            self.loading.configure(bg='white')
            self.loading.grab_set()  

            # Function for progress bar Exit Protocol
            def on_closing_saving(self):

                # Don't allow exit while saving
                return
            
            # Create the window exit protocal
            self.loading.protocol("WM_DELETE_WINDOW", lambda:on_closing_saving(self))

            # Create the loading label
            self.loading_label = ttk.Label(
                    self.loading, 
                    text="Reading the Model Data - Please Wait.", 
                    style = "Modern2.TLabel"
                    )
            self.loading_label.pack(pady=10)

            # Create the progress bar
            pb = ttk.Progressbar(
                                self.loading, 
                                mode='indeterminate',
                                style = "Modern.Horizontal.TProgressbar"
                                )
            pb.pack(fill='x', padx=20, pady=10)
            pb.start(10)

            # Function to close window when task is completed
            def on_task_done():

                # Stop Progress bar
                pb.stop()

                # Destroy Window
                self.loading.destroy()

            # Begin save on background thread
            threading.Thread(target=import_data, args=(on_task_done,), daemon=True).start()

            # Wait until loading window is closed
            window.wait_window(self.loading)

        # Start Search
        show_import_window()

        if self.msg != '':
            if self.return_flag == 0:
                messagebox.showerror(message= self.msg)
            else:
                messagebox.showinfo(message=self.msg)

    # Create Import Button
    self.btn_mod_imp = ttk.Button(
                            window, 
                            text = "Import to DB", 
                            command = import_to_db,
                            style = 'Modern2.TButton',
                            )
    self.btn_mod_imp.place(
                    anchor = 'n',  
                    relx = self.Placement['Gateway']['ButtonImpMod'][0], 
                    rely = self.Placement['Gateway']['ButtonImpMod'][1],
                    relwidth = self.Placement['Gateway']['ButtonImpMod'][2],
                    relheight = self.Placement['Gateway']['ButtonImpMod'][3],
                    )
    self.att_list.append('self.btn_mod_imp')
    
    # Create Back Button
    self.btn_back = ttk.Button(
                            window, 
                            text = "Back", 
                            command = self.back,
                            style = 'Modern2.TButton',
                            )
    self.btn_back.place(
                    anchor = 'nw',  
                    relx = self.Placement['Gateway']['ButtonBack'][0], 
                    rely = self.Placement['Gateway']['ButtonBack'][1],
                    relwidth = self.Placement['Gateway']['ButtonBack'][2],
                    relheight = self.Placement['Gateway']['ButtonBack'][3],
                    )
    self.att_list.append('self.btn_back')



