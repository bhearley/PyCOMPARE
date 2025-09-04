#--------------------------------------------------------------------------------------------------------------------------------------------------------------------
#
# ExportData.py
#
# PURPOSE: Export record data to excel for PyCOMPARE
#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------
def ExportData(record, temp_path):
    # Import Modules
    from Model.UnitConversion import UnitConversion
    from openpyxl import load_workbook
    import os
    import numpy as np

    # Initialize messsage
    msg = ''
    flag = 0

    # Open the template file
    wb = load_workbook(os.path.join(os.getcwd(),'Templates','ImportTemplate.xlsx'))

    # Open the worksheet
    ws = wb['Data']

    # Get the test direction
    test_dir = record.attributes['Material Test Direction'].value
    if test_dir is None:
        flag = 1
        msg = msg + f"Material Test Direction not defined in {record.name} \n"

    # Get the test summary
    test_summary = record.attributes['Test Summary']

    # Write the Test information
    # -- Name
    if record.attributes['Specimen ID'].value is not None:
        ws.cell(row = 3, column = 3).value = record.attributes['Specimen ID'].value
    else:
        flag = 1
        msg = msg + f"Specimen ID not defined in {record.name} \n"

    # -- Test Type
    if record.attributes['Test Type'].value is not None:
        ws.cell(row = 4, column = 3).value = record.attributes['Test Type'].value
        if record.attributes['Test Type'].value == 'Stress Relaxation':
            ws.cell(row = 4, column = 3).value = 'Relaxation'
    else:
        flag = 1
        msg = msg + f"Test Type not defined in {record.name} \n"

    # -- Test Temperature
    try:
        ws.cell(row = 5, column = 3).value = round(
                                                    UnitConversion(
                                                            record.attributes['Test Temperature'].unit, 
                                                            record.attributes['Test Temperature'].value,
                                                            "°C"
                                                            )
                                                    )
    except:
        flag = 1
        msg = msg + f"Test Temperature not defined in {record.name} \n"
    
    # -- Anisotropy Angle
    try:
        ws.cell(row = 6, column = 3).value = round(
                                                    UnitConversion(
                                                            record.attributes['Orientation'].unit, 
                                                            record.attributes['Orientation'].value,
                                                            "°"
                                                            )
                                                    ,1)
    except:
        flag = 1
        msg = msg + f"Orientation not defined in {record.name} \n"
    
    # -- Control Mode
    try:
        control_mode = test_summary.value[0][test_summary.columns.index('Control Mode')][0]
        if test_dir == '11':
            ws.cell(row = 7, column = 3).value = control_mode
        elif test_dir == '22':
            ws.cell(row = 8, column = 3).value = control_mode
        else:
            ws.cell(row = 9, column = 3).value = control_mode
    except:
        flag = 1
        msg = msg + f"Test Summary not defined in {record.name} \n"

    # Determine the max time and stage end times
    try:
        end_times = []
        one_row = ['Tensile', 'Compressive', 'Shear']
        two_rows = ['Creep', 'Stress Relaxation']
        if record.attributes['Test Type'].value in one_row:
            max_rows = 1
        elif record.attributes['Test Type'].value in two_rows:
            max_rows = 2
        else:
            max_rows = test_summary.shape[1]

        for i in range(max_rows):
            # -- Write End Time
            end_time = test_summary.value[i][test_summary.columns.index('End Time')][0]
            end_time_unit = test_summary.units.data[i][test_summary.columns.index('End Time')]
            end_time = UnitConversion(end_time_unit, end_time, 's')
            end_times.append(end_time)
    except:
        flag = 1
        msg = msg + f"Could not determine end times from the Test Summary in {record.name} \n"
       
    # Write Functional Data
    # -- Get Strain vs Time data arrays
    try:
        data = record.attributes['Strain (' + test_dir + ' axis) vs Time']
        time = [row[2] for row in data.value][1:]
        time_unit = data.parameters['Time'].unit
        if time_unit != 's':
            time = UnitConversion(time_unit, np.array(time), 's')
        idx = np.where(np.array(time) <= end_times[-1])[0][-1]
        time = time[:idx+1]
        
        strain = [row[0] for row in data.value][1:]
        strain_unit = data.unit
        if strain_unit != '-':
            strain = UnitConversion(strain_unit, np.array(strain), '-')
        strain = strain[:idx+1]
    except:
        flag = 1
        msg = msg + f"No Strain vs Time Data given in {record.name} \n"

    # -- Get Stress vs Time data arrays
    try:
        data = record.attributes['Stress (' + test_dir + ' axis) vs Time']
        stress = [row[0] for row in data.value][1:]
        stress_unit = data.unit
        if stress_unit != 'MPa':
            stress = UnitConversion(stress_unit, np.array(stress), 'MPa')
        stress = stress[:idx+1]
    except:
        flag = 1
        msg = msg + f"No Stress vs Time Data given in {record.name} \n"

    # -- Check for Poisson Strain Data
    try:
        pr_dir = None
        if test_dir == '11':
            pr_dir = '22'
        if test_dir == '22':
            pr_dir = '11'
        if pr_dir is not None:
            data = record.attributes['Strain (' + pr_dir + ' axis) vs Time']
            trans_strain = []
            if len(data.value) >= len(time):
                trans_strain = [row[0] for row in data.value][1:]
                trans_strain_unit = data.unit
                if trans_strain_unit != '-':
                    trans_strain = UnitConversion(trans_strain_unit, np.array(trans_strain), '-')
                trans_strain = trans_strain[:idx+1]
    except:
        pass

    if flag == 0:
        # -- Get columns
        time_col = 12
        if test_dir == '11':
            strain_col = 13
            stress_col = 14
            trans_strain_col = 15
        if test_dir == '22':
            strain_col = 15
            stress_col = 16
            trans_strain_col = 13
        if test_dir == '12':
            strain_col = 17
            stress_col = 18

        # -- Write Data
        for i in range(len(time)):
            ws.cell(row = 3+i, column = time_col).value = time[i]
            ws.cell(row = 3+i, column = strain_col).value = strain[i]
            ws.cell(row = 3+i, column = stress_col).value = stress[i]
            if pr_dir is not None:
                if len(trans_strain) > 0:
                    ws.cell(row = 3+i, column = trans_strain_col).value = trans_strain[i]


        # Write Stage Table Information
        for i in range(max_rows):
            # -- Write End Time
            ws.cell(row = 4+i, column = 5).value = end_times[i]

            # -- Get the index of the current and previous stage
            idx_c = np.where(np.array(time) <= end_times[i])[0][-1]
            if i == 0:
                idx_p = 0
            else:
                idx_p = np.where(np.array(time) <= end_times[i-1])[0][-1]

            # -- Write Loading Direction
            ws.cell(row = 4+i, column = 6).value = test_dir

            # -- Write Control Mode
            ws.cell(row = 4+i, column = 7).value = test_summary.value[i][test_summary.columns.index('Control Mode')][0]

            # -- Write Loading Rate
            if test_summary.value[i][test_summary.columns.index('Control Mode')][0] == 'Strain':
                load_rate = test_summary.value[i][test_summary.columns.index('Strain Rate (' + test_dir + ' axis)')][0]
                load_rate_unit = test_summary.units.data[i][test_summary.columns.index('Strain Rate (' + test_dir + ' axis)')]
                load_rate = UnitConversion(load_rate_unit, load_rate, '1/s')
            else:
                load_rate = test_summary.value[i][test_summary.columns.index('Stress Rate (' + test_dir + ' axis)')][0]
                load_rate_unit = test_summary.units.data[i][test_summary.columns.index('Stress Rate (' + test_dir + ' axis)')]
                load_rate = UnitConversion(load_rate_unit, load_rate, 'MPa/s')
            ws.cell(row = 4+i, column = 8).value = load_rate

            # -- Write Target
            # ---- Creep or Relaxation - Target is time
            if test_summary.value[i][test_summary.columns.index('Stage Type')][0] == 'Creep' or test_summary.value[i][test_summary.columns.index('Stage Type')][0] == 'Relaxation':
                ws.cell(row = 4+i, column = 9).value = 'Time'
                ws.cell(row = 4+i, column = 10).value = time[idx_c] - time[idx_p] 
            else:
                # ---- Strain Target Value
                if test_summary.value[i][test_summary.columns.index('Target Strain (' + test_dir + ' axis)')] is not None:
                    ws.cell(row = 4+i, column = 9).value = 'Strain'
                    ws.cell(row = 4+i, column = 10).value = strain[idx_c]

                # ---- Stress Target Value
                elif test_summary.value[i][test_summary.columns.index('Target Stress (' + test_dir + ' axis)')] is not None:
                    ws.cell(row = 4+i, column = 9).value = 'Stress'
                    ws.cell(row = 4+i, column = 10).value = stress[idx_c]

                # ---- Target Not Defined - default to control mode
                else:
                    if ws.cell(row = 4+i, column = 7).value == 'Strain':
                        ws.cell(row = 4+i, column = 9).value = 'Strain'
                        ws.cell(row = 4+i, column = 10).value = strain[time.index(ws.cell(row = 4+i, column = 5).value)]
                    if ws.cell(row = 4+i, column = 7).value == 'Stress':
                        ws.cell(row = 4+i, column = 9).value = 'Stress'
                        ws.cell(row = 4+i, column = 10).value = stress[time.index(ws.cell(row = 4+i, column = 5).value)]
        # Save the file
        fname = os.path.join(temp_path, record.attributes['Specimen ID'].value + "_PyCOMP_IN.xlsx")
        wb.save(fname)
    else:
        fname = None

    return fname, msg
