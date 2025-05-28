#--------------------------------------------------------------------------------------------------------------------------------------------------------------------
#
# CreateModelTab.py
#
# PURPOSE: Create the Optimize Model tab. The Optimize Model tab allows users to define a model to fit and run COMPARE
#          to determine the optimal parameter values
#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------
def CreateModelTab(self,window,frmt):
    # Import Modules
    import copy
    import json
    from openpyxl import load_workbook
    import tkinter as tk
    from tkinter import messagebox
    from tkinter import simpledialog
    from tkinter import ttk 
    from tkinter import scrolledtext 
    import tksheet
    
    # Import Functions
    from General.DeleteWidgets import DeleteTab
    from Model.UpdateModelData import UpdateModelData

    #Unpack Formatting
    bg_color = frmt[0] 
    fontname = frmt[1]
    fsize_s = frmt[2]

    # Initialize Model
    if 'Model' not in self.Compare.keys():
        self.Compare['Model'] = {}
        self.Compare['Model']['Status'] = 0

    if 'Model ID' not in self.Compare.keys() == False:
        self.Compare['Model ID'] = None

    # Get available model information
    model_info = load_workbook(self.Compare['Paths']['Model Library'], data_only=True)

    # Delete all tab attributes
    if hasattr(self,"tab_att_list"):
        if hasattr(self,'clicked') == False:
            UpdateModelData(None, self, 3, 'Model')
        else:
            if self.clicked == 1:
                self.clicked = 0
            else:
                UpdateModelData(None, self, 3, 'Model')
        DeleteTab(self)

    if hasattr(self, 'canvas'):
        self.toolbar.destroy()
        self.canvas.get_tk_widget().destroy()
        del self.canvas

    # Preallocate the att list and other variables
    self.att_list = []
    self.loc_att_list = []
    self.tab_att_list = []
    self.optimize = 0
    self.clicked = 0

    # Preallocate Saved Models
    if 'Model Library' not in self.Compare.keys():
        self.Compare['Model Library'] = {}

    # Define Available Models
    self.Models = model_info.sheetnames

    def change_model(value):
        #----------------------------------------------------------------------
        #
        #   PURPOSE: Recreate page based on model choice.
        #
        #----------------------------------------------------------------------

        # Clear the model information if the model type changed
        try:
            if value != self.Compare['Model']['Model Name']:
                # Prompt user to save model
                self.save_model(None)

                # Clear model information
                self.Compare['Model'] = {}
                self.Compare['Model']['Status'] = 0
        except:
            pass
        
        # Delete local attributes
        DeleteTab(self)

        # Clear sheets
        if hasattr(self,'sheet1'):
            self.sheet1.destroy()
            del self.sheet1
        if hasattr(self,'sheet2'):
            self.sheet2.destroy()
            del self.sheet2

        # Read the model info
        ws = model_info[value]
        self.Compare['Model']['Model Name'] = value
        self.Compare['Model']['Model Info'] = {}
        for i in range(1,ws.max_row+1):
            self.Compare['Model']['Model Info'][ws.cell(row=i,column=1).value] = []
            j = 2
            if "Units" not in ws.cell(row=i,column=1).value:
                while ws.cell(row=i,column=j).value != None:
                    self.Compare['Model']['Model Info'][ws.cell(row=i,column=1).value].append(ws.cell(row=i,column=j).value)
                    j= j+1
            else:
                for j in range(2,len(self.Compare['Model']['Model Info'][ws.cell(row=i-1,column=1).value])+2):
                    self.Compare['Model']['Model Info'][ws.cell(row=i,column=1).value].append(ws.cell(row=i,column=j).value)
            
        # Get available reversible models
        self.RevModels = self.Compare['Model']['Model Info']['Reversible Models']
        
        if len(self.RevModels) > 0:
            # Create the label
            text_var = tk.StringVar()
            text_var.set("Reversible Model:")
            self.desc2 = tk.Label(window, 
                            textvariable=text_var, 
                            anchor=tk.CENTER,       
                            bg=bg_color,                  
                            font=(fontname, fsize_s),                    
                            )
            self.desc2.place(anchor = 'w', relx = 0.225, rely = 0.205)
            self.tab_att_list.append('self.desc2')

            # Create the option variable
            self.opt2 = tk.StringVar(window)
            rmod_opt = self.RevModels[0]

            # Check if previous data exists
            if 'Model' in list(self.Compare.keys()):
                # Set the reversible model type
                if 'Reversible Model Name' in list(self.Compare['Model'].keys()):
                    if self.Compare['Model']['Reversible Model Name'] in self.RevModels:
                        rmod_opt = self.Compare['Model']['Reversible Model Name']

            # Create the reversible model drop down   
            self.opt2.set(rmod_opt) 
            self.optmenu2 = tk.OptionMenu(window, self.opt2, *self.RevModels, command = lambda event:UpdateModelData(event, self, 1, 'Model')) 
            self.optmenu2.place(anchor = 'w', relx = 0.35, rely = 0.205)
            self.optmenu2.configure(font = fsize_s)
            self.tab_att_list.append('self.optmenu2')

            # Initialize Parameter List
            self.Params_VE = self.Compare['Model']['Model Info']['Reversible Deformation Parameters'] + self.Compare['Model']['Model Info']['Reversible Damage Parameters']
            self.Params_VE_Units = self.Compare['Model']['Model Info']['Reversible Deformation Parameter Units'] + self.Compare['Model']['Model Info']['Reversible Damage Parameter Units']

        # Get available irreversible models
        self.IrrevModels = self.Compare['Model']['Model Info']['Irreversible Models']

        if len(self.IrrevModels) > 0:
            # Create the label
            text_var = tk.StringVar()
            text_var.set("Irreversible Model:")
            self.desc3 = tk.Label(window, 
                            textvariable=text_var, 
                            anchor=tk.CENTER,       
                            bg=bg_color,                  
                            font=(fontname, fsize_s),                    
                            )
            self.desc3.place(anchor = 'w', relx = 0.525, rely = 0.205)
            self.tab_att_list.append('self.desc3')

            # Create the option variable
            self.opt3 = tk.StringVar(window)
            irmod_opt = self.IrrevModels[0]

            # Check if previous data exists
            if 'Model' in list(self.Compare.keys()):
                # Set the reversible model type
                if 'Irreversible Model Name' in list(self.Compare['Model'].keys()):
                    if self.Compare['Model']['Irreversible Model Name'] in self.IrrevModels:
                        irmod_opt = self.Compare['Model']['Irreversible Model Name']

            # Create the irreversible model drop down
            self.opt3.set(irmod_opt) 
            self.optmenu3 = tk.OptionMenu(window, self.opt3, *self.IrrevModels, command = lambda event:UpdateModelData(event, self, 1, 'Model')) 
            self.optmenu3.place(anchor = 'w', relx = 0.65, rely = 0.205)
            self.optmenu3.configure(font = fsize_s)
            self.tab_att_list.append('self.optmenu3')

            # Initialize Parameter List
            self.Params_VP = self.Compare['Model']['Model Info']['Irreversible Deformation Parameters'] + self.Compare['Model']['Model Info']['Irreversible Damage Parameters']
            self.Params_VP_Units = self.Compare['Model']['Model Info']['Irreversible Deformation Parameter Units'] + self.Compare['Model']['Model Info']['Irreversible Damage Parameter Units']

        # Update Reversible Parameters
        def update_reversible_table(self):
            #------------------------------------------------------------------
            #
            #   PURPOSE: Update the reversible model parameters table.
            #
            #------------------------------------------------------------------

            # Delete table if it exists
            if hasattr(self,"sheet1"):
                if hasattr(self,"res_flag1") == True:
                    if self.res_flag1 == 0:
                        # Store data
                        self.sheet1_data = self.sheet1.data
                else:
                    # Store data
                    self.sheet1_data = self.sheet1.data

                # Delete sheet
                self.sheet1.destroy()
                del self.sheet1

            # Check if previous data exists
            if 'Model' in list(self.Compare.keys()):
                # Set the reversible model type
                if 'VE_Param' in list(self.Compare['Model'].keys()):
                    if hasattr(self,"res_flag1") == True:
                        if self.res_flag1 == 0:
                            self.sheet1_data = self.Compare['Model']['VE_Param']
                        else:
                            self.res_flag1 = 0
                    else:
                        self.sheet1_data = self.Compare['Model']['VE_Param']

            # Set the columns
            Cols = ['Parameter', 'Units','Lower Bound','Initial Guess','Upper Bound','Active/Passive', 'COMPARE']

            # Create the table
            self.sheet1 = tksheet.Sheet(window, total_rows = len(self.Params_VE), total_columns = len(Cols), 
                            headers = Cols,
                            width = 700, height = 400, show_x_scrollbar = False, show_y_scrollbar = True,
                            font = (fontname,12,"normal"),
                            header_font = (fontname,12,"bold"))
            self.sheet1.place(anchor = 'n', relx = 0.245, rely = 0.33)
            self.tab_att_list.append('self.sheet1')
            self.sheet1.change_theme("blue")
            self.sheet1.set_index_width(0)

            def sort_cols(self):
                #----------------------------------------------------------------------
                #
                #   PURPOSE: Custom soring function.
                #
                #----------------------------------------------------------------------

                # Get the currently selected element
                currently_selected = self.sheet1.get_currently_selected()
                
                # Get the list of values
                sort_list = []
                for i in range(self.sheet1.visible_rows[1]):
                    sort_list.append(self.sheet1.data[i][currently_selected.column])
                index_list = sorted(range(len(sort_list)), key=lambda k: sort_list[k])
                
                # Rewrite the table
                temp_data = copy.deepcopy(self.sheet1.data)
                for i in range(self.sheet1.visible_rows[1]):
                    for j in range(self.sheet1.visible_columns[1]):
                        self.sheet1.set_cell_data(i,j,temp_data[index_list[i]][j])
                self.sheet1.redraw()

            def all_active(self):
                #----------------------------------------------------------------------
                #
                #   PURPOSE: Set all parameters to active.
                #
                #----------------------------------------------------------------------

                # Set all rows to Active
                for i in range(len(self.sheet1.data)):
                    self.sheet1.set_cell_data(i,5, 'Active')

            def all_passive(self):
                #----------------------------------------------------------------------
                #
                #   PURPOSE: Set all parameters to passive.
                #
                #----------------------------------------------------------------------

                # Set all rows to Passive
                for i in range(len(self.sheet1.data)):
                    self.sheet1.set_cell_data(i,5, 'Passive')

            def genbounds(self):
                #----------------------------------------------------------------------
                #
                #   PURPOSE: Automatically generate bounds.
                #
                #----------------------------------------------------------------------

                # Get the bounds slider value
                value = float(self.slider1.get())

                # Generate bounds
                for i in range(len(self.sheet1.data)):
                    try:
                        val = float(self.sheet1.data[i][3])
                        lb = val-val*float(value)/100
                        self.sheet1.set_cell_data(i,2, '{:0.4e}'.format(lb))
                        ub = val+val*float(value)/100
                        self.sheet1.set_cell_data(i,4, '{:0.4e}'.format(ub))
                        self.sheet1.redraw() 
                    except:
                        pass

            # Enable bindings
            self.sheet1.enable_bindings('single_select','cell_select', 'column_select', 'edit_cell',"arrowkeys", "right_click_popup_menu")
            self.sheet1.popup_menu_add_command('Sort', lambda : sort_cols(self), table_menu = True, index_menu = True, header_menu = True)
            self.sheet1.popup_menu_add_command('Change All to Active', lambda : all_active(self), table_menu = True, index_menu = True, header_menu = True)
            self.sheet1.popup_menu_add_command('Change All to Passive', lambda : all_passive(self), table_menu = True, index_menu = True, header_menu = True)
            self.sheet1.popup_menu_add_command('Auto-Generate Bounds', lambda : genbounds(self), table_menu = True, index_menu = True, header_menu = True)
            self.sheet1.extra_bindings([("cell_select", lambda event: self.cell_select_opt(event, 'sheet1'))])

            # Set Column Widths
            self.sheet1.column_width(column = 0, width = 85, redraw = True)
            self.sheet1.column_width(column = 1, width = 60, redraw = True)
            self.sheet1.column_width(column = 2, width = 100, redraw = True)
            self.sheet1.column_width(column = 3, width = 100, redraw = True)
            self.sheet1.column_width(column = 4, width = 100, redraw = True)
            self.sheet1.column_width(column = 5, width = 110, redraw = True)
            self.sheet1.column_width(column = 6, width = 95, redraw = True)
            self.sheet1.table_align(align = 'c',redraw=True)

            # Set unit dictionary
            Units = {'Stress':['GPa','MPa','kPa','Pa','msi','ksi','psi'],
                    'Time':['s'],
                    'Time-1':['1/s']
                    }

            # Set Rows
            for i in range(len(self.Params_VE)):
                self.sheet1.set_cell_data(i,0, self.Params_VE[i])
                self.sheet1.create_dropdown(r=i, c = 5,values=['Active','Passive'])
                if self.Params_VE_Units[i] != None:
                    for key in list(Units.keys()):
                        if self.Params_VE_Units[i] in Units[key]:
                            units_list = Units[key]
                else:
                    units_list = []
                def_val = self.Params_VE_Units[i]
                    
                self.sheet1.create_dropdown(r=i, c = 1,values=units_list)
                if def_val != None:
                    self.sheet1.set_cell_data(i,1, def_val)

            # Set Optimization Flag
            try:
                if self.sheet1_data[0][6] != '' and self.sheet1_data[0][6] != None:
                    self.optimize = 1
                    self.Compare['Model']['Status'] = 1
            except:
                pass

            # Add Existing Data
            for i in range(len(self.sheet1_data)):
                try:
                    # Find the corresponding index
                    rown = None
                    for j in range(len(self.sheet1.data)):
                        if self.sheet1.data[j][0] == self.sheet1_data[i][0]:
                            rown = j

                    if rown != None:
                        for j in range(1,len(Cols)):
                            try:
                                self.sheet1.set_cell_data(rown,j, self.sheet1_data[i][j])
                                if j == 6:
                                    if self.optimize == 1:
                                        try:
                                            if float(self.sheet1_data[i][6]) > float(self.sheet1_data[i][2]) and float(self.sheet1_data[i][6]) < float(self.sheet1_data[i][4]):
                                                self.sheet1.highlight((rown,6),fg = 'green', bg = 'white')
                                            else:
                                                self.sheet1.highlight((rown,6),fg = 'red', bg = 'white')
                                        except:
                                            pass
                                    else:
                                        self.sheet1.set_cell_data(rown,j, '')
                            except:
                                pass
                except:
                    pass

            # Redraw the table
            self.sheet1.redraw()

            # Update the Model Data
            UpdateModelData(None, self, 1, 'Model')

        def VE_param(value):
            #------------------------------------------------------------------
            #
            #   PURPOSE: Get List of of ViscoElastic Mechanisms.
            #
            #------------------------------------------------------------------

            # Initialize Parameters
            self.Params_VE = []
            self.Params_VE_Units = []

            # Add non-mechanism dependent parameters
            for i in range(len(self.Compare['Model']['Model Info']['Reversible Deformation Parameters'])):
                param = self.Compare['Model']['Model Info']['Reversible Deformation Parameters'][i]
                unit = self.Compare['Model']['Model Info']['Reversible Deformation Parameter Units'][i]
                if '_[M]' not in param:
                    self.Params_VE.append(param)
                    self.Params_VE_Units.append(unit)
                else:
                    for i in range(int(value)):
                        param_mech = param.replace("_[M]",str(i+1))
                        self.Params_VE.append(param_mech)
                        self.Params_VE_Units.append(unit)

            # Update the reversible mechanisms table
            update_reversible_table(self)

        def update_irreversible_table(self):
            #------------------------------------------------------------------
            #
            #   PURPOSE: Update the irreversible model parameters table.
            #
            #------------------------------------------------------------------
                        
            # Delete table if it exists
            if hasattr(self,"sheet2"):
                if hasattr(self,"res_flag2") == True:
                    if self.res_flag2 == 0:
                        # Store data
                        self.sheet2_data = self.sheet2.data
                else:
                    # Store data
                    self.sheet2_data = self.sheet2.data

                # Delete sheet
                self.sheet2.destroy()
                del self.sheet2

            # Check if previous data exists
            if 'Model' in list(self.Compare.keys()):
                # Set the reversible model type
                if 'VP_Param' in list(self.Compare['Model'].keys()):
                    if hasattr(self,"res_flag2") == True:
                        if self.res_flag2 == 0:
                            self.sheet2_data = self.Compare['Model']['VP_Param']
                        else:
                            self.res_flag2 = 0
                    else:
                        self.sheet2_data = self.Compare['Model']['VP_Param']


            # Set the columns
            Cols = ['Parameter', 'Units','Lower Bound','Initial Guess','Upper Bound','Active/Passive','COMPARE']

            # Create the table
            self.sheet2 = tksheet.Sheet(window, total_rows = len(self.Params_VP), total_columns = len(Cols), 
                            headers = Cols,
                            width = 700, height = 400, show_x_scrollbar = False, show_y_scrollbar = True,
                            font = (fontname,12,"normal"),
                            header_font = (fontname,12,"bold"))
            self.sheet2.place(anchor = 'n', relx = 0.75, rely = 0.33)
            self.tab_att_list.append('self.sheet2')
            self.sheet2.change_theme("blue")
            self.sheet2.set_index_width(0)

            def sort_cols(self):
                #----------------------------------------------------------------------
                #
                #   PURPOSE: Custom soring function.
                #
                #----------------------------------------------------------------------

                # Get currently selected element
                currently_selected = self.sheet2.get_currently_selected()
                
                # Get the list of values
                sort_list = []
                for i in range(self.sheet2.visible_rows[1]):
                    sort_list.append(self.sheet2.data[i][currently_selected.column])
                index_list = sorted(range(len(sort_list)), key=lambda k: sort_list[k])
                
                # Rewrite the table
                temp_data = copy.deepcopy(self.sheet2.data)
                for i in range(self.sheet2.visible_rows[1]):
                    for j in range(self.sheet2.visible_columns[1]):
                        self.sheet2.set_cell_data(i,j,temp_data[index_list[i]][j])
                self.sheet2.redraw()
        
            def all_active(self):
                #----------------------------------------------------------------------
                #
                #   PURPOSE: Set all parameters to active.
                #
                #----------------------------------------------------------------------

                # Set all parameters to Active
                for i in range(len(self.sheet2.data)):
                    self.sheet2.set_cell_data(i,5, 'Active')

            def all_passive(self):
                #----------------------------------------------------------------------
                #
                #   PURPOSE: Set all parameters to passive.
                #
                #----------------------------------------------------------------------

                # Set all parameters to Passive
                for i in range(len(self.sheet2.data)):
                    self.sheet2.set_cell_data(i,5, 'Passive')
                
            # Auto generate bounds
            def genbounds(self):
                #----------------------------------------------------------------------
                #
                #   PURPOSE: Automatically generate bounds.
                #
                #----------------------------------------------------------------------

                # Get the bounds slider value
                value = float(self.slider1.get())
                for i in range(len(self.sheet2.data)):
                    try:
                        val = float(self.sheet2.data[i][3])
                        lb = val-val*float(value)/100
                        self.sheet2.set_cell_data(i,2, '{:0.4e}'.format(lb))
                        ub = val+val*float(value)/100
                        self.sheet2.set_cell_data(i,4, '{:0.4e}'.format(ub))
                        self.sheet2.redraw() 
                    except:
                        pass

            # Enable Bindings
            self.sheet2.enable_bindings('single_select','cell_select', 'column_select', 'edit_cell',"arrowkeys", "right_click_popup_menu")
            self.sheet2.popup_menu_add_command('Sort', lambda : sort_cols(self), table_menu = True, index_menu = True, header_menu = True)
            self.sheet2.popup_menu_add_command('Change All to Active', lambda : all_active(self), table_menu = True, index_menu = True, header_menu = True)
            self.sheet2.popup_menu_add_command('Change All to Passive', lambda : all_passive(self), table_menu = True, index_menu = True, header_menu = True)
            self.sheet2.popup_menu_add_command('Auto-Generate Bounds', lambda : genbounds(self), table_menu = True, index_menu = True, header_menu = True)
            self.sheet2.extra_bindings([("cell_select", lambda event: self.cell_select_opt(event, 'sheet2'))])

            # Set Column Widths
            self.sheet2.column_width(column = 0, width = 85, redraw = True)
            self.sheet2.column_width(column = 1, width = 65, redraw = True)
            self.sheet2.column_width(column = 2, width = 100, redraw = True)
            self.sheet2.column_width(column = 3, width = 100, redraw = True)
            self.sheet2.column_width(column = 4, width = 100, redraw = True)
            self.sheet2.column_width(column = 5, width = 110, redraw = True)
            self.sheet2.column_width(column = 6, width = 95, redraw = True)
            self.sheet2.table_align(align = 'c',redraw=True)

            # Set unit dictionary
            Units = {'Stress':['GPa','MPa','kPa','Pa','msi','ksi','psi'],
                     'Stress-Time':['GPa-s','MPa-s','kPa-s','Pa-s','msi-s','ksi-s','psi-s'],
                    'Time':['s'],
                    'Time-1':['1/s'],
                    }

            # Set Rows
            for i in range(len(self.Params_VP)):
                self.sheet2.set_cell_data(i,0, self.Params_VP[i])
                self.sheet2.create_dropdown(r=i, c = 5,values=['Active','Passive'])
                if self.Params_VP_Units[i] != None:
                    for key in list(Units.keys()):
                        if self.Params_VP_Units[i] in Units[key]:
                            units_list = Units[key]
                else:
                    units_list = []
                def_val = self.Params_VP_Units[i]
                    
                self.sheet2.create_dropdown(r=i, c = 1,values=units_list)
                if def_val != None:
                    self.sheet2.set_cell_data(i,1, def_val)

            # Set Optimization Flag
            try:
                if self.sheet2_data[0][6] != '' and self.sheet2_data[0][6] != None:
                    self.optimize = 1
                    self.Compare['Model']['Status'] = 1
            except:
                pass

            # Add Existing Data
            for i in range(len(self.sheet2_data)):
                try:
                    # Find the corresponding index
                    rown = None
                    for j in range(len(self.sheet2.data)):
                        if self.sheet2.data[j][0] == self.sheet2_data[i][0]:
                            rown = j

                    if rown != None:
                        for j in range(1,len(Cols)):
                            try:
                                self.sheet2.set_cell_data(rown,j, self.sheet2_data[i][j])
                                if j == 6:
                                    if self.optimize == 1:
                                        try:
                                            if float(self.sheet2_data[i][6]) > float(self.sheet2_data[i][2]) and float(self.sheet2_data[i][6]) < float(self.sheet2_data[i][4]):
                                                self.sheet2.highlight((rown,6),fg = 'green', bg = 'white')
                                            else:
                                                self.sheet2.highlight((rown,6),fg = 'red', bg = 'white')
                                        except:
                                            pass
                                    else:
                                        self.sheet2.set_cell_data(rown,j, '')
                            except:
                                pass
                except:
                    pass

            #Redraw the table
            self.sheet2.redraw()

            # Update the Model Data
            UpdateModelData(None, self, 2, 'Model')

        # Get Number of ViscoPlastic Mechanisms
        def VP_param(value):
            #------------------------------------------------------------------
            #
            #   PURPOSE: Get List of of ViscoPlastic Mechanisms.
            #
            #------------------------------------------------------------------

            # Initialize Parameters
            self.Params_VP = []
            self.Params_VP_Units = []

            # Add non-mechanism dependent parameters
            for i in range(len(self.Compare['Model']['Model Info']['Irreversible Deformation Parameters'])):
                param = self.Compare['Model']['Model Info']['Irreversible Deformation Parameters'][i]
                unit = self.Compare['Model']['Model Info']['Irreversible Deformation Parameter Units'][i]
                if '_[N]' not in param:
                    self.Params_VP.append(param)
                    self.Params_VP_Units.append(unit)
                else:
                    for j in range(int(value)):
                        param_mech = param.replace("_[N]",str(j+1))
                        self.Params_VP.append(param_mech)
                        self.Params_VP_Units.append(unit)

            # Update the reversible mechanisms table
            update_irreversible_table(self)


        # Get number of viscoelastic parameters
        self.VEMech = self.Compare['Model']['Model Info']['Reversible Mechanisms']
        if len(self.VEMech) > 0:
            # Create the label
            self.desc4 = tk.Label(window, 
                            text="Viscoelastic Mechanisms (M):", 
                            anchor=tk.CENTER,       
                            bg=bg_color,                  
                            font=(fontname, fsize_s),                    
                            )
            self.desc4.place(anchor = 'w', relx = 0.225, rely = 0.28)
            self.tab_att_list.append('self.desc4')

            self.opt4 = tk.StringVar(window)
            ve_opt = self.VEMech[0]

            # Check if previous data exists
            if 'Model' in list(self.Compare.keys()):
                # Set the reversible model type
                if 'M' in list(self.Compare['Model'].keys()):
                    if int(self.Compare['Model']['M']) in self.VEMech:
                        ve_opt = int(self.Compare['Model']['M'])

            # Create the drop down
            self.opt4.set(ve_opt) 
            self.optmenu4 = tk.OptionMenu(window, self.opt4, *self.VEMech, command = VE_param) 
            self.optmenu4.place(anchor = 'w', relx = 0.4, rely = 0.28)
            self.optmenu4.configure(font = fsize_s)
            self.tab_att_list.append('self.optmenu4')

            # Get list of viscoelastic parameters
            VE_param(ve_opt)

        # Get Number of Viscoplastic Mechanisms
        self.VPMech = self.Compare['Model']['Model Info']['Irreversible Mechanisms']
        if len(self.VPMech) > 0:
            # Create the label
            self.desc5 = tk.Label(window, 
                            text="Viscoplastic Mechanisms (N):", 
                            anchor=tk.CENTER,       
                            bg=bg_color,                  
                            font=(fontname, fsize_s),                    
                            )
            self.desc5.place(anchor = 'w', relx = 0.525, rely = 0.28)
            self.tab_att_list.append('self.desc5')

            self.opt5 = tk.StringVar(window)
            vp_opt = self.VPMech[0]

            # Check if previous data exists
            if 'Model' in list(self.Compare.keys()):
                # Set the reversible model type
                if 'N' in list(self.Compare['Model'].keys()):
                    if int(self.Compare['Model']['N']) in self.VPMech:
                        vp_opt = int(self.Compare['Model']['N'])

            # Create the drop down menu
            self.opt5.set(vp_opt) 
            self.optmenu5 = tk.OptionMenu(window, self.opt5, *self.VPMech, command = VP_param) 
            self.optmenu5.place(anchor = 'w', relx = 0.7, rely = 0.28)
            self.optmenu5.configure(font = fsize_s)
            self.tab_att_list.append('self.optmenu5')

            # Get list of viscoplastic parameters
            VP_param(vp_opt)

        # Create the bounds slider
        # -- Format slider text
        def update_value(value):
            formatted_value = f"Bounds: ± {int(value)}%"  # Format to integer decimal places with %
            self.desc6.config(text=formatted_value)

        # -- Create the label
        self.desc6 = tk.Label(window, 
                        text="Bounds: ± 5%", 
                        anchor=tk.CENTER,       
                        bg=bg_color,                  
                        font=(fontname, fsize_s),                    
                        )
        self.desc6.place(anchor = 'w', relx = 0.825, rely = 0.205)
        self.tab_att_list.append('self.desc6')

        # -- Create the slider
        self.slider1 = tk.Scale(window, from_=5, to=50, orient=tk.HORIZONTAL, length=200, command = update_value, showvalue=False)
        self.slider1.place(anchor = 'w', relx = 0.8, rely = 0.28)
        self.tab_att_list.append('self.slider1')

        # Local Formatting
        starty = 0.895

        # Create the Load from Excel button
        self.btn_load_db = tk.Button(window, text = "Load from Excel", command = lambda:self.load_from_db('Optimize'), 
                                    font = (fontname, fsize_s), bg = '#fc3d21', fg='white',
                                    width = 18)
        self.btn_load_db.place(anchor = 'w', relx = self.startx+ self.delx*0, rely = starty)
        self.tab_att_list.append('self.btn_load_db')

        # Create button to view/delete models
        self.btn_modlib = tk.Button(window, text = "Model Library", command = lambda : self.Model_Library('Optimize'), 
                                    font = (fontname, fsize_s), bg = '#fc3d21', fg='white',
                                    width = 18)
        self.btn_modlib.place(anchor = 'w', relx = self.startx+ self.delx*1, rely = starty)
        self.tab_att_list.append('self.btn_modlib')

        # Create the Optimize button
        self.btn_opt = tk.Button(window, text = "Optimize", command = self.optimizer, 
                                    font = (fontname, fsize_s), bg = '#fc3d21', fg='white',
                                    width = 18)
        self.btn_opt.place(anchor = 'w', relx = self.startx+ self.delx*2, rely = starty)
        self.tab_att_list.append('self.btn_opt')

        
        def save_model_local():
            #------------------------------------------------------------------
            #
            #   PURPOSE: Save models to the project library.
            #
            #------------------------------------------------------------------

            # Save notes if they exist
            nflag = 0
            if 'Note' in self.Compare['Model'].keys():
                note =  self.Compare['Model']['Note']
                nflag = 1

            # Reset the Model
            self.Compare['Model'] = {}

            # Update Model Data
            UpdateModelData(None, self, 3, 'Model')

            # Set the model type
            self.Compare['Model']['Compare Type'] = 'Optimize'

            # Add the note back
            if nflag == 1:
                self.Compare['Model']['Note'] = note

            # Get the save name
            save_flag = 0
            while save_flag == 0:
                user_input = simpledialog.askstring("Save Model", "Enter the model name:")

                if user_input in list(self.Compare['Model Library'].keys()):
                    askyn = messagebox.askyesno(title = 'Save Model', message = 'Do you want to overwrite ' + user_input + ' ?')
                    if askyn == True:
                        save_flag = 1
                else:
                    save_flag = 1
            
            # Save to binary in the model library
            json_string = json.dumps(self.Compare['Model'])
            binary_data = json_string.encode('utf-8')
            self.Compare['Model Library'][user_input] = binary_data

            # Set the model name
            self.Compare['Model ID']= user_input

        # Create button to save a model
        self.btn_savemod = tk.Button(window, text = "Save Model", command = save_model_local, 
                                    font = (fontname, fsize_s), bg = '#fc3d21', fg='white',
                                    width = 18)
        self.btn_savemod.place(anchor = 'w', relx = self.startx+ self.delx*3, rely = starty)
        self.tab_att_list.append('self.btn_savemod')

        def add_note():
            #------------------------------------------------------------------
            #
            #   PURPOSE: Add a note to the model.
            #
            #------------------------------------------------------------------

            # Set the flag for the window
            if hasattr(self,'note_click') == False:
                self.note_click = 0

            # Open the window if it does not exist
            if self.note_click == 0:
                self.note_click = 1

                # Create the window
                root = tk.Tk() 
                root.geometry("600x400")
                root.title("Enter Model Notes") 
                
                # Create the label
                ttk.Label(root, text="Enter Model Notes:", 
                            font=(fontname, fsize_s)).place(anchor='n', relx = 0.5, rely = 0.1) 
                
                # Create the note area
                text_area = scrolledtext.ScrolledText(root, wrap=tk.WORD, 
                                                    width=40, height=8, 
                                                    font=(fontname, fsize_s)) 
                text_area.place(anchor='c', relx = 0.5, rely = 0.5)

                # Display any existing notes
                if 'Note' in list(self.Compare['Model'].keys()):
                    text_area.insert("end", self.Compare['Model']['Note']) 
                
                # placing cursor in text area 
                text_area.focus()

                def on_closing_root(self):
                    #--------------------------------------------------------------
                    #
                    #   PURPOSE: Create exit protocol for the note window.
                    #
                    #--------------------------------------------------------------
                    
                    # Save the note
                    try:
                        self.Compare['Model']['Note'] = text_area.get("1.0",'end-1c')
                    except:
                        pass

                    # Reset the window
                    self.note_click = 0
                    root.destroy()

                # Add the exit protocol to the root
                root.protocol("WM_DELETE_WINDOW", lambda:on_closing_root(self))

        # Create button to add a note
        self.btn_addnote = tk.Button(window, text = "Model Notes", command = add_note, 
                                    font = (fontname, fsize_s), bg = '#fc3d21', fg='white',
                                    width = 12)
        self.btn_addnote.place(anchor = 'w', relx = self.startx+ self.delx*0.25, rely = 0.275)
        self.tab_att_list.append('self.btn_addnote')

        # Update Model Data
        UpdateModelData(None, self, 3, 'Model')

        # Update Tables
        if len(self.Compare['Model']['Model Info']['Reversible Models']) > 0:
            update_reversible_table(self)
        if len(self.Compare['Model']['Model Info']['Irreversible Models']) > 0:
            update_irreversible_table(self)

    # Create the label for Model Type
    self.desc1 = tk.Label(window, 
                     text="Select the Model:", 
                     anchor=tk.CENTER,       
                     bg=bg_color,                  
                     font=(fontname, fsize_s),                    
                    )
    self.desc1.place(anchor = 'w', relx = 0.025, rely = 0.205)
    self.loc_att_list.append('self.desc1')

    # Create the model text variable
    self.opt1 = tk.StringVar(window)
    mod_opt = self.Models[0]

    # Check if previous value exists
    if 'Model' in list(self.Compare.keys()):
        # Set the model name
        if 'Model Name' in list(self.Compare['Model'].keys()):
            if self.Compare['Model']['Model Name'] in self.Models:
                mod_opt = self.Compare['Model']['Model Name']

    # Create Option Menu for Model Type
    self.opt1.set(mod_opt) 
    self.optmenu1 = tk.OptionMenu(window, self.opt1, *self.Models, command= change_model) 
    self.optmenu1.place(anchor = 'w', relx = 0.15, rely = 0.205)
    self.optmenu1.configure(font = fsize_s)
    change_model(mod_opt)
    self.loc_att_list.append('self.optmenu1')