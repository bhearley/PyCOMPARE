#--------------------------------------------------------------------------------------------------------------------------------------------------------------------
#
# PyCOMPARE.py
#   Brandon Hearley - LMS
#   brandon.l.hearley@nasa.gov
#   4/15/2025
#
# PURPOSE: Run the PyCOMPARE GUI
#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Import Modules
import copy
import json
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import (FigureCanvasTkAgg, NavigationToolbar2Tk)
from matplotlib.ticker import ScalarFormatter
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side
import os
import pandas as pd
from pathlib import Path
import pickle
from PIL import ImageTk, Image
from scipy.interpolate import CubicSpline
import shutil
import tkinter as tk
from tkinter import filedialog 
from tkinter import messagebox
from tkinter import simpledialog
from tkinter import scrolledtext
from tkinter import ttk 
import tksheet

# Import Functions
from Data.CreateCharacterizationTab import *
from Data.CreateDataTab import *
from General.BuildGeneralPage import *
from General.BuildHomePage import *
from General.GetProjectFile import *
from General.DeleteWidgets import *
from Model.CreateAnalysisTab import*
from Model.CreateModelTab import *
from Model.GVIPS.WriteDSG_GVIPS_ANLY import *
from Model.GVIPS.WriteDSG_GVIPS_OPT import *
from Model.ReadModel import *
from Model.UpdateModelData import *
from Model.WriteNLP import *
from Model.WriteSIM import *
from UnitConversion.UnitConversion import *
from Visualization.CreateVisualizationTab import *

# Set Directories
home = os.getcwd()

# GUI Formatting
bg_color = 'white'          #background color
fontname = 'calibri'        #font name
fsize_s = 16                #small font fize
fsize_l = 18                #large font size
fsize_t = 24                #title font size
frmt = [bg_color, fontname,fsize_s,fsize_l,fsize_t]    # Pack Formatting into list

# Define Images
title_img = os.path.join(home,'GUI','TitleHeader.png') # Set the title image path
logo_img = os.path.join(home,'GUI','NasaLogo.png')     # Set the logo image path

#Create the GUI
class PY_COMPARE:
    #Initialize
    def __init__(self):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: Initialize the GUI.
        #
        #--------------------------------------------------------------------------

        # Set global variales
        global window, frmt

        #Create Background Window
        window = tk.Tk()
        window.title("PyCOMPARE")
        window.state('zoomed')
        window.configure(bg='white')

        #Add the Title
        img = Image.open(title_img)
        scale = 0.9
        img = img.resize((int(img.width*scale), int(img.height*scale)))
        self.img_hdr = ImageTk.PhotoImage(img)
        self.panel_hdr = tk.Label(window, image = self.img_hdr, bg = bg_color)
        self.panel_hdr.place(anchor = 'n', relx = 0.5, rely = 0.005)

        #Add the NASA Logo
        img = Image.open(logo_img)
        scale = 0.8
        img = img.resize((int(img.width*scale), int(img.height*scale)))
        self.img_nasa = ImageTk.PhotoImage(img)
        self.panel_nasa = tk.Label(window, image = self.img_nasa, bg = bg_color)
        self.panel_nasa.place(anchor = 'e', relx = 0.999, rely = 0.045)

        # Build the home page
        BuildHomePage(self, window, frmt)

        # Ask for save when closing
        def on_closing(self):
            #----------------------------------------------------------------------
            #
            #   PURPOSE: Set exit protocol for the GUI.
            #
            #----------------------------------------------------------------------

            # Check if project file exists
            if hasattr(self,'proj_file'):

                # Prompt usert o save
                if messagebox.askyesno("Quit", "Do you want to save before exiting?"):

                    # Get the project file
                    file = self.proj_file

                    # Update the model info
                    UpdateModelData(None, self, 3, 'Model')

                    # Get the Data
                    data = self.Compare

                    # Write data to file
                    with open(file, 'wb') as file:
                        pickle.dump(data, file)

                    # Display save message to user
                    messagebox.showinfo(title = 'Save', message = 'Project Saved!')

            # Destory the window
            window.destroy()

        # Create Window Exit Protocol
        window.protocol("WM_DELETE_WINDOW", lambda:on_closing(self))

        #Main Loop
        window.mainloop()

    #------------------------------------------------------------------------------------------------------------------------------------------
    #
    #   START PAGE
    #   The start page allows users to define a new project file or load a previous project file
    #
    #------------------------------------------------------------------------------------------------------------------------------------------

    def new_project(self):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: Initialize a new project and load the general page.
        #
        #--------------------------------------------------------------------------

        # Create new project file
        CreateNewProject(self)

        # Initialize the data structure and build the home page
        if self.proj_file != None:

            # Initialize the data structure
            self.Compare = {}

            # Initialize the Settings Path Dependencies
            model_library = os.path.join(home,'Model','AvailableModels.xlsx')
            import_template = os.path.join(home,'Templates','ImportTemplate.xlsx')
            export_template = os.path.join(home,'Templates','ExportTemplate.xlsx')
            compare_path = os.path.join(home,'compnasardamage.exe')
            self.Compare['Paths'] = {'Model Library':model_library,
                                     'Import Template':import_template,
                                     'Export Template':export_template,
                                     'Compare Executable':compare_path,}
            
            # Delete the Home Page
            DeletePages(self)

            # Build the General Page
            BuildGeneralPage(self, window, frmt)

    def load_project(self):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: Load an existing proejct file and the general page.
        #
        #--------------------------------------------------------------------------
        
        # Load existing project file
        LoadProject(self)

        # Initialize the data structure and build the home page
        if self.proj_file != None:
            # Delete the Home Page
            DeletePages(self)

            # Build the General Page
            BuildGeneralPage(self, window, frmt)

    def save(self):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: Save the project file.
        #
        #--------------------------------------------------------------------------

        # Get the file name
        file = self.proj_file

        # Update the model info
        UpdateModelData(None, self, 3, 'Model')

        # Get the Compare Data
        data = self.Compare

        # Write data to file
        with open(file, 'wb') as file:
            pickle.dump(data, file)

        # Show save message to user
        messagebox.showinfo(title = 'Save', message = 'Project Saved!')

    #------------------------------------------------------------------------------------------------------------------------------------------
    #
    #   DATA PAGES
    #   Functions for the Database and Characterization Pages
    #   The Database page allows users to upload test data to their project database
    #   The Characterization page allows users to view data used for characterization and edit data reduction
    #
    #------------------------------------------------------------------------------------------------------------------------------------------

    def data_tab(self):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: Create the Database Page.
        #
        #--------------------------------------------------------------------------

        # Delete any local attributes
        DeleteLocal(self)

        # Create the database tab
        CreateDataTab(self,window,frmt)

    def char_tab(self):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: Create the Characterization Page.
        #
        #--------------------------------------------------------------------------

        # Delete any local attributes
        DeleteLocal(self)

        # Preallocate characterization
        if 'Characterization' not in list(self.Compare.keys()):
            self.Compare['Characterization'] = {}

        # Create the characterization tab
        CreateCharacterizationTab(self,window,frmt)

    def cell_select_char(self, response):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: Create Custom Column Disablement for Characterization Test Set
        #
        #--------------------------------------------------------------------------

        # Define locked columns (all but weight)
        locked_cols = [0,1,2,3,4,5,6]

        # Enable/Disable user ability to edit cells
        if response.selected.column != None:
            if response.selected.column in locked_cols:
                self.sheet_char.disable_bindings(("edit_cell"))
            else:
                self.sheet_char.enable_bindings(("edit_cell"))
                self.sheet_char.extra_bindings([("edit_cell", self.save_weight)])
            self.sheet_char.redraw()

    def save_weight(self, response):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: Save the test weights
        #
        #--------------------------------------------------------------------------
        
        # Get Characterization Table Data
        data = self.sheet_char.data

        # Get all weights
        weights = []
        for i in range(len(data)):
            try:
                weights.append(float(data[i][7]))
            except: 
                weights.append(0)

        # Write relative weights
        for i in range(len(data)):
            test_name = data[i][0]
            self.Compare['Characterization'][test_name]['RelWeight'] = weights[i]

    def cell_select_stage(self, response):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: Create Custom Column Disablement for Stage Table 
        #            (Data Reduction Window) 
        #
        #--------------------------------------------------------------------------

        # Get locked columns
        locked_cols = [0]

        # Enable/Disable user ability to edit cells
        if response.selected.column != None:
            if response.selected.column in locked_cols:
                self.stage_pts_sheet.disable_bindings(("edit_cell"))
            else:
                self.stage_pts_sheet.enable_bindings(("edit_cell"))
                self.stage_pts_sheet.extra_bindings([("edit_cell", self.edit_stage_pts)])
            self.stage_pts_sheet.redraw()

    def edit_stage_pts(self, response):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: Edit the Stage Table (Characterization Tab) with the new 
        #            reduced number of data points 
        #
        #--------------------------------------------------------------------------

        # Get the selected row and column
        c = response.selected.column
        r = response.selected.row

        # Set the number of division points
        try:
            int(self.stage_pts_sheet.data[r][c])
        except:
            self.stage_pts_sheet.set_cell_data(r,c,10)
            self.stage_pts_sheet.redraw()

    def plotter(self):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: Create plots of individual test response curves
        #
        #--------------------------------------------------------------------------

        #  Delete the canvas and drop down if it exists
        if hasattr(self, 'canvas'):
            self.toolbar.destroy()
            self.canvas.get_tk_widget().destroy()
            del self.canvas

        # Create the plot
        self.fig = Figure(figsize=(5,3.6), dpi = 125, constrained_layout = True)
        self.plot1 = self.fig.add_subplot(111)

        # Get the arrays
        x_val = self.opt1_plt.get()
        y_val = self.opt2_plt.get()

        # X Value
        if 'Time' in x_val:
            x = self.Compare['Data'][self.test_name]['Time']
            xs = self.Compare['Data'][self.test_name]['Reduced Data']['Time']
            xu = 'Time [s]'
        else:
            x_val = x_val.split('-')
            x = self.Compare['Data'][self.test_name][x_val[0]][int(x_val[1])]
            xs = self.Compare['Data'][self.test_name]['Reduced Data'][x_val[0]][int(x_val[1])]
            if x_val[0] == 'Strain':
                xu = 'Strain'
            else:
                xu = 'Stress [MPa]'

         # Y Value
        if 'Time' in y_val:
            y = self.Compare['Data'][self.test_name]['Time']
            ys = self.Compare['Data'][self.test_name]['Reduced Data']['Time']
            yu = 'Time [s]'
        else:
            y_val = y_val.split('-')
            y = self.Compare['Data'][self.test_name][y_val[0]][int(y_val[1])]
            ys = self.Compare['Data'][self.test_name]['Reduced Data'][y_val[0]][int(y_val[1])]
            if y_val[0] == 'Strain':
                yu = 'Strain'
            else:
                yu = 'Stress [MPa]'

        # Plot the data
        for i in range(len(self.Compare['Data'][self.test_name]['Stage Index'])):
            if i == 0:
                start_idx = 0
            else:
                start_idx = int(self.Compare['Data'][self.test_name]['Stage Index'][i-1])
            self.plot1.plot(x[start_idx:int(self.Compare['Data'][self.test_name]['Stage Index'][i])],
                            y[start_idx:int(self.Compare['Data'][self.test_name]['Stage Index'][i])],
                            label=self.Compare['Data'][self.test_name]['Stage Type'][i])

        # Plot Reduced Data Points on the Characterization Tab
        if hasattr(self,'sheet_char'):
            if self.sheet_char.winfo_exists():
                if xs is not None:
                    self.xdata = x
                    self.ydata = y
                    self.xsdata = xs
                    self.ysdata = ys
                    self.plot1.plot(xs,ys,'ro',label='Sample Points')

                    def AddPoint():
                        #----------------------------------------------------------
                        #
                        #   PURPOSE: Bind button click to adding a reduced data 
                        #            point.
                        #
                        #----------------------------------------------------------

                        # Bind the button click
                        if self.clicked == 0:
                            self.clicked = 1
                            self.plot1.figure.canvas.mpl_connect('button_press_event', self.add_point)

                    # Create Add Point Button
                    if hasattr(self, 'btn_loc4'):
                        self.btn_loc4.destroy()
                    self.btn_loc4 = tk.Button(window, text = "Add", command = AddPoint, 
                                font = (fontname, 10), bg = '#A9A9A9', fg='white',
                                width = 8)
                    self.btn_loc4.place(anchor = 'e', relx = 0.675, rely = 0.95)
                    self.tab_att_list.append('self.btn_loc4')

                    def DelPoint():
                        #----------------------------------------------------------
                        #
                        #   PURPOSE: Bind button click to deleting a reduced data 
                        #            point.
                        #
                        #----------------------------------------------------------

                        # Bind the button click
                        if self.clicked == 0:
                            self.clicked = 1
                            self.plot1.figure.canvas.mpl_connect('button_press_event', self.del_point)

                    # Create Delete Point Button
                    if hasattr(self, 'btn_loc5'):
                        self.btn_loc5.destroy()
                    self.btn_loc5 = tk.Button(window, text = "Delete", command = DelPoint, 
                                font = (fontname, 10), bg = '#A9A9A9', fg='white',
                                width = 8)
                    self.btn_loc5.place(anchor = 'e', relx = 0.725, rely = 0.95)
                    self.tab_att_list.append('self.btn_loc5')

        # Set Plot Formatting
        xlab = xu
        ylab = yu
        xlab_frmt = ScalarFormatter() 
        ylab_frmt = ScalarFormatter()

        # Format the plot
        self.plot1.set_xlabel(xlab)
        self.plot1.set_ylabel(ylab)
        self.plot1.xaxis.set_major_formatter(xlab_frmt)
        self.plot1.yaxis.set_major_formatter(ylab_frmt)
        if "Strain" in xlab or "Time" in xlab:
            self.plot1.ticklabel_format(style='sci',scilimits=(-6,-3),axis='x')
        if "Strain" in ylab or "Time" in ylab:
            self.plot1.ticklabel_format(style='sci',scilimits=(-6,-3),axis='y')
        self.plot1.legend()

        # Create the Tkinter canvas
        self.canvas = FigureCanvasTkAgg(self.fig, master = window)
        self.canvas.draw()

        # Create the Matplotlib toolbar
        self.toolbar = NavigationToolbar2Tk(self.canvas, window)
        self.toolbar.update()

        # Format Toolbar
        self.toolbar.config(bg=bg_color)
        self.toolbar._message_label.config(background=bg_color)
        self.toolbar.place(anchor = 'e', relx = 0.975, rely = 0.9)

        # Add the figure to the GUI
        self.canvas.get_tk_widget().place(anchor = 'n', relx = 0.78, rely = 0.32)
        if 'self.canvas' not in self.tab_att_list:
            self.tab_att_list.append('self.canvas')

        # Enable Data Reduction on the Characterization Table
        if hasattr(self,'sheet_char'):
            if self.sheet_char.winfo_exists():
                def ReduceData():
                    #--------------------------------------------------------------
                    #
                    #   PURPOSE: Reduce data based on stage type
                    #
                    #--------------------------------------------------------------
                
                    # Get the selected row and name
                    row = None
                    for i in range(len(self.sheet_char.data)):
                        if self.sheet_char.get_row_options(None)[i]['highlight'].bg == 'lightblue1':
                            row = i
                    try:
                        # Get the test information
                        self.test_name = self.sheet_char.data[row][0]
                        self.test_type = self.sheet_char.data[row][1]
                        self.load_dir = int(self.sheet_char.data[row][3])
                        rows = self.Compare['Data'][self.test_name]['Stage Type']

                        # Create the window to ask user for number of points for each stage
                        # -- Only create a window if it is not open
                        try:
                            if self.window == 1:
                                flag = 0
                            else:
                                flag = 1
                        except:
                            flag = 1
                    except:
                        flag = 0
                        messagebox.showinfo(message="No test selected.")

                    # Create the Segmentation Control Panel
                    if flag == 1:
                        self.window = 1
                        root = tk.Tk()
                        root.geometry("300x600")
                        root.configure(bg='white')
                        root.title("Segmentation Control Panel")

                        # Create a sheet with number of stages
                        Cols = ['Stage', 'Points']
                        self.stage_pts_sheet = tksheet.Sheet(root, total_rows = len(rows), total_columns = len(Cols), 
                            headers = Cols,
                            width = 220, height = 400, show_x_scrollbar = False, show_y_scrollbar = True,
                            font = (fontname,12,"normal"),
                            header_font = (fontname,12,"bold"))
                        self.stage_pts_sheet.place(anchor = 'c', relx = 0.5, rely = 0.4)

                        # Format the sheet
                        self.stage_pts_sheet.change_theme("blue")
                        self.stage_pts_sheet.set_index_width(0)
                        self.stage_pts_sheet.column_width(column = 0, width = 100, redraw = True)
                        self.stage_pts_sheet.column_width(column = 1, width = 100, redraw = True)
                        self.stage_pts_sheet.table_align(align = 'c',redraw=True)

                        # Enable Bindings
                        self.stage_pts_sheet.enable_bindings('single_select','cell_select', 'column_select',"arrowkeys")
                        self.stage_pts_sheet.extra_bindings([("cell_select", self.cell_select_stage)]) 
                        
                        # Fill existing values values
                        for i in range(len(rows)):
                            self.stage_pts_sheet.set_cell_data(i,0,self.Compare['Data'][self.test_name]['Stage Type'][i])
                            self.stage_pts_sheet.set_cell_data(i,1,self.Compare['Data'][self.test_name]['Stage Divisions'][i]) 
                        
                        # Update the sheet
                        self.stage_pts_sheet.redraw()

                        # Reduce Data Points
                        def GetReducedPts():
                            #------------------------------------------------------
                            #
                            #   PURPOSE: Get reduced data points from the button
                            #            press
                            #------------------------------------------------------

                            # Reset the window
                            self.window = 0

                            # Get the number of divisions for each stage
                            for i in range(len(self.stage_pts_sheet.data)):
                                divp = int(self.stage_pts_sheet.data[i][1])
                                if divp < 0:
                                    divp = 0
                                self.Compare['Data'][self.test_name]['Stage Divisions'][i] = divp

                            # Reduce the data
                            self.reduce_data(self.test_name, self.load_dir)

                            # Destory the window
                            root.destroy()

                            # Update the data points
                            for i in range(len(self.Compare['Data'][self.test_name]['Stage Divisions'])):
                                self.stage_table.set_cell_data(i, 6, self.Compare['Data'][self.test_name]['Stage Divisions'][i])
                            self.stage_table.redraw()

                            # Recall the plotting function
                            self.plotter()
                                
                        # Create button to get the reduced data points
                        self.btn_get_red = tk.Button(root, text = "Get Data Points", command = GetReducedPts, 
                                        font = (fontname, fsize_s), bg = '#fc3d21', fg='white',
                                        width = 12)
                        self.btn_get_red.place(anchor = 'c', relx = 0.5, rely = 0.9)

                        def on_closing_root(self):
                            #------------------------------------------------------
                            #
                            #   PURPOSE: Create exit protocol for the Segmentation
                            #            Control Panel
                            #
                            #-------------------------------------------------------

                            # Reset the window
                            self.window = 0
                            root.destroy()

                        # Add the exit protocol to the root
                        root.protocol("WM_DELETE_WINDOW", lambda:on_closing_root(self))

                # Create button to reduce data
                if hasattr(self, 'btn_loc3'):
                    self.btn_loc3.destroy()
                self.btn_loc3 = tk.Button(window, text = "Reduce Data", command = ReduceData, 
                                font = (fontname, fsize_s), bg = '#fc3d21', fg='white',
                                width = 12)
                self.btn_loc3.place(anchor = 'e', relx = 0.725, rely = 0.9)
                self.tab_att_list.append('self.btn_loc3')

    def reduce_data(self, test, load_dir):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: Create the reduced data arrays
        #
        #--------------------------------------------------------------------------

        # Reset the Reduced Arrays
        self.Compare['Data'][test]['Reduced Data']['Time'] = []
        dir_keys = list(self.Compare['Data'][test]['Strain'].keys())
        for key in dir_keys:
            self.Compare['Data'][test]['Reduced Data']['Strain'][key] = []
        dir_keys = list(self.Compare['Data'][test]['Stress'].keys())
        for key in dir_keys:
            self.Compare['Data'][test]['Reduced Data']['Stress'][key] = []

        # Loop through the stages to fit points
        for i in range(len(self.Compare['Data'][test]['Stage Type'])):
            # Get Index information
            if i == 0:
                sindex = 0
            else:
                sindex = self.Compare['Data'][test]['Stage Index'][i-1]
            eindex = self.Compare['Data'][test]['Stage Index'][i]

            # Prealloacte data
            data = []

            # Get the reduced Time
            data.append(self.Compare['Data'][test]['Time'][sindex:eindex])
            dir_keys = list(self.Compare['Data'][test]['Strain'].keys())
            for key in dir_keys:
                data.append(self.Compare['Data'][test]['Strain'][key][sindex:eindex])
            dir_keys = list(self.Compare['Data'][test]['Stress'].keys())
            for key in dir_keys:
                data.append(self.Compare['Data'][test]['Stress'][key][sindex:eindex])

            # Get the X and Y Data for the Stage (Response Curves)
            # - Tensile -> Stress vs Strain
            if self.Compare['Data'][test]['Stage Type'][i] == 'Tensile':
                x = self.Compare['Data'][test]['Strain'][load_dir][sindex:eindex]
                y = self.Compare['Data'][test]['Stress'][load_dir][sindex:eindex]

            # - Creep -> Strain vs Time
            if self.Compare['Data'][test]['Stage Type'][i] == 'Creep':
                x = self.Compare['Data'][test]['Time'][sindex:eindex]
                y = self.Compare['Data'][test]['Strain'][load_dir][sindex:eindex]

            # - Relaxation -> Stress vs Time
            if self.Compare['Data'][test]['Stage Type'][i] == 'Relaxation':
                x = self.Compare['Data'][test]['Time'][sindex:eindex]
                y = self.Compare['Data'][test]['Strain'][load_dir][sindex:eindex]

            # Reduce and Smooth Data with Cubic Spline
            xall = x
            yall = y
            x = [xall[0]]
            y = [yall[0]]
            ct = 1
            while ct < len(xall) -1:
                if self.Compare['Data'][test]['Load Rate'][i][0] >= 0:
                    if xall[ct] > x[-1]:
                        x.append(xall[ct])
                        y.append(yall[ct])
                    rev_flag = 0
                else: 
                    if xall[ct] < x[-1]:
                        x.append(xall[ct])
                        y.append(yall[ct])
                    rev_flag = 1
                ct = ct + 1
            x = np.array(x)
            y = np.array(y)

            # Reverse unloading for increasing X
            if rev_flag == 1:
                x = x[::-1]
                y = y[::-1]

            # Fit cubic spline
            x_cs = np.linspace(x.min(), x.max(), 10000)
            y_cs = CubicSpline(x, y, bc_type='not-a-knot')
            y_cs = y_cs(x_cs)

            # Flip unloading x and y back
            if rev_flag == 1:
                x = x[::-1]
                y = y[::-1]

            # Normalize
            x_cs_n = (x_cs - min(x_cs))/(max(x_cs)-min(x_cs))
            y_cs_n = (y_cs - min(y_cs))/(max(y_cs)-min(y_cs))

            # Initialize total curve length and set number of points
            LT = 0
            divp = self.Compare['Data'][test]['Stage Divisions'][i]

            # Get the reduced data points
            if divp > 0:
                # Get the total curve length
                for j in range(1,len(x_cs_n)):
                    LT = LT + ((x_cs_n[j]-x_cs_n[j-1])**2+(y_cs_n[j]-y_cs_n[j-1])**2)**(0.5)

                # Get the curve length of each segment
                Lbar = LT/divp

                # Initialize X and Y
                x_pts = [x_cs_n[0]]
                y_pts = [y_cs_n[0]]

                # Iterate next point until the curve length is met
                idx = 0
                k = idx+1
                for j in range(divp):
                    flag = 0
                    Li = 0
                    while flag == 0:
                        Li = Li + ((x_cs_n[k]-x_cs_n[k-1])**(2)+(y_cs_n[k]-y_cs_n[k-1])**(2))**(0.5)

                        if Li > Lbar:
                            idx = k
                            vec = [x_cs_n[k]-x_cs_n[k-1],y_cs_n[k]-y_cs_n[k-1]]
                            vec = vec/((vec[0])**2+(vec[1])**2)**(0.5)
                            vec = vec*(Lbar-LiPrev)
                            x_pts.append(x_cs_n[k-1]+vec[0])
                            y_pts.append(y_cs_n[k-1]+vec[1])
                            flag = 1

                        if k == len(y_cs_n)-1:
                            x_pts.append(x_cs_n[k-1]+vec[0])
                            y_pts.append(y_cs_n[k-1]+vec[1])
                            flag = 1

                        k = k+1
                        LiPrev = Li

                # Unnormalize
                x_pts = np.array(x_pts)*(max(x_cs)-min(x_cs)) + min(x_cs)
                y_pts = np.array(y_pts)*(max(y_cs)-min(y_cs)) + min(y_cs)

                # Inerpolate data points
                data_interp = []
                for i in range(len(data)):
                    new_vec = np.interp(x_pts, xall, data[i])
                    data_interp.append(new_vec)

                # Store Interpolated Points
                # -- Time
                for j in range(len(data_interp[0])):
                    self.Compare['Data'][test]['Reduced Data']['Time'].append(data_interp[0][j])
                ct = 1
                # -- Strain
                dir_keys = list(self.Compare['Data'][test]['Strain'].keys())
                for key in dir_keys:
                    for j in range(len(data_interp[ct])):
                        self.Compare['Data'][test]['Reduced Data']['Strain'][key].append(data_interp[ct][j])
                    ct= ct+1
                # -- Stress
                dir_keys = list(self.Compare['Data'][test]['Stress'].keys())
                for key in dir_keys:
                    for j in range(len(data_interp[ct])):
                        self.Compare['Data'][test]['Reduced Data']['Stress'][key].append(data_interp[ct][j])
                    ct= ct+1

    def add_point(self, event):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: Add a data reduction point
        #
        #--------------------------------------------------------------------------

        # Check Clicked
        if self.clicked == 1:

            # Update Clicked for adding a point
            self.clicked = 2

            # Get the selected point
            self.x = event.xdata
            self.y= event.ydata

            # Plot the closest point
            mind = 10e3
            mindidx = -1
            for i in range(len(self.xdata)):
                d = ((self.x-self.xdata[i])**2+(self.y-self.ydata[i])**2)**(1/2)
                if d < mind:
                    mind = d
                    mindidx = i

            # Store the data
            self.idx = mindidx
            self.xc = self.xdata[mindidx]
            self.yc= self.ydata[mindidx]

            # Update the canvas
            self.plot1.plot(self.xc, self.yc,'go')
            self.canvas.draw()

            # Bind the Right and Left Arrow Keys
            self.canvas.get_tk_widget().bind("<Right>", self.move_right)
            self.canvas.get_tk_widget().bind("<Left>", self.move_left)

            # Bind 'Return'
            self.canvas.get_tk_widget().bind("<Return>", self.update_pts)

            # Set Bindings
            self.canvas.get_tk_widget().focus_set()
        
    def del_point(self, event):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: Delete a data reduction point
        #
        #--------------------------------------------------------------------------

        # Check Clicked
        if self.clicked == 1:

            # Update the clicked for deleting a point
            self.clicked = 3

            # Get the selected point
            self.x = event.xdata
            self.y= event.ydata

            # Plot the closest point
            mind = 10e3
            mindidx = -1
            for i in range(len(self.xsdata)):
                d = ((self.x-self.xsdata[i])**2+(self.y-self.ysdata[i])**2)**(1/2)
                if d < mind:
                    mind = d
                    mindidx = i

            # Store the data
            self.idx = mindidx
            self.xc = self.xsdata[mindidx]
            self.yc= self.ysdata[mindidx]

            # Update the canvas
            self.plot1.plot(self.xc, self.yc,'go')
            self.canvas.draw()

            # Bind the Right and Left Arrow Keys
            self.canvas.get_tk_widget().bind("<Right>", self.move_right)
            self.canvas.get_tk_widget().bind("<Left>", self.move_left)

            # Bind 'Return'
            self.canvas.get_tk_widget().bind("<Return>", self.update_pts)

            # Set Bidnings
            self.canvas.get_tk_widget().focus_set()

    def move_left(self, event):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: Move Data Reduction Point Left
        #
        #--------------------------------------------------------------------------

        # Remove the Previous Point
        self.plot1.lines[len(self.plot1.lines)-1].remove()
        self.canvas.draw()

        # Update the index
        # -- For adding a point
        if self.clicked == 2:
            self.idx  = self.idx - 1
            if self.idx < 0:
                self.idx = 0
            self.xc = self.xdata[self.idx]
            self.yc= self.ydata[self.idx]

        # -- For deleting a point
        if self.clicked == 3:
            self.idx  = self.idx - 1
            if self.idx < 0:
                self.idx = 0
            self.xc = self.xsdata[self.idx]
            self.yc= self.ysdata[self.idx]

        # Replot the data
        self.plot1.plot(self.xc, self.yc,'go')
        self.canvas.draw()

    def move_right(self, event):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: Move Data Reduction Point Right
        #
        #--------------------------------------------------------------------------
        
        # Remove the Previous Point
        self.plot1.lines[len(self.plot1.lines)-1].remove()
        self.canvas.draw()

        # Update the index
        self.idx  = self.idx + 1
            
        # -- For Adding a point
        if self.clicked == 2:
            if self.idx > len(self.xdata)-1:
                self.idx = len(self.xdata)-1
            self.xc = self.xdata[self.idx]
            self.yc= self.ydata[self.idx]

        # -- For deleting a point
        if self.clicked == 3:
            if self.idx > len(self.xsdata)-1:
                self.idx = len(self.xsdata)-1
            self.xc = self.xsdata[self.idx]
            self.yc= self.ysdata[self.idx]

        # Replot the data
        self.plot1.plot(self.xc, self.yc,'go')
        self.canvas.draw()

    def update_pts(self, event):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: Update the current data reduction point
        #
        #--------------------------------------------------------------------------

        # Adding a point
        if self.clicked == 2:

            # Get the current index
            idx = self.idx

            # Update Time
            new_time = []
            time = self.Compare['Data'][self.test_name]['Time'][idx]
            idx_insert = None
            i = 0
            while time > self.Compare['Data'][self.test_name]['Reduced Data']['Time'][i]:
                new_time.append(self.Compare['Data'][self.test_name]['Reduced Data']['Time'][i])
                i = i+1
            idx_insert = i
            new_time.append(time)
            i = i+1
            for i in range(len(self.Compare['Data'][self.test_name]['Reduced Data']['Time'])):
                if time < self.Compare['Data'][self.test_name]['Reduced Data']['Time'][i]:
                    new_time.append(self.Compare['Data'][self.test_name]['Reduced Data']['Time'][i])
            self.Compare['Data'][self.test_name]['Reduced Data']['Time'] = new_time

            # Update Strain
            keys = list(self.Compare['Data'][self.test_name]['Reduced Data']['Strain'].keys())
            for key in keys:
                strain = self.Compare['Data'][self.test_name]['Strain'][key][idx]
                new_strain = []
                for i in range(idx_insert):
                    new_strain.append(self.Compare['Data'][self.test_name]['Reduced Data']['Strain'][key][i])
                new_strain.append(strain)
                for i in range(idx_insert,len(self.Compare['Data'][self.test_name]['Reduced Data']['Strain'][key])):
                    new_strain.append(self.Compare['Data'][self.test_name]['Reduced Data']['Strain'][key][i])
                self.Compare['Data'][self.test_name]['Reduced Data']['Strain'][key] = new_strain

            # Update Stress
            keys = list(self.Compare['Data'][self.test_name]['Reduced Data']['Stress'].keys())
            for key in keys:
                stress = self.Compare['Data'][self.test_name]['Stress'][key][idx]
                new_stress = []
                for i in range(idx_insert):
                    new_stress.append(self.Compare['Data'][self.test_name]['Reduced Data']['Stress'][key][i])
                new_stress.append(stress)
                for i in range(idx_insert,len(self.Compare['Data'][self.test_name]['Reduced Data']['Stress'][key])):
                    new_stress.append(self.Compare['Data'][self.test_name]['Reduced Data']['Stress'][key][i])
                self.Compare['Data'][self.test_name]['Reduced Data']['Stress'][key] = new_stress

            # Find the stage that was changed
            stage_idx = [0] + self.Compare['Data'][self.test_name]['Stage Index']
            for i in range(len(self.Compare['Data'][self.test_name]['Stage Index'])):
                if  idx > stage_idx[i] and idx < stage_idx[i+1]:
                    self.Compare['Data'][self.test_name]['Stage Divisions'][i] = self.Compare['Data'][self.test_name]['Stage Divisions'][i]+1

            # Update the stage table
            for i in range(len(self.Compare['Data'][self.test_name]['Stage Divisions'])):
                self.stage_table.set_cell_data(i, 6, self.Compare['Data'][self.test_name]['Stage Divisions'][i])
            self.stage_table.redraw()

        # Deleting a point
        if self.clicked == 3:
            # Confirm delete
            askyn = messagebox.askyesno(title = 'Delete point', message = 'Do you want to delete this point?')
            if askyn == True:
                # Delete the Time Point
                self.Compare['Data'][self.test_name]['Reduced Data']['Time'] = np.delete(self.Compare['Data'][self.test_name]['Reduced Data']['Time'],[self.idx])
                
                # Delete the Strain Point
                keys = list(self.Compare['Data'][self.test_name]['Reduced Data']['Strain'].keys())
                for key in keys:
                    self.Compare['Data'][self.test_name]['Reduced Data']['Strain'][key] = np.delete(self.Compare['Data'][self.test_name]['Reduced Data']['Strain'][key],[self.idx])
                
                # Delete the Stress Point
                keys = list(self.Compare['Data'][self.test_name]['Reduced Data']['Stress'].keys())
                for key in keys:
                    self.Compare['Data'][self.test_name]['Reduced Data']['Stress'][key] = np.delete(self.Compare['Data'][self.test_name]['Reduced Data']['Stress'][key],[self.idx])
        
                # Find the stage that was changed
                stage_idx = [0] + self.Compare['Data'][self.test_name]['Stage Index']
                for i in range(len(self.Compare['Data'][self.test_name]['Stage Index'])):
                    if  self.Compare['Data'][self.test_name]['Time'][self.idx] > self.Compare['Data'][self.test_name]['Time'][stage_idx[i]] and self.Compare['Data'][self.test_name]['Time'][self.idx] < self.Compare['Data'][self.test_name]['Time'][stage_idx[i+1]]:
                        self.Compare['Data'][self.test_name]['Stage Divisions'][i] = self.Compare['Data'][self.test_name]['Stage Divisions'][i]-1

                # Update the stage table
                for i in range(len(self.Compare['Data'][self.test_name]['Stage Divisions'])):
                    self.stage_table.set_cell_data(i, 6, self.Compare['Data'][self.test_name]['Stage Divisions'][i])
                self.stage_table.redraw()

        # Reset the clicked variable
        self.clicked = 0

        # Replot data
        self.plotter()

    def plotter_all(self, value, tag):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: Plot all curves on the same plot.
        #
        #--------------------------------------------------------------------------

        #  Delete the canvas and drop down if it exists
        if hasattr(self, 'canvas'):
            self.toolbar.destroy()
            self.canvas.get_tk_widget().destroy()
            del self.canvas

        # Delete local attributes
        atts = ['self.btn_loc3', 'self.btn_loc4', 'self.btn_loc5']
        for widget in atts:
            try:
                eval(widget).destroy()
            except:
                pass

        # Reset any formatting on the test sheet
        table_name = 'self.' + tag
        for i in range(len(eval(table_name).data)):
            eval(table_name).highlight_rows(i,'white','black')

        # Create the plot
        self.fig = Figure(figsize=(5,3.6), dpi = 125, constrained_layout = True)
        self.plot1 = self.fig.add_subplot(111)

        # Initialize the display curves
        disp_tests = []
        disp_load = []
        disp_type = []

        # Get display curves for the database tab
        if hasattr(self,'sheet_db'):
            if self.sheet_db.winfo_exists():
                for i in range(len(self.sheet_db.data)):
                    if self.sheet_db.data[i][0] == True:
                        if value in [self.sheet_db.data[i][2], 'All']:
                            disp_tests.append(self.sheet_db.data[i][1])
                            disp_load.append(int(self.sheet_db.data[i][4]))
                            disp_type.append(self.sheet_db.data[i][2])

        # Get display curves for the characterization tab
        if hasattr(self,'sheet_char'):
            if self.sheet_char.winfo_exists():
                for i in range(len(self.sheet_char.data)):
                    if value in [self.sheet_char.data[i][1], 'All']:
                        disp_tests.append(self.sheet_char.data[i][0])
                        disp_load.append(int(self.sheet_char.data[i][3]))
                        disp_type.append(self.sheet_char.data[i][1])

        # Loop through all tests selected
        for test in disp_tests:
            if disp_type[disp_tests.index(test)] == 'Tensile':
                end_idx = self.Compare['Data'][test]['Stage Index'][0]
            if disp_type[disp_tests.index(test)] == 'Creep':
                end_idx = self.Compare['Data'][test]['Stage Index'][1]
            if disp_type[disp_tests.index(test)] == 'Relaxation':
                end_idx = self.Compare['Data'][test]['Stage Index'][1]
            if disp_type[disp_tests.index(test)] == 'Generic':
                end_idx = self.Compare['Data'][test]['Stage Index'][-1]
            
            # Plot the data
            if value == 'Tensile' or value == 'Generic' or value == 'All':
                self.plot1.plot(self.Compare['Data'][test]['Strain'][disp_load[disp_tests.index(test)]][:end_idx],
                                self.Compare['Data'][test]['Stress'][disp_load[disp_tests.index(test)]][:end_idx],
                                label=test)
                xlab = 'Strain'
                ylab = 'Stress [MPa]'
            elif value == 'Creep':
                self.plot1.plot(self.Compare['Data'][test]['Time'][:end_idx],
                                self.Compare['Data'][test]['Strain'][disp_load[disp_tests.index(test)]][:end_idx],
                                label=test)
                xlab = 'Time [s]'
                ylab = 'Strain'
            elif value == 'Relaxation':
                self.plot1.plot(self.Compare['Data'][test]['Time'][:end_idx],
                                self.Compare['Data'][test]['Stress'][disp_load[disp_tests.index(test)]][:end_idx],
                                label=test)
                xlab = 'Time [s]'
                ylab = 'Stress [MPa]'

        if len(disp_tests) > 0:
            # Format the plot
            xlab_frmt = ScalarFormatter() 
            ylab_frmt = ScalarFormatter()
            self.plot1.set_xlabel(xlab)
            self.plot1.set_ylabel(ylab)
            self.plot1.xaxis.set_major_formatter(xlab_frmt)
            self.plot1.yaxis.set_major_formatter(ylab_frmt)
            if "Strain" in xlab or "Time" in xlab:
                self.plot1.ticklabel_format(style='sci',scilimits=(-6,-3),axis='x')
            if "Strain" in ylab or "Time" in ylab:
                self.plot1.ticklabel_format(style='sci',scilimits=(-6,-3),axis='y')
            self.plot1.legend()

            # Create the Tkinter canvas
            self.canvas = FigureCanvasTkAgg(self.fig, master = window)
            self.canvas.draw()

            # Create the Matplotlib toolbar
            self.toolbar = NavigationToolbar2Tk(self.canvas, window)
            self.toolbar.update()

            # Format Toolbar
            self.toolbar.config(bg=bg_color)
            self.toolbar._message_label.config(background=bg_color)
            self.toolbar.place(anchor = 'e', relx = 0.975, rely = 0.9)

            # Add the figure to the GUI
            self.canvas.get_tk_widget().place(anchor = 'n', relx = 0.78, rely = 0.32)
            if 'self.canvas' not in self.tab_att_list:
                self.tab_att_list.append('self.canvas')

    #------------------------------------------------------------------------------------------------------------------------------------------
    #
    #   MODEL PAGES
    #   Model-Optimize and Model-Analysis Pages
    #   The Model-Optimize page allows users to define a model and run the COMPARE enginer for parameter optimization.
    #   The Model-Analysis page allows users to manually define parameters and run evaluation
    #
    #------------------------------------------------------------------------------------------------------------------------------------------

    def model_tab(self):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: Create the Optimize Model Page.
        #
        #--------------------------------------------------------------------------

        # Delete any local attributes
        DeleteLocal(self)

        # Create the Optimize Model page
        CreateModelTab(self,window,frmt)

    def analyze_tab(self):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: Create the Analyze Model Page.
        #
        #--------------------------------------------------------------------------

        # Delete any local attributes
        DeleteLocal(self)

        # Create the Analyze Model page
        CreateAnalysisTab(self,window,frmt)

    def cell_select_opt(self, response, tag):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: Create Custom Column Disablement for Model-Optimization 
        #            Parameters.
        #
        #--------------------------------------------------------------------------

        # Define the table
        table_name = "self." + tag

        # Define locked columns (parameter name and COMPARE value)
        locked_cols = [0,6]

        # Enable/Disable user ability to edit cells
        if response.selected.column != None:
            if response.selected.column in locked_cols:
                eval(table_name).disable_bindings(("edit_cell"))
            else:
                self.sheet_data = eval(table_name).data
                self.sheet_tag = tag
                eval(table_name).enable_bindings(("edit_cell"))
                eval(table_name).extra_bindings([("edit_cell", self.save_model)])

                # Check bounds
                tags = ['sheet1', 'sheet2']
                for tag_i in tags:
                    if hasattr(self,tag_i):
                        tag_name = 'self.' + tag_i
                        for i in range(len(eval(tag_name).data)):

                            # -- Check lower bound
                            eval(tag_name).highlight((i,2),fg = 'black', bg = 'white')
                            try:
                                if float(eval(tag_name).data[i][2]) > float(eval(tag_name).data[i][3]):
                                    eval(tag_name).highlight((i,2),fg = 'red', bg = 'white')
                            except:
                                pass

                            # -- Check upper bound
                            eval(tag_name).highlight((i,4),fg = 'black', bg = 'white')
                            try:
                                if float(eval(tag_name).data[i][4]) < float(eval(tag_name).data[i][3]):
                                    eval(tag_name).highlight((i,4),fg = 'red', bg = 'white')
                            except:
                                pass
            eval(table_name).redraw()

    def save_model(self, response):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: Save an optimized model before editing parameters.
        #
        #--------------------------------------------------------------------------

        # Determine if optimized parameters were found
        if self.optimize == 1:

            # Ask user if they want to save
            askyn = messagebox.askyesno(title = 'Save Model', message = 'Editing the model tables will remove any optimized parameters. Do you want to save the model?')
            if askyn == True:
                # Reset the model status
                self.Compare['Model']['Status'] = 0

                # Get the model name
                save_flag = 0
                while save_flag == 0:
                    user_input = simpledialog.askstring("Save Model", "Enter the model name:")

                    if user_input in list(self.Compare['Model Library'].keys()):
                        askyn = messagebox.askyesno(title = 'Save Model', message = 'Do you want to overwrite ' + user_input + ' ?')
                        if askyn == True:
                            save_flag = 1
                    else:
                        save_flag = 1

                # Get the model type
                if len(self.Compare['Model']['VE_Params'][0]) > 3:
                    self.Compare['Model']['Compare Type'] = 'Optimize'
                else:
                    self.Compare['Model']['Compare Type'] = 'Analysis'
                
                # Write model data to binary in the mode library
                json_string = json.dumps(self.Compare['Model'])
                binary_data = json_string.encode('utf-8')
                self.Compare['Model Library'][user_input] = binary_data

            # Reset Optimization
            self.optimize = 0

        # Reset Parameter Values in the selected row
        # -- Viscoelastic Parameters
        if hasattr(self,'sheet1'):
            try:
                currently_selected = self.sheet1.get_currently_selected()
                self.sheet1.set_cell_data(currently_selected.row,self.sheet1.visible_columns[1]-1,'')
            except:
                pass

        # -- Viscoplastic Parameters
        if hasattr(self,'sheet2'):
            try:
                currently_selected = self.sheet2.get_currently_selected()
                self.sheet2.set_cell_data(currently_selected.row,self.sheet2.visible_columns[1]-1,'')
            except:
                pass


    def cell_select_anly(self, response, tag):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: Create Custom Column Disablement for Model-Analysis 
        #            Parameters.
        #
        #--------------------------------------------------------------------------

        # Define the table
        table_name = "self." + tag

        # Define locked columns (parameter name)
        locked_cols = [0]

        # Enable/Disable user ability to edit cells
        if response.selected.column != None:
            if response.selected.column in locked_cols:
                eval(table_name).disable_bindings(("edit_cell"))
            else:
                eval(table_name).enable_bindings(("edit_cell"))
            eval(table_name).redraw()

    def load_from_db(self, tag):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: Load a model from the excel template
        #
        #--------------------------------------------------------------------------

        # Get the Excel File
        file = filedialog.askopenfile(title = "Model Import", filetypes= [('Excel', '*.xlsx')], mode ='r',)

        # Get Data
        model, flag, msg = ReadModel(file.name, self)

        # Set NaN units as empty
        if 'VE_Param' in model.keys():
            for i in range(len(model['VE_Param'])):
                if pd.isna(model['VE_Param'][i][1]) == True:
                    model['VE_Param'][i][1] = ""
        if 'VP_Param' in model.keys():
            for i in range(len(model['VP_Param'])):
                if pd.isna(model['VP_Param'][i][1]) == True:
                    model['VP_Param'][i][1] = ""

        # Update the model and corresponding page
        if flag == 0:
            if tag == "Optimize":
                # Save the Model
                self.Compare['Model'] = copy.deepcopy(model)

                # Reformat the Parameters with empty bounds
                self.Compare['Model']['VE_Param'] = []
                for i in range(len(model['VE_Param'])):
                    self.Compare['Model']['VE_Param'].append([model['VE_Param'][i][0],
                                                            model['VE_Param'][i][1],
                                                            '',
                                                            model['VE_Param'][i][2],
                                                            '',
                                                            'Active',
                                                            ''
                                                            ])                                        
                    self.Compare['Model']['VP_Param'] = []
                    for i in range(len(model['VP_Param'])):
                        self.Compare['Model']['VP_Param'].append([model['VP_Param'][i][0],
                                                                model['VP_Param'][i][1],
                                                                '',
                                                                model['VP_Param'][i][2],
                                                                '',
                                                                'Active',
                                                                ''
                                                                ])

                # Set the Sheet Data
                self.sheet1_data = self.Compare['Model']['VE_Param']
                self.sheet2_data = self.Compare['Model']['VP_Param']
                self.res_flag1 = 1
                self.res_flag2 = 1
                
                # Recreate the Optimize Page
                DeleteLocal(self)
                CreateModelTab(self,window,frmt)

            else:
                # Save the Model
                self.Compare['Analysis'] = copy.deepcopy(model)

                # Set the Sheet Data
                self.sheet_anly1_data = self.Compare['Analysis']['VE_Param']
                self.sheet_anly2_data = self.Compare['Analysis']['VP_Param']
                self.res_flag1 = 1
                self.res_flag2 = 1

                # Recreate teh Analysis Page
                DeleteLocal(self)
                CreateAnalysisTab(self,window,frmt)
        else:
            messagebox.showerror(message=msg)

    def Model_Library(self, tag):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: View/Edit Model Libary.
        #
        #--------------------------------------------------------------------------

        # Only open if not already open
        if self.clicked != 1:
            self.clicked = 1

            # Get Existing Models
            self.models = list(self.Compare['Model Library'].keys())

            if len(self.models) > 0:
                # Create a window to view the model libary
                root = tk.Tk()
                root.config(width=900, height=700)
                root.title("Model Libary")

                def on_closing_ML(self):
                    #--------------------------------------------------------------
                    #
                    #   PURPOSE: Set exit protocol for closing the model window.
                    #
                    #--------------------------------------------------------------

                    # Reset clicked and destory the window
                    self.clicked = 0
                    root.destroy()

                # Create Window Exit Protocol
                root.protocol("WM_DELETE_WINDOW", lambda:on_closing_ML(self))

                def create_sheet():
                    #--------------------------------------------------------------
                    #
                    #   PURPOSE: Create the Material Library Sheet.
                    #
                    #--------------------------------------------------------------

                    # Set column names
                    Cols = ['Name', 'Type', 'Reversible Model','Irreversible Model','Method']

                    # Create the table
                    self.sheet_lib = tksheet.Sheet(root, total_rows = len(self.models), total_columns = len(Cols), 
                                    headers = Cols,
                                    width = 800, height = 600, show_x_scrollbar = False, show_y_scrollbar = True,
                                    font = (fontname,12,"normal"),
                                    header_font = (fontname,12,"bold"))
                    self.sheet_lib.place(anchor = 'c', relx = 0.5, rely = 0.5)
                    self.sheet_lib.change_theme("blue")
                    self.sheet_lib.set_index_width(0)
           
                    def rename_model(self):
                        #----------------------------------------------------------
                        #
                        #   PURPOSE: Rename a model.
                        #
                        #----------------------------------------------------------

                        # Get the previous name
                        currently_selected = self.sheet_lib.get_currently_selected()
                        prev_name = self.sheet_lib.data[currently_selected.row][0]

                        # Get the model name
                        save_flag = 0
                        while save_flag == 0:
                            user_input = simpledialog.askstring("Save Model", "Enter the model name:")

                            if user_input in list(self.Compare['Model Library'].keys()):
                                askyn = messagebox.askyesno(title = 'Save Model', message = 'Do you want to overwrite ' + user_input + ' ?')
                                if askyn == True:
                                    save_flag = 1
                            else:
                                save_flag = 1

                        # Save the new data
                        self.Compare['Model Library'][user_input] = self.Compare['Model Library'][prev_name] 

                        # Delete the old data
                        del self.Compare['Model Library'][prev_name] 

                        # Reset the model list
                        self.models = list(self.Compare['Model Library'].keys())

                        # Recreate the sheet
                        self.sheet_lib.destroy()
                        del self.sheet_lib
                        create_sheet()
                    
                    def delete_model(self):
                        #----------------------------------------------------------
                        #
                        #   PURPOSE: Delete a model from the library.
                        #
                        #----------------------------------------------------------

                        # Confirm delete
                        askyn = messagebox.askyesno(title = 'Delete Model', message = 'Do you want to delete this model?')
                        if askyn == True:
                            # Get the previous name
                            currently_selected = self.sheet_lib.get_currently_selected()
                            prev_name = self.sheet_lib.data[currently_selected.row][0]

                            # Delete the data
                            del self.Compare['Model Library'][prev_name] 

                            # Reset the model list
                            self.models = list(self.Compare['Model Library'].keys())

                            # Recreate the sheet
                            self.sheet_lib.destroy()
                            del self.sheet_lib
                            create_sheet()


                    def load_model(self):
                        #----------------------------------------------------------
                        #
                        #   PURPOSE: Load a model from the library into the page.
                        #
                        #----------------------------------------------------------

                        # Get the previous name
                        currently_selected = self.sheet_lib.get_currently_selected()
                        name = self.sheet_lib.data[currently_selected.row][0]

                        # Set the model name
                        self.Compare['Model ID'] = name

                        # Close the window
                        root.destroy()

                        # Get the data
                        json_string = self.Compare['Model Library'][name].decode('utf-8')
                        data = json.loads(json_string)

                        # Populate the Optimization Page
                        if tag == 'Optimize':
                            if data['Compare Type'] == 'Optimize':
                                self.Compare['Model'] = data   
                            else:
                                self.Compare['Model'] = dict(data) 
                                # Reformat Parameters
                                if 'VE_Param' in list(data.keys()):
                                    self.Compare['Model']['VE_Param'] = []
                                    for i in range(len(data['VE_Param'])):
                                        self.Compare['Model']['VE_Param'].append([data['VE_Param'][i][0],
                                                                                data['VE_Param'][i][1],
                                                                                '',
                                                                                data['VE_Param'][i][2],
                                                                                '',
                                                                                'Active',
                                                                                ''
                                                                                ])                                      
                                if 'VP_Param' in list(data.keys()):
                                    self.Compare['Model']['VP_Param'] = []
                                    for i in range(len(data['VP_Param'])):
                                        self.Compare['Model']['VP_Param'].append([data['VP_Param'][i][0],
                                                                                data['VP_Param'][i][1],
                                                                                '',
                                                                                data['VP_Param'][i][2],
                                                                                '',
                                                                                'Active',
                                                                                ''
                                                                                ])
                            
                            # Recreate the Optimize Page
                            DeleteLocal(self)
                            CreateModelTab(self,window,frmt)

                        # Populate the Analysis Page
                        else:
                            if data['Compare Type'] == 'Analysis':
                                self.Compare['Analysis'] = data   
                            else:
                                self.Compare['Analysis'] = dict(data) 
                                # Reformat Parameters
                                if 'VE_Param' in list(data.keys()):
                                    self.Compare['Analysis']['VE_Param'] = []
                                    for i in range(len(data['VE_Param'])):
                                        self.Compare['Analysis']['VE_Param'].append([data['VE_Param'][i][0],
                                                                                data['VE_Param'][i][1],
                                                                                data['VE_Param'][i][6],
                                                                                ])                                       
                                if 'VP_Param' in list(data.keys()):
                                    self.Compare['Analysis']['VP_Param'] = []
                                    for i in range(len(data['VP_Param'])):
                                        self.Compare['Analysis']['VP_Param'].append([data['VP_Param'][i][0],
                                                                                data['VP_Param'][i][1],
                                                                                data['VP_Param'][i][6],
                                                                                ])

                            # Recreate Analysis Page
                            DeleteLocal(self)
                            CreateAnalysisTab(self,window,frmt)

                    def view_notes(self):
                        #----------------------------------------------------------
                        #
                        #   PURPOSE: View any model notes.
                        #
                        #----------------------------------------------------------

                        # Set the structure name
                        currently_selected = self.sheet_lib.get_currently_selected()
                        name = self.sheet_lib.data[currently_selected.row][0]
                        json_string = self.Compare['Model Library'][name].decode('utf-8')
                        data = json.loads(json_string)

                        # Get the note if it exists
                        if 'Note' in data.keys():
                            
                            # Get the note
                            note = data['Note']

                        else:
                            note = ''

                        # Set the window flag
                        if hasattr(self,'note_click_v') == False:
                            self.note_click_v = 0

                        # Open the notes window if it doesn't already exit
                        if self.note_click_v == 0:

                            # Reset the flag
                            self.note_click_v = 1

                            # Create the window to display the note
                            root_v = tk.Tk() 
                            root_v.geometry("600x400")
                            root_v.title("Model Notes") 
                            
                            # Create the label
                            ttk.Label(root_v, text="Model Notes:", 
                                        font=(fontname, fsize_s)).place(anchor='n', relx = 0.5, rely = 0.1) 
                            
                            # Create the note
                            text_area = scrolledtext.ScrolledText(root_v, wrap=tk.WORD, 
                                                                width=40, height=8, 
                                                                font=(fontname, fsize_s)) 
                            
                            text_area.place(anchor='c', relx = 0.5, rely = 0.5)
                            text_area.insert(tk.END, note) 
                            text_area.config(state="disabled")

                            # Create Exit Protocol
                            def on_closing_root_v(self):
                                #------------------------------------------------------
                                #
                                #   PURPOSE: Set exit protocol for closing the model 
                                #            window.
                                #
                                #------------------------------------------------------
                                
                                # Reset clicked and destory the window
                                self.note_click_v = 0
                                root_v.destroy()

                            # Add the exit protocol to the root
                            root_v.protocol("WM_DELETE_WINDOW", lambda:on_closing_root_v(self))

                    # Enable Bindings
                    self.sheet_lib.enable_bindings('single_select','cell_select', 'column_select', "arrowkeys", "right_click_popup_menu")
                    self.sheet_lib.popup_menu_add_command('Rename Model', lambda : rename_model(self), table_menu = True, index_menu = True, header_menu = True)
                    self.sheet_lib.popup_menu_add_command('Delete Model', lambda : delete_model(self), table_menu = True, index_menu = True, header_menu = True)
                    self.sheet_lib.popup_menu_add_command('Load Model', lambda : load_model(self), table_menu = True, index_menu = True, header_menu = True)
                    self.sheet_lib.popup_menu_add_command('View Notes', lambda : view_notes(self), table_menu = True, index_menu = True, header_menu = True)

                    # Set Column Widths
                    self.sheet_lib.column_width(column = 0, width = 150, redraw = True)
                    self.sheet_lib.column_width(column = 1, width = 100, redraw = True)
                    self.sheet_lib.column_width(column = 2, width = 200, redraw = True)
                    self.sheet_lib.column_width(column = 3, width = 200, redraw = True)
                    self.sheet_lib.column_width(column = 4, width = 100, redraw = True)
                    self.sheet_lib.table_align(align = 'c',redraw=True)

                    # Populate the sheet
                    for i in range(len(self.models)):
                        json_string = self.Compare['Model Library'][self.models[i]].decode('utf-8')
                        data = json.loads(json_string)
                        self.sheet_lib.set_cell_data(i,0,self.models[i])
                        keys = ['Model Name', 'Reversible Model Name', 'Irreversible Model Name', 'Compare Type']
                        for j in range(len(keys)):
                            if keys[j] in list(data.keys()):
                                self.sheet_lib.set_cell_data(i,j+1,data[keys[j]])

                # Create the material library sheet
                create_sheet()

            else:
                self.clicked = 0
                messagebox.showinfo(message = 'There are no models in the library.')

    def optimizer(self):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: Check all conditions before running optimization.
        #
        #--------------------------------------------------------------------------

        # Set error checking flag
        flag = 0

        # Check all elastic inital guesses and bounds are populated
        if hasattr(self,'sheet1'):
            for i in range(len(self.sheet1.data)):
                try:
                    float(self.sheet1.data[i][2])
                    float(self.sheet1.data[i][3])
                    float(self.sheet1.data[i][4])
                except:
                    flag = 1
                    msg = 'Invalid values for the elastic parameter initial guess and/or bounds.'

        # Check all elastic bounds are valid
        if hasattr(self,'sheet1'):
            if flag == 0:
                for i in range(len(self.sheet1.data)):
                    if float(self.sheet1.data[i][2]) > float(self.sheet1.data[i][3]):
                        flag = 1
                        msg = 'Invalid values for elastic parameter bounds.'
                    if float(self.sheet1.data[i][4]) < float(self.sheet1.data[i][3]):
                        flag = 1
                        msg = 'Invalid values for elastic parameter bounds.'

        # Check all plastic inital guesses and bounds are populated
        if hasattr(self,'sheet2'):
            for i in range(len(self.sheet2.data)):
                try:
                    float(self.sheet2.data[i][2])
                    float(self.sheet2.data[i][3])
                    float(self.sheet2.data[i][4])
                except:
                    flag = 1
                    msg = 'Invalid values for the plastic parameter initial guess and/or bounds.'

        # Check all elastic bounds are valid
        if hasattr(self,'sheet2'):
            if flag == 0:
                for i in range(len(self.sheet2.data)):
                    if float(self.sheet2.data[i][2]) > float(self.sheet2.data[i][3]):
                        flag = 1
                        msg = 'Invalid values for plastic parameter bounds.'
                    if float(self.sheet2.data[i][4]) < float(self.sheet2.data[i][3]):
                        flag = 1
                        msg = 'Invalid values for plastic parameter bounds.'

        # Run Optimization
        if flag == 0:
            self.run_compare_opt()
        else:
            messagebox.showinfo(message=msg)

                    
    def run_compare_opt(self):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: Run COMPARE Optimization.
        #
        #--------------------------------------------------------------------------

        try:
            # Create and clear the Temp Directory
            temp_dir = os.path.join(os.getcwd(),'Temp')
            try:
                os.mkdir(temp_dir)
            except:
                pass
            for filename in os.listdir(temp_dir):
                file_path = os.path.join(temp_dir, filename)
                try:
                    if os.path.isfile(file_path) or os.path.islink(file_path):
                        os.unlink(file_path)
                    elif os.path.isdir(file_path):
                        shutil.rmtree(file_path)
                except Exception as e:
                    print(f"Failed to delete {file_path}. Reason: {e}")

            # Copy the executable
            shutil.copy(self.Compare['Paths']['Compare Executable'], temp_dir)

            # Determine Model Type
            if self.Compare['Model']['Model Name'] == 'GVIPS':
                if self.Compare['Model']['Reversible Model Name'] == 'Isotropic Viscoelastic' and self.Compare['Model']['Irreversible Model Name'] == 'Isotropic Viscoplastic':
                    
                    # Write the DGS file
                    mod, Param, Param_U, Param_N = WriteDSG_GVIPS_OPT_IN(self, temp_dir)


            # Write the Simulation files
            ct = 1
            sim_tests = list(self.Compare['Characterization'].keys())
            for sim_test in sim_tests:
                WriteSIM(self, self.Compare['Characterization'][sim_test], temp_dir, ct, mod, Param)
                ct = ct + 1

            # Write the NLP files
            WriteNLP(temp_dir)

            # Run Compare
            command = 'cmd /k "cd ' + temp_dir + ' & compnasardamage & exit"'
            os.system(command)

            # Read Values
            Vals=  []
            with open(os.path.join(temp_dir,"final.val"), "r") as file:
                for line in file:
                    # Process each line here
                    line_all = line.strip()
                    val = line_all.split(' ')[-1]
                    Vals.append(val)
            
            #Update the Viscoelastic parameters
            if hasattr(self,'sheet1'):
                VE = self.sheet1.data
                for i in range(self.sheet1.visible_rows[1]):
                    val = float(Vals[Param_N[Param.index(VE[i][0])]-1])
                    try:
                        val = UnitConversion(Param_U[Param.index(VE[i][0])], val, VE[i][1], os.path.join(os.getcwd()))
                    except:
                        pass
                    self.sheet1.set_cell_data(i,6,'{:0.4e}'.format(val))
                    if val >= 0.99*float(VE[i][4]) or val <= 1.01*float(VE[i][2]):
                        clr = 'red'
                    else:
                        clr = 'green'
                    self.sheet1.highlight((i,6),fg= clr, bg = 'white')

                self.sheet1.redraw()

            #Update the Viscoplastic parameters
            if hasattr(self,'sheet2'):
                VP = self.sheet2.data
                for i in range(self.sheet2.visible_rows[1]):
                    val = float(Vals[Param_N[Param.index(VP[i][0])]-1])
                    try:
                        val = UnitConversion(Param_U[Param.index(VP[i][0])], val, VP[i][1], os.path.join(os.getcwd()))
                    except:
                        pass
                    self.sheet2.set_cell_data(i,6,'{:0.4e}'.format(val))
                    if val >= 0.99*float(VP[i][4]) or val <= 1.01*float(VP[i][2]):
                        clr = 'red'
                    else:
                        clr = 'green'
                    self.sheet2.highlight((i,6),fg= clr, bg = 'white')

                self.sheet2.redraw()

            # Check that a name exists
            if 'Model ID' not in self.Compare.keys():
               self.Compare['Model ID'] = None 
            if self.Compare['Model ID'] == None:
                # Set the type
                self.Compare['Model']['Compare Type'] = 'Optimize' 

                # Get the save name
                save_flag = 0
                while save_flag == 0:
                    user_input = simpledialog.askstring("Save Model", "Enter the model name:")

                    if user_input in list(self.Compare['Model Library'].keys()):
                        askyn = messagebox.askyesno(title = 'Save Model', message = 'Do you want to overwrite ' + user_input + ' ?')
                        if askyn == True:
                            save_flag = 1
                    else:
                        save_flag = 1

                # Set the model name
                self.Compare['Model ID'] = user_input

            # Save to binary in the model library
            json_string = json.dumps(self.Compare['Model'])
            binary_data = json_string.encode('utf-8')
            self.Compare['Model Library'][self.Compare['Model ID']] = binary_data

            # Set model status to 1
            self.Compare['Model']['Status'] = 1

            # Get Test Error
            out_file = os.path.join(temp_dir,"comp.out")
            line_out = []
            line_ct = 1
            line_exp = []
            with open(out_file, "r") as file:
                    for line in file:
                        line = line.strip()
                        line_out.append(line)
                        if "Experiment number:" in line:
                            line_exp.append(line_ct+4)
                        line_ct = line_ct + 1

            
            # Evaluate all tests in the characterization set
            tests = list(self.Compare['Characterization'].keys())
            self.Compare['Prediction'] = dict.fromkeys(self.Compare['Data'])
            ct = 1
            for test in tests:
                # Preallocate the predictions for a test
                self.Compare['Prediction'][test] = dict.fromkeys(self.Compare['Data'][test])
                self.Compare['Prediction'][test]['Strain'] = dict.fromkeys(self.Compare['Data'][test]['Strain'])
                self.Compare['Prediction'][test]['Stress'] = dict.fromkeys(self.Compare['Data'][test]['Stress'])

                # Read the plot file
                data_plot = {'Time':[],
                            'Strain-11':[],
                            'Stress-11':[],
                            'Strain-22':[],
                            'Stress-22':[],
                            'Strain-12':[],
                            'Stress-12':[],}
                
                with open(os.path.join(temp_dir,"u" + str(ct) + ".plot"), "r") as file:
                    for line in file:
                        # Process each line here
                        line_all = line.strip()
                        line_all = line_all.split()
                        i = 0
                        for key in data_plot.keys():
                            data_plot[key].append(float(line_all[i]))
                            i=i+1

                # Populate Time
                self.Compare['Prediction'][test]['Time'] = data_plot['Time']

                # Populate Strain
                keys = list(self.Compare['Prediction'][test]['Strain'].keys())
                for key in keys:
                    self.Compare['Prediction'][test]['Strain'][key] = data_plot['Strain-' + str(key)]
                    
                # Populate Stress
                keys = list(self.Compare['Prediction'][test]['Stress'].keys())
                for key in keys:
                    self.Compare['Prediction'][test]['Stress'][key] = data_plot['Stress-' + str(key)]

                # Calculate Error
                err = float(line_out[line_exp[ct-1]].split(' ')[-1])
                self.Compare['Prediction'][test]['Error'] = err

                # Update ct
                ct = ct + 1

            messagebox.showinfo(message = 'Optimization Complete!')
        except:
            messagebox.showerror(message = 'Optimization Failed')

    def analyze(self):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: Check all conditions before running analysis.
        #
        #--------------------------------------------------------------------------

        # Set error checking flag
        flag = 0

        # Check all elastic values are populated
        if hasattr(self,'sheet1_anly'):
            for i in range(len(self.sheet1_anly.data)):
                try:
                    float(self.sheet1_anly.data[i][3])
                except:
                    flag = 1
                    msg = 'Invalid values for the elastic parameters.'

        # Check all plastic values are populated
        if hasattr(self,'sheet2_anly'):
            for i in range(len(self.sheet2_anly.data)):
                try:
                    float(self.sheet2_anly.data[i][3])
                except:
                    flag = 1
                    msg = 'Invalid values for the plastic parameters.'


        if flag == 0:
            if len(list(self.Compare['Characterization'].keys())) > 0:
                tests = list(self.Compare['Characterization'].keys())
                self.run_compare_anly(tests)
            else:
                messagebox.showeror(message='No tests have been added to the Characterization set.')
        else:
            messagebox.showinfo(message=msg)

    def run_compare_anly(self, tests):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: Run COMPARE Analysis.
        #
        #--------------------------------------------------------------------------

        try:
            # Check for analyze
            if 'Model' in self.Compare.keys():
                data = copy.deepcopy(self.Compare['Model'])
                if 'Analysis' not in self.Compare.keys():                    
                    self.Compare['Analysis'] = data
                # Reformat Parameters
                if 'VE_Param' in list(data.keys()):
                    self.Compare['Analysis']['VE_Param'] = []
                    for i in range(len(data['VE_Param'])):
                        self.Compare['Analysis']['VE_Param'].append([data['VE_Param'][i][0],
                                                                data['VE_Param'][i][1],
                                                                data['VE_Param'][i][6],
                                                                ])
                        
                if 'VP_Param' in list(data.keys()):
                    self.Compare['Analysis']['VP_Param'] = []
                    for i in range(len(data['VP_Param'])):
                        self.Compare['Analysis']['VP_Param'].append([data['VP_Param'][i][0],
                                                                data['VP_Param'][i][1],
                                                                data['VP_Param'][i][6],
                                                                ])


            # Create and clear the Temp Directory
            temp_dir = os.path.join(os.getcwd(),'Temp')
            try:
                os.mkdir(temp_dir)
            except:
                pass
            for filename in os.listdir(temp_dir):
                file_path = os.path.join(temp_dir, filename)
                try:
                    if os.path.isfile(file_path) or os.path.islink(file_path):
                        os.unlink(file_path)
                    elif os.path.isdir(file_path):
                        shutil.rmtree(file_path)
                except Exception as e:
                    print(f"Failed to delete {file_path}. Reason: {e}")

            # Copy the executable
            shutil.copy(self.Compare['Paths']['Compare Executable'], temp_dir)

            # Determine Model Type
            if self.Compare['Analysis']['Model Name'] == 'GVIPS':
                if self.Compare['Analysis']['Reversible Model Name'] == 'Isotropic Viscoelastic' and self.Compare['Analysis']['Irreversible Model Name'] == 'Isotropic Viscoplastic':
                    
                    # Write the DGS file
                    mod, Param, Param_U, Param_N = WriteDSG_GVIPS_ANLY_IN(self, temp_dir, tests)


            # Write the Simulation files
            ct = 1
            sim_tests = tests
            for sim_test in sim_tests:
                WriteSIM(self, self.Compare['Data'][sim_test], temp_dir, ct, mod, Param)
                ct = ct + 1

            # Write the NLP files
            WriteNLP(temp_dir)

            # Run Compare
            command = 'cmd /k "cd ' + temp_dir + ' & compnasardamage & exit"'
            os.system(command)

            # Set model status to 1
            self.Compare['Model']['Status'] = 1

            # Get Test Error
            out_file = os.path.join(temp_dir,"comp.out")
            line_out = []
            line_ct = 1
            line_exp = []
            with open(out_file, "r") as file:
                    for line in file:
                        line = line.strip()
                        line_out.append(line)
                        if "Experiment number:" in line:
                            line_exp.append(line_ct+4)
                        line_ct = line_ct + 1

            
            # Evaluate all tests in the characterization set
            if "Prediction" not in self.Compare.keys():
                self.Compare['Prediction'] = dict.fromkeys(self.Compare['Data'])
            ct = 1
            for test in tests:
                # Preallocate the predictions for a test
                self.Compare['Prediction'][test] = dict.fromkeys(self.Compare['Data'][test])
                self.Compare['Prediction'][test]['Strain'] = dict.fromkeys(self.Compare['Data'][test]['Strain'])
                self.Compare['Prediction'][test]['Stress'] = dict.fromkeys(self.Compare['Data'][test]['Stress'])

                # Read the plot file
                data_plot = {'Time':[],
                            'Strain-11':[],
                            'Stress-11':[],
                            'Strain-22':[],
                            'Stress-22':[],
                            'Strain-12':[],
                            'Stress-12':[],}
                
                with open(os.path.join(temp_dir,"u" + str(ct) + ".plot"), "r") as file:
                    for line in file:
                        # Process each line here
                        line_all = line.strip()
                        line_all = line_all.split()
                        i = 0
                        for key in data_plot.keys():
                            data_plot[key].append(float(line_all[i]))
                            i=i+1

                # Populate Time
                self.Compare['Prediction'][test]['Time'] = data_plot['Time']

                # Populate Strain
                keys = list(self.Compare['Prediction'][test]['Strain'].keys())
                for key in keys:
                    self.Compare['Prediction'][test]['Strain'][key] = data_plot['Strain-' + str(key)]
                    
                # Populate Stress
                keys = list(self.Compare['Prediction'][test]['Stress'].keys())
                for key in keys:
                    self.Compare['Prediction'][test]['Stress'][key] = data_plot['Stress-' + str(key)]

                # Calculate Error
                err = float(line_out[line_exp[ct-1]].split(' ')[-1])
                self.Compare['Prediction'][test]['Error'] = err

            messagebox.showinfo(message = 'Analysis Complete!')
        except:
            messagebox.showerror(message = 'Analysis Failed')

    #------------------------------------------------------------------------------------------------------------------------------------------
    #
    #   VISUALIZATION PAGE
    #   The Visualization page allows users to visualize the model fit to the 
    #   characterization set and to test data not in the characterization set
    #
    #------------------------------------------------------------------------------------------------------------------------------------------

    def viz_tab(self):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: Create the Visualization Page.
        #
        #--------------------------------------------------------------------------

        # Delete any local attributes
        DeleteLocal(self)

        # Create the visualizaton tab
        CreateVisualizationTab(self,window,frmt)

    def plotter_viz(self):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: Plot Visualization curves.
        #
        #--------------------------------------------------------------------------

        #  Delete the canvas and drop down if it exists
        if hasattr(self, 'canvas'):
            self.toolbar.destroy()
            self.canvas.get_tk_widget().destroy()
            del self.canvas

        # Create the plot
        self.fig = Figure(figsize=(5,3.6), dpi = 125, constrained_layout = True)
        self.plot1 = self.fig.add_subplot(111)

        # Get the arrays
        x_val = self.opt1_viz.get()
        y_val = self.opt2_viz.get()

        # X Value
        xp = None
        if 'Time' in x_val:
            x = self.Compare['Data'][self.test_name]['Time'][:self.Compare['Data'][self.test_name]['Stage Index'][-1]]
            xs = self.Compare['Data'][self.test_name]['Reduced Data']['Time']
            if 'Prediction' in list(self.Compare.keys()):
                xp = self.Compare['Prediction'][self.test_name]['Time']
            xu = 'Time [s]'
        else:
            x_val = x_val.split('-')
            x = self.Compare['Data'][self.test_name][x_val[0]][int(x_val[1])][:self.Compare['Data'][self.test_name]['Stage Index'][-1]]
            xs = self.Compare['Data'][self.test_name]['Reduced Data'][x_val[0]][int(x_val[1])]
            if 'Prediction' in list(self.Compare.keys()):
                xp = self.Compare['Prediction'][self.test_name][x_val[0]][int(x_val[1])]
            if x_val[0] == 'Strain':
                xu = 'Strain'
            else:
                xu = 'Stress [MPa]'

         # Y Value
        yp = None
        if 'Time' in y_val:
            y = self.Compare['Data'][self.test_name]['Time'][:self.Compare['Data'][self.test_name]['Stage Index'][-1]]
            ys = self.Compare['Data'][self.test_name]['Reduced Data']['Time']
            if 'Prediction' in list(self.Compare.keys()):
                yp = self.Compare['Prediction'][self.test_name]['Time']
            yu = 'Time [s]'
        else:
            y_val = y_val.split('-')
            y = self.Compare['Data'][self.test_name][y_val[0]][int(y_val[1])][:self.Compare['Data'][self.test_name]['Stage Index'][-1]]
            ys = self.Compare['Data'][self.test_name]['Reduced Data'][y_val[0]][int(y_val[1])]
            if 'Prediction' in list(self.Compare.keys()):
                yp = self.Compare['Prediction'][self.test_name][y_val[0]][int(y_val[1])]
            if y_val[0] == 'Strain':
                yu = 'Strain'
            else:
                yu = 'Stress [MPa]'

        # Plot the data
        self.plot1.plot(x,y,'k',label = 'Raw Data')
        if xs is not None:
            self.plot1.plot(xs,ys,'ko',label = 'Reduced Data')
        if xp is not None:
            self.plot1.plot(xp,yp, color = 'r', marker='o', markerfacecolor='r',label ='Prediction')

        # Set Formatting
        xlab = xu
        ylab = yu
        xlab_frmt = ScalarFormatter() 
        ylab_frmt = ScalarFormatter()

        # Format the plot
        self.plot1.set_xlabel(xlab)
        self.plot1.set_ylabel(ylab)
        self.plot1.xaxis.set_major_formatter(xlab_frmt)
        self.plot1.yaxis.set_major_formatter(ylab_frmt)
        if "Strain" in xlab or "Time" in xlab:
            self.plot1.ticklabel_format(style='sci',scilimits=(-6,-3),axis='x')
        if "Strain" in ylab or "Time" in ylab:
            self.plot1.ticklabel_format(style='sci',scilimits=(-6,-3),axis='y')
        self.plot1.legend()

        # Create the Tkinter canvas
        self.canvas = FigureCanvasTkAgg(self.fig, master = window)
        self.canvas.draw()

        # Create the Matplotlib toolbar
        self.toolbar = NavigationToolbar2Tk(self.canvas, window)
        self.toolbar.update()

        # Format Toolbar
        self.toolbar.config(bg=bg_color)
        self.toolbar._message_label.config(background=bg_color)
        self.toolbar.place(anchor = 'e', relx = 0.975, rely = 0.9)

        # Add the figure to the GUI
        self.canvas.get_tk_widget().place(anchor = 'n', relx = 0.78, rely = 0.32)
        if 'self.canvas' not in self.tab_att_list:
            self.tab_att_list.append('self.canvas')

    #------------------------------------------------------------------------------------------------------------------------------------------
    #
    #   Export Page
    #   The Export page allows users to populate database records and produce
    #   excel output files for the project
    #
    #------------------------------------------------------------------------------------------------------------------------------------------

    def export_tab(self):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: Export data to excel.
        #
        #--------------------------------------------------------------------------

        # Check that a model exists
        if  'Model ID' in self.Compare.keys():
            if self.Compare['Model ID'] != None:
                def export_ex():
                    #--------------------------------------------------------------------------
                    #
                    #   PURPOSE: Create the excel output file.
                    #
                    #--------------------------------------------------------------------------

                    # Get the model information
                    json_string = self.Compare['Model Library'][self.Compare['Model ID']].decode('utf-8')
                    data = json.loads(json_string)

                    # Get the characterization test names
                    char_tests = list(self.Compare['Characterization'].keys())
                    pred_tests = list(self.Compare['Prediction'].keys())

                    # Open the excel template
                    wb = load_workbook(self.Compare['Paths']['Export Template'], data_only=True)

                    def FormatArea(ws, x1, x2, y1, y2):
                        #----------------------------------------------------------
                        #
                        #   PURPOSE: Format an area in the excel output file.
                        #
                        #----------------------------------------------------------

                        # Make all cells white
                        for i in range(x1, x2+1):
                            for j in range(y1, y2+1):
                                ws.cell(row=i, column=j).fill = PatternFill(fill_type='solid', start_color='FFFFFF', end_color='FFFFFF')

                        # Create the borders
                        border_thin = Side(style='thin')
                        border_thick = Side(style='thick')

                        # Format Border
                        for i in range(x1, x2+1):
                            ws.cell(row = i, column = y1).border = Border(top=border_thin, left=border_thick, right=border_thin, bottom=border_thin)
                            ws.cell(row = i, column = y2).border = Border(top=border_thin, left=border_thin, right=border_thick, bottom=border_thin)

                        for i in range(y1, y2+1):
                            ws.cell(row = x1, column = i).border = Border(top=border_thick, left=border_thin, right=border_thin, bottom=border_thin)
                            ws.cell(row = x2, column = i).border = Border(top=border_thin, left=border_thin, right=border_thin, bottom=border_thick)

                        # Format Interior
                        for i in range(x1+1,x2):
                            for j in range(y1+1,y2):
                                ws.cell(row = i, column = j).border = Border(top=border_thin, left=border_thin, right=border_thin, bottom=border_thin)

                        # Format Corners
                        ws.cell(row = x1, column = y1).border = Border(top=border_thick, left=border_thick, right=border_thin, bottom=border_thin)
                        ws.cell(row = x2, column = y1).border = Border(top=border_thin, left=border_thick, right=border_thin, bottom=border_thick)
                        ws.cell(row = x1, column = y2).border = Border(top=border_thick, left=border_thin, right=border_thick, bottom=border_thin)
                        ws.cell(row = x2, column = y2).border = Border(top=border_thin, left=border_thin, right=border_thick, bottom=border_thick)

                    
                    # Populate the General Model Sheet
                    # -- General Information
                    ws = wb['Model Information']
                    ws.cell(row = 3, column = 3).value = self.Compare['Model ID']
                    ws.cell(row = 4, column = 3).value = data['Model Name']
                    ws.cell(row = 5, column = 3).value = self.Compare['Characterization'][char_tests[0]]['Temperature'][0]
                    ws.cell(row = 6, column = 3).value = data['Compare Type']
                    if 'Note' in data.keys():
                        ws.cell(row = 7, column = 3).value = data['Note']
                    FormatArea(ws, 2, 7, 2, 8)

                    # -- Reversible Model
                    if 'Reversible Model Name' in data.keys():
                        ws.cell(row = 10, column = 3).value = data['Reversible Model Name']
                    if 'M' in data.keys():
                        ws.cell(row = 11, column = 3).value = int(data['M'])
                    ct = 0
                    if 'VE_Param' in data.keys():
                        ct = len(data['VE_Param'])
                        for i in range(len(data['VE_Param'])):
                            ws.cell(row=14+i,column=2).value = data['VE_Param'][i][0]
                            ws.cell(row=14+i,column=3).value = data['VE_Param'][i][1]
                            if data['Compare Type'] == 'Optimize':
                                ws.cell(row=14+i,column=4).value = data['VE_Param'][i][6]
                            else:
                                ws.cell(row=14+i,column=4).value = data['VE_Param'][i][2]
                    FormatArea(ws, 9, 13+ ct, 2, 4)

                    # -- Irreversible Model
                    if 'Irreversible Model Name' in data.keys():
                        ws.cell(row = 10, column = 7).value = data['Irreversible Model Name']
                    if 'N' in data.keys():
                        ws.cell(row = 11, column = 7).value = int(data['N'])
                    ct = 0
                    if 'VP_Param' in data.keys():
                        ct = len(data['VP_Param'])
                        for i in range(len(data['VP_Param'])):
                            ws.cell(row=14+i,column=6).value = data['VP_Param'][i][0]
                            ws.cell(row=14+i,column=7).value = data['VP_Param'][i][1]
                            if data['Compare Type'] == 'Optimize':
                                ws.cell(row=14+i,column=8).value = data['VP_Param'][i][6]
                            else:
                                ws.cell(row=14+i,column=8).value = data['VP_Param'][i][2]
                    FormatArea(ws, 9, 13+ ct, 6, 8)

                    # Test information
                    ws = wb['Test Information']

                    # Write Characaterization Test Data
                    for i in range(len(char_tests)):
                        ws.cell(row = 3+i, column = 2).value = char_tests[i]
                        ws.cell(row = 3+i, column = 3).value = self.Compare['Characterization'][char_tests[i]]['Test Type']
                        ws.cell(row = 3+i, column = 4).value = self.Compare['Characterization'][char_tests[i]]['Loading Direction'][0]
                        ws.cell(row = 3+i, column = 5).value = self.Compare['Characterization'][char_tests[i]]['Control'][0]
                        ws.cell(row = 3+i, column = 6).value = self.Compare['Characterization'][char_tests[i]]['Load Rate'][0][0]
                        ws.cell(row = 3+i, column = 7).value = self.Compare['Characterization'][char_tests[i]]['Load Rate'][0][1]
                        ws.cell(row = 3+i, column = 8).value = self.Compare['Characterization'][char_tests[i]]['Angle']
                        ws.cell(row = 3+i, column = 9).value = 'Characterization'
                        ws.cell(row = 3+i, column = 10).value = self.Compare['Characterization'][char_tests[i]]['RelWeight']
                        ws.cell(row = 3+i, column = 11).value = self.Compare['Prediction'][char_tests[i]]['Error']
                    ct = len(char_tests)

                    # Write Verification Test Data
                    for i in range(len(pred_tests)):
                        if self.Compare['Prediction'][pred_tests[i]] is not None:
                            if pred_tests[i] not in char_tests:
                                if self.Compare['Prediction'][pred_tests[i]]['Error'] is not None:
                                    ws.cell(row = 3+ct, column = 2).value = pred_tests[i]
                                    ws.cell(row = 3+ct, column = 3).value = self.Compare['Data'][pred_tests[i]]['Test Type']
                                    ws.cell(row = 3+ct, column = 4).value = self.Compare['Data'][pred_tests[i]]['Loading Direction'][0]
                                    ws.cell(row = 3+ct, column = 5).value = self.Compare['Data'][pred_tests[i]]['Control'][0]
                                    ws.cell(row = 3+ct, column = 6).value = self.Compare['Data'][pred_tests[i]]['Load Rate'][0][0]
                                    ws.cell(row = 3+ct, column = 7).value = self.Compare['Data'][pred_tests[i]]['Load Rate'][0][1]
                                    ws.cell(row = 3+ct, column = 8).value = self.Compare['Data'][pred_tests[i]]['Angle']
                                    ws.cell(row = 3+ct, column = 9).value = 'Verification'
                                    ws.cell(row = 3+ct, column = 11).value = self.Compare['Prediction'][pred_tests[i]]['Error']
                                    ct = ct + 1

                    FormatArea(ws, 2, 2+ct, 2, 11)

                    # Response Curves
                    ws = wb['Response Curves']
                    max_x = 2

                    # Write response curves
                    start_col = 2
                    for i in range(len(pred_tests)):
                        start_col_test = start_col
                        if self.Compare['Prediction'][pred_tests[i]] is not None:
                            # -- Write the test name
                            ws.cell(row = 2, column=start_col).value = pred_tests[i]

                            # -- Write Time
                            if 'Time' in self.Compare['Prediction'][pred_tests[i]].keys():
                                ws.cell(row = 3, column=start_col).value = 'Time (s)'
                                for j in range(len(self.Compare['Prediction'][pred_tests[i]]['Time'])):
                                    ws.cell(row = 4+j, column=start_col).value = self.Compare['Prediction'][pred_tests[i]]['Time'][j]
                                if j > max_x:
                                    max_x = j+4
                                start_col = start_col + 1

                            # -- Write Strain
                            if 'Strain' in self.Compare['Prediction'][pred_tests[i]].keys():
                                keys = list(self.Compare['Prediction'][pred_tests[i]]['Strain'].keys())
                                for key in keys:
                                    ws.cell(row = 3, column=start_col).value = 'Strain-' + str(key) + '(-)'
                                    for j in range(len(self.Compare['Prediction'][pred_tests[i]]['Strain'][key])):
                                        ws.cell(row = 4+j, column=start_col).value = self.Compare['Prediction'][pred_tests[i]]['Strain'][key][j]
                                    if j > max_x:
                                        max_x = j+4
                                    start_col = start_col + 1

                            # -- Write Stress
                            if 'Stress' in self.Compare['Prediction'][pred_tests[i]].keys():
                                keys = list(self.Compare['Prediction'][pred_tests[i]]['Stress'].keys())
                                for key in keys:
                                    ws.cell(row = 3, column=start_col).value = 'Stress-' + str(key) + '(MPa)'
                                    for j in range(len(self.Compare['Prediction'][pred_tests[i]]['Stress'][key])):
                                        ws.cell(row = 4+j, column=start_col).value = self.Compare['Prediction'][pred_tests[i]]['Stress'][key][j]
                                    if j > max_x:
                                        max_x = j+4
                                    start_col = start_col + 1

                            # -- Merge Test Cells
                            ws.merge_cells(start_row=2, start_column=start_col_test, end_row=2, end_column=start_col-1)

                    FormatArea(ws, 2, max_x, 2, start_col-1)

                    # Ask for the save name from the user
                    file_path = filedialog.asksaveasfilename(
                                    defaultextension=".xlsx",
                                    filetypes=[("Excel Files", "*.xlsx")],
                                    title="Save As",
                                    confirmoverwrite=True
                                )
                    
                    if file_path:
                        try:
                            wb.save(file_path)
                            messagebox.showinfo(message = 'Export File Saved!')
                        except:
                            messagebox.showerror(message = "An error occurred. Please check that the file is closed.")

                # Create the excel export
                export_ex()

            else:
                    messagebox.showinfo(message = 'No Model Selected.')
        else:
            messagebox.showinfo(message = 'No Model Selected.')

    #------------------------------------------------------------------------------------------------------------------------------------------
    #
    #   Settings Page
    #   The Settings page allows users alter settings regarding their instance of PYMI COMPARE
    #
    #------------------------------------------------------------------------------------------------------------------------------------------

    def settings_tab(self):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: Create the settings window.
        #
        #--------------------------------------------------------------------------

        # Set Paths Button
        def set_paths(self):
            #----------------------------------------------------------------------
            #
            #   PURPOSE: Create the settings window.
            #
            #----------------------------------------------------------------------

            # Set flag to check for open setings tab
            if hasattr(self,"settings_window") == False:
                self.settings_window = 0
            if self.settings_window == 0:
                self.settings_window = 1

                # Create the window
                root = tk.Tk()
                root.config(width=900, height=700)
                root.title("Path Dependencies")

                # Reset Window
                def reset_window():
                    #--------------------------------------------------------------
                    #
                    #   PURPOSE: Reset the path window.
                    #
                    #--------------------------------------------------------------

                    # Define attribute lst
                    atts =  ["self.comp_path", "self.mod_path", "self.imp_path", "self.exp_path",
                            "self.comp_path_btn", "self.mod_path_btn", "self.imp_path_btn", "self.exp_path_btn",
                            "self.exp_dwnld_btn", "self.imp_dwnld_btn"]
                    
                    # Destory attributes
                    for att in atts:
                        try:
                            eval(atts).destroy()
                        except:
                            pass

                    # Set the paths
                    # -- Compare Executable
                    def set_comp_path(self):
                        #----------------------------------------------------------
                        #
                        #   PURPOSE: Set the compare path file.
                        #
                        #----------------------------------------------------------

                        # Ask for the new file path
                        file = filedialog.askopenfile(title = "Compare Executable", filetypes= [('Executable', '*.exe')], mode ='r',)
                        
                        # Set new path and reset window
                        if file is not None:
                            self.Compare['Paths']['Compare Executable'] = file.name
                            self.comp_path.configure(text='')
                            reset_window()

                    # Create label
                    self.comp_path = tk.Label(root, text="Compare Executable Path: " + self.Compare['Paths']['Compare Executable'] , 
                                                font=(fontname, 10))
                    self.comp_path.place(anchor='w', relx = 0.025, rely = 0.1)

                    # Create button to edit path
                    self.comp_path_btn = tk.Button(root, text = "Edit", command = lambda:set_comp_path(self), 
                                                font = (fontname, 10), bg = '#fc3d21', fg='white',
                                                width = 10)
                    self.comp_path_btn.place(anchor = 'w', relx = 0.025, rely = 0.15)

                    
                    # -- Model Library
                    def set_mod_path(self):
                        #----------------------------------------------------------
                        #
                        #   PURPOSE: Set the model library path file.
                        #
                        #----------------------------------------------------------

                        # Ask for the new file path
                        file = filedialog.askopenfile(title = "Available Models", filetypes= [('Excel', '*.xlsx')], mode ='r',)
                        
                        # Set new path and reset window
                        if file is not None:
                            self.Compare['Paths']['Model Library'] = file.name
                            self.mod_path.configure(text='')
                            reset_window()

                    # Create the label
                    self.mod_path = tk.Label(root, text="Available Models: " + self.Compare['Paths']['Model Library'] , 
                                                font=(fontname, 10))
                    self.mod_path.place(anchor='w', relx = 0.025, rely = 0.2)

                    # Create Button to edit the path
                    self.mod_path_btn = tk.Button(root, text = "Edit", command = lambda:set_mod_path(self), 
                                                font = (fontname, 10), bg = '#fc3d21', fg='white',
                                                width = 10)
                    self.mod_path_btn.place(anchor = 'w', relx = 0.025, rely = 0.25)
                    
                    # -- Excel Import Template
                    def set_imp_path(self):
                        #----------------------------------------------------------
                        #
                        #   PURPOSE: Set the excel import template file path file.
                        #
                        #----------------------------------------------------------

                        # Ask for the new file path
                        file = filedialog.askopenfile(title = "Import Template", filetypes= [('Excel', '*.xlsx')], mode ='r',)
                        
                        # Set new path and reset window
                        if file is not None:
                            self.Compare['Paths']['Import Template'] = file.name
                            self.imp_path.configure(text='')
                            reset_window()


                    def download_imp(self):
                        #----------------------------------------------------------
                        #
                        #   PURPOSE: Download the import template file.
                        #
                        #----------------------------------------------------------

                        # Ask where to save the file
                        try:
                            file = filedialog.asksaveasfile(title = "Import Template", filetypes=[('Excel', '*.xlsx')],
                                                        initialdir = str(Path.home() / "Downloads"),
                                                        initialfile="ImportTemplate.xlsx")
                        except:
                            file = filedialog.asksaveasfile(title = "Import Template", filetypes=[('Excel', '*.xlsx')],
                                                            initialfile="ImportTemplate.xlsx")

                        # Copy the file
                        if file is not None:
                            shutil.copy(self.Compare['Paths']['Import Template'], file.name)

                    # Create the lable
                    self.imp_path = tk.Label(root, text="Excel Import Template: " + self.Compare['Paths']['Import Template'] , 
                                                font=(fontname, 10))
                    self.imp_path.place(anchor='w', relx = 0.025, rely = 0.3)

                    # Create button to edit the path
                    self.imp_path_btn = tk.Button(root, text = "Edit", command = lambda:set_imp_path(self), 
                                                font = (fontname, 10), bg = '#fc3d21', fg='white',
                                                width = 10)
                    self.imp_path_btn.place(anchor = 'w', relx = 0.025, rely = 0.35)

                    # Create button to download file
                    self.imp_dwnld_btn = tk.Button(root, text = "Download", command = lambda:download_imp(self), 
                                                font = (fontname, 10), bg = '#fc3d21', fg='white',
                                                width = 10)
                    self.imp_dwnld_btn.place(anchor = 'w', relx = 0.125, rely = 0.35)
                    
                    # -- Excel Export Template
                    def set_exp_path(self):
                        #----------------------------------------------------------
                        #
                        #   PURPOSE: Set the export template path file.
                        #
                        #----------------------------------------------------------

                        # Ask for the new file path
                        file = filedialog.askopenfile(title = "Export Template", filetypes= [('Excel', '*.xlsx')], mode ='r',)
                        
                        # Set new path and reset window
                        if file is not None:
                            self.Compare['Paths']['Export Template'] = file.name
                            self.exp_path.configure(text='')
                            reset_window()

                    def download_exp(self):
                        #----------------------------------------------------------
                        #
                        #   PURPOSE: Download the export template.
                        #
                        #----------------------------------------------------------

                        # Ask for the new file path
                        try:
                            file = filedialog.asksaveasfile(title = "Export Template", filetypes=[('Excel', '*.xlsx')],
                                                        initialdir = str(Path.home() / "Downloads"),
                                                        initialfile="ExportTemplate.xlsx")
                        except:
                            file = filedialog.asksaveasfile(title = "Export Template", filetypes=[('Excel', '*.xlsx')],
                                                            initialfile="ExportTemplate.xlsx")

                        # Copy file
                        if file is not None:
                            shutil.copy(self.Compare['Paths']['Export Template'], file.name)

                    # Create the label
                    self.exp_path = tk.Label(root, text="Excel Export Template: " + self.Compare['Paths']['Export Template'] , 
                                                font=(fontname, 10))
                    self.exp_path.place(anchor='w', relx = 0.025, rely = 0.4)

                    # Create button to edit the path
                    self.exp_path_btn = tk.Button(root, text = "Edit", command = lambda:set_exp_path(self), 
                                                font = (fontname, 10), bg = '#fc3d21', fg='white',
                                                width = 10)
                    self.exp_path_btn.place(anchor = 'w', relx = 0.025, rely = 0.45)

                    # Create button to download the file
                    self.exp_dwnld_btn = tk.Button(root, text = "Download", command = lambda:download_exp(self), 
                                                font = (fontname, 10), bg = '#fc3d21', fg='white',
                                                width = 10)
                    self.exp_dwnld_btn.place(anchor = 'w', relx = 0.125, rely = 0.45)
                
                # Create the window
                reset_window()

                # Create Exit Protocol
                def on_closing_root(self):
                    #--------------------------------------------------------------
                    #
                    #   PURPOSE: Set exit protocol for the setting window.
                    #
                    #--------------------------------------------------------------

                    # Reset clicked and destory the window
                    self.settings_window = 0
                    root.destroy()

                # Add the exit protocol to the root
                root.protocol("WM_DELETE_WINDOW", lambda:on_closing_root(self))

        # Call Path Selection
        set_paths(self)

# Run the GUI
PY_COMPARE()