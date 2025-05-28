#--------------------------------------------------------------------------------------------------------------------------------------------------------------------
#
# ReadExcelInput.py
#
# PURPOSE: Read an Excel Input File and perform error checking.
#
# INPUTS
#   file      Excel model file
#   self      GUI data structure
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------

def ReadModel(file, self):
    # Import Modules
    import pandas as pd 
    from openpyxl import load_workbook
 
    # Read the file
    df = pd.read_excel(file, sheet_name = 'Model Information')

    # Set the error flag and message
    flag  = 0
    msg = ''
    model = {}

    # Check for the model name 
    if pd.isna(df.values[2][2]):
        flag = 1
        msg = msg + '   ERROR: Model Type Not Defined'
    else:
        if df.values[2][2] not in self.Models:
            flag = 1
            msg = msg + '   ERROR: Model Type not found in available models.'
        else:
            model['Model Name'] = df.values[2][2]
            model_info = load_workbook(self.Compare['Paths']['Model Library'], data_only=True)
            ws = model_info[df.values[2][2]]
            model['Model Info'] = {}
            for i in range(1,ws.max_row+1):
                model['Model Info'][ws.cell(row=i,column=1).value] = []
                j = 2
                if "Units" not in ws.cell(row=i,column=1).value:
                    while ws.cell(row=i,column=j).value != None:
                        model['Model Info'][ws.cell(row=i,column=1).value].append(ws.cell(row=i,column=j).value)
                        j= j+1
                else:
                    for j in range(2,len(model['Model Info'][ws.cell(row=i-1,column=1).value])+2):
                        model['Model Info'][ws.cell(row=i,column=1).value].append(ws.cell(row=i,column=j).value)

    # Get the reversible model name
    if flag == 0:
        if pd.isna(df.values[8][2]):
            flag = 1
            msg = msg + '   ERROR: Reversible Model Name Not Defined'
        else:
            if df.values[8][2] in model['Model Info']['Reversible Models']:
                model['Reversible Model Name'] = df.values[8][2]
            else:
                flag = 1
                msg = msg + '   ERROR: Reversible Model Name not in available reversible models'

    # Get the irreversible model name
    if flag == 0:
        if pd.isna(df.values[8][6]):
            flag = 0
            msg = msg + '   WARNING: Irreversible Model Name Not Defined'
        else:
            if df.values[8][6] in model['Model Info']['Irreversible Models']:
                model['Irreversible Model Name'] = df.values[8][6]
            else:
                flag = 1
                msg = msg + '   ERROR: Irreversible Model Name not in available irreversible models'

    # Get the reversible mechanisms
    if flag == 0:
        if pd.isna(df.values[9][2]):
            flag = 1
            msg = msg + '   ERROR: Reversible Mechanisms Name Not Defined'
        else:
            if int(df.values[9][2]) in model['Model Info']['Reversible Mechanisms']:
                model['M'] = df.values[9][2]
            else:
                flag = 1
                msg = msg + '   ERROR: Reversible Mechanisms not in available reversible mechanisms'
            
    # Get the irreversible mechanisms
    if flag == 0:
        if pd.isna(df.values[9][6]):
            flag = 1
            msg = msg + '   ERROR: Ireversible Mechanisms Name Not Defined'
        else:
            if int(df.values[9][6]) in model['Model Info']['Irreversible Mechanisms']:
                model['N'] = df.values[9][6]
            else:
                flag = 1
                msg = msg + '   ERROR: Irreversible Mechanisms not in available irreversible mechanisms'

    # Get the reversible parameters
    if flag == 0:
        row = 12
        VE = []
        while pd.isna(df.values[row][1]) == False:
            VE.append([df.values[row][1], df.values[row][2], df.values[row][3]]) 
            row = row + 1
            if row == len(df.values):
                break

        for i in range(len(VE)):
            # Check the variable
            if VE[i][0][0] not in model['Model Info']['Reversible Deformation Parameters'] and VE[i][0][0]+'_[M]' not in model['Model Info']['Reversible Deformation Parameters']:
                flag = 1
                msg = msg + "   ERROR: Incorrect Reversible Parameter names found."

        if flag == 0:
            model['VE_Param'] = VE

    # Get the irreversible parameters
    if flag == 0:
        row = 12
        VP = []
        while pd.isna(df.values[row][5]) == False:
            VP.append([df.values[row][5], df.values[row][6], df.values[row][7]]) 
            row = row + 1
            if row == len(df.values):
                break

        for i in range(len(VP)):
            # Check the variable
            if VP[i][0][0] not in  model['Model Info']['Irreversible Deformation Parameters'] and VP[i][0][0]+'_[N]' not in model['Model Info']['Irreversible Deformation Parameters']:
                flag = 1
                msg = msg + "   ERROR: Incorrect Irreversible Parameter names found."

        if flag == 0:
            model['VP_Param'] = VP

    # Set the model type
    model['Compare Type'] = 'Analysis'

    return model, flag, msg