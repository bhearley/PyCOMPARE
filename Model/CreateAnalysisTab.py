#--------------------------------------------------------------------------------------------------------------------------------------------------------------------
#
# CreateAnalysisTab.py
#
# PURPOSE: Create the Analyze Model tab. The Analyze Model tab allows users to define a model manually and evaluate.
#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------
def CreateAnalysisTab(self,window):
    # Import Modules
    import copy
    import json
    from openpyxl import load_workbook
    import tkinter as tk
    from tkinter import messagebox
    from tkinter import simpledialog
    from tkinter import ttk 
    from tkinter import scrolledtext 
    import tksheet

    # Import Functions
    from Model.UpdateModelData import UpdateModelData

    # Initialize Model
    if 'Analysis' not in self.Compare.keys():
        self.Compare['Analysis'] = {}
        
    if 'Model ID' not in self.Compare.keys() == False:
        self.Compare['Model ID'] = None

    # Get available model information
    if hasattr(self,"model_info_all") == False:
        self.model_info_all = load_workbook(self.Compare['Paths']['Model Library'], data_only=True)
    
    self.clicked = 0

    # Preallocate Saved Models
    if 'Model Library' not in self.Compare.keys():
        self.Compare['Model Library'] = {}

    # Define Available Models
    self.Models = self.model_info_all.sheetnames

    # Deselect Function
    def on_click(event):
        #--------------------------------------------------------------------------
        #
        #   PURPOSE: Deselect from sheets not currently pressed in
        #
        #--------------------------------------------------------------------------

        widget = event.widget

        # If the click is *not* inside the sheet, deselect it
        try:
            if widget != self.sheet1_analy.MT:
                self.sheet1_analy.deselect("all")
        except:
            pass

        try:
            if widget != self.sheet2_analy.MT:
                self.sheet2_analy.deselect("all")
        except:
            pass

    # Bind the deselect function to the window
    window.bind_all("<Button-1>", on_click, add="+")

    def change_model(value):
        #----------------------------------------------------------------------
        #
        #   PURPOSE: Recreate page based on model choice.
        #
        #----------------------------------------------------------------------

        # Clear the model information if the model type changed
        try:
            if value != self.Compare['Analysis']['Model Name']:
                # Prompt user to save model
                self.save_model(None)

                # Clear model information
                self.Compare['Analysis'] = {}
        except:
            pass
        
        # Delete local attributes
        for att in self.atts['Analysis']['Local']:
            try:
                eval(f"{att}").destroy()
            except:
                pass

        # Read the model info
        ws = self.model_info_all[value]
        self.Compare['Analysis']['Model Name'] = value
        self.Compare['Analysis']['Model Info'] = {}
        for i in range(1,ws.max_row+1):
            self.Compare['Analysis']['Model Info'][ws.cell(row=i,column=1).value] = []
            j = 2
            if "Units" not in ws.cell(row=i,column=1).value:
                while ws.cell(row=i,column=j).value != None:
                    self.Compare['Analysis']['Model Info'][ws.cell(row=i,column=1).value].append(ws.cell(row=i,column=j).value)
                    j= j+1
            else:
                for j in range(2,len(self.Compare['Analysis']['Model Info'][ws.cell(row=i-1,column=1).value])+2):
                    self.Compare['Analysis']['Model Info'][ws.cell(row=i,column=1).value].append(ws.cell(row=i,column=j).value)
            
        # Get available reversible models
        self.RevModels = self.Compare['Analysis']['Model Info']['Reversible Models']
        
        if len(self.RevModels) > 0:
            # Create the label
            self.desc2_analy = ttk.Label(
                                self.nb_tab_tab4, 
                                text="Reversible Model:", 
                                anchor=tk.NW,       
                                style = 'Modern1.TLabel'                    
                                )
            self.desc2_analy.place(
                            anchor = 'n', 
                            relx = self.Placement['Analysis']['LabelRev'][0], 
                            rely = self.Placement['Analysis']['LabelRev'][1],
                            )
            
            if 'self.desc2_analy' not in self.atts['Analysis']['Local']:
                self.atts['Analysis']['Local'].append('self.desc2_analy') 

            # Initialize the model
            rmod_opt = self.RevModels[0]

            # Check if previous data exists
            if 'Analysis' in list(self.Compare.keys()):
                # Set the reversible model type
                if 'Reversible Model Name' in list(self.Compare['Analysis'].keys()):
                    if self.Compare['Analysis']['Reversible Model Name'] in self.RevModels:
                        rmod_opt = self.Compare['Analysis']['Reversible Model Name']

            # Create the reversible model drop down   
            self.optmenu2_analy = ttk.Combobox(
                                        self.nb_tab_tab4,
                                        values=self.RevModels,
                                        style="Modern.TCombobox",
                                        state="readonly"
                                        )
            self.optmenu2_analy.configure(font = self.style_man['Combo'])
            self.optmenu2_analy.option_add('*TCombobox*Listbox.font', self.style_man['Combo'])
            self.optmenu2_analy.place(
                                anchor='n', 
                                relx = self.Placement['Analysis']['ComboRev'][0], 
                                rely = self.Placement['Analysis']['ComboRev'][1],
                                relwidth = self.Placement['Analysis']['ComboRev'][2], 
                                relheight = self.Placement['Analysis']['ComboRev'][3]
                                )
            self.optmenu2_analy.set(rmod_opt)
            self.optmenu2_analy.bind("<<ComboboxSelected>>",  lambda event:UpdateModelData(event, self, 1, 'Model'))
            if 'self.optmenu2_analy' not in self.atts['Analysis']['Local']:
                self.atts['Analysis']['Local'].append('self.optmenu2_analy') 

            # Initialize Parameter List
            self.Params_VE = self.Compare['Analysis']['Model Info']['Reversible Deformation Parameters'] + self.Compare['Analysis']['Model Info']['Reversible Damage Parameters']
            self.Params_VE_Units = self.Compare['Analysis']['Model Info']['Reversible Deformation Parameter Units'] + self.Compare['Analysis']['Model Info']['Reversible Damage Parameter Units']

        # Get available irreversible models
        self.IrrevModels = self.Compare['Analysis']['Model Info']['Irreversible Models']

        if len(self.IrrevModels) > 0:
            # Create the label
            self.desc3_analy = ttk.Label(
                                self.nb_tab_tab4, 
                                text= "Irreversible Model:", 
                                anchor=tk.NW,       
                                style = "Modern1.TLabel"
                                )
            self.desc3_analy.place(
                            anchor = 'n', 
                            relx = self.Placement['Analysis']['LabelIrrev'][0], 
                            rely = self.Placement['Analysis']['LabelIrrev'][1]
                            )
            if 'self.desc3_analy' not in self.atts['Analysis']['Local']:
                self.atts['Analysis']['Local'].append('self.desc3_analy') 

            # Initialize the irreversible model
            irmod_opt = self.IrrevModels[0]

            # Check if previous data exists
            if 'Analysis' in list(self.Compare.keys()):
                # Set the reversible model type
                if 'Irreversible Model Name' in list(self.Compare['Analysis'].keys()):
                    if self.Compare['Analysis']['Irreversible Model Name'] in self.IrrevModels:
                        irmod_opt = self.Compare['Analysis']['Irreversible Model Name']

            # Create the irreversible model drop down
            self.optmenu3_analy = ttk.Combobox(
                                        self.nb_tab_tab4,
                                        values=self.IrrevModels,
                                        style="Modern.TCombobox",
                                        state="readonly"
                                        )
            self.optmenu3_analy.configure(font = self.style_man['Combo'])
            self.optmenu3_analy.option_add('*TCombobox*Listbox.font', self.style_man['Combo'])
            self.optmenu3_analy.place(
                                anchor='n', 
                                relx = self.Placement['Analysis']['ComboIrrev'][0], 
                                rely = self.Placement['Analysis']['ComboIrrev'][1],
                                relwidth = self.Placement['Analysis']['ComboIrrev'][2], 
                                relheight = self.Placement['Analysis']['ComboIrrev'][3]
                                )
            self.optmenu3_analy.set(irmod_opt)
            self.optmenu3_analy.bind("<<ComboboxSelected>>",  lambda event:UpdateModelData(event, self, 2, 'Model'))
            if 'self.optmenu3_analy' not in self.atts['Optimize']['Local']:
                self.atts['Optimize']['Local'].append('self.optmenu3_analy') 

            # Initialize Parameter List
            self.Params_VP = self.Compare['Analysis']['Model Info']['Irreversible Deformation Parameters'] + self.Compare['Analysis']['Model Info']['Irreversible Damage Parameters']
            self.Params_VP_Units = self.Compare['Analysis']['Model Info']['Irreversible Deformation Parameter Units'] + self.Compare['Analysis']['Model Info']['Irreversible Damage Parameter Units']

        def update_reversible_table(self):
            #------------------------------------------------------------------
            #
            #   PURPOSE: Update the reversible model parameters table.
            #
            #------------------------------------------------------------------

            # Delete table if it exists
            if hasattr(self,"sheet1_analy"):
                if hasattr(self,"res_flag1") == True:
                    if self.res_flag1 == 0:
                        # Store data
                        self.sheet1_analy_data = self.sheet1_analy.data
                else:
                    # Store data
                    self.sheet1_analy_data = self.sheet1_analy.data

                # Delete sheet
                self.sheet1_analy.destroy()
                del self.sheet1_analy

            # Check if previous data exists
            if 'Analysis' in list(self.Compare.keys()):
                # Set the reversible model type
                if 'VE_Param' in list(self.Compare['Analysis'].keys()):
                    if hasattr(self,"res_flag1") == True:
                        if self.res_flag1 == 0:
                            self.sheet1_analy_data = self.Compare['Analysis']['VE_Param']
                        else:
                            self.res_flag1 = 0
                    else:
                        self.sheet1_analy_data = self.Compare['Analysis']['VE_Param']

            if hasattr(self,"sheet1_analy_data") == False:
                self.sheet1_analy_data = []

            # Set the columns
            Cols = ['Parameter', 'Units', 'Value']

            # Create the table
            self.sheet1_analy = tksheet.Sheet(
                                            self.nb_tab_tab4, 
                                            total_rows = len(self.Params_VE), 
                                            total_columns = len(Cols), 
                                            headers = Cols,
                                            show_x_scrollbar = False, 
                                            show_y_scrollbar = True,
                                            font = ("Segoe UI", max([self.min_font, int(12*self.scale)]),"normal"),
                                            header_font = ("Segoe UI", max([self.min_font, int(12*self.scale)]),"bold"),
                                            )
            self.sheet1_analy.place(
                            anchor = 'n', 
                            relx = self.Placement['Analysis']['Sheet1'][0], 
                            rely = self.Placement['Analysis']['Sheet1'][1],
                            relwidth = self.Placement['Analysis']['Sheet1'][2], 
                            relheight = self.Placement['Analysis']['Sheet1'][3], 
                            )
            if 'self.sheet1_analy' not in self.atts['Analysis']['Local']:
                self.atts['Analysis']['Local'].append('self.sheet1_analy') 

            self.sheet1_analy.change_theme("blue")
            self.sheet1_analy.set_index_width(0)

            def sort_cols(self):
                #----------------------------------------------------------------------
                #
                #   PURPOSE: Custom soring function.
                #
                #----------------------------------------------------------------------

                # Get the currently selected element
                currently_selected = self.sheet1_analy.get_currently_selected()
                
                # Get the list of values
                sort_list = []
                for i in range(self.sheet1_analy.visible_rows[1]):
                    sort_list.append(self.sheet1_analy.data[i][currently_selected.column])
                index_list = sorted(range(len(sort_list)), key=lambda k: sort_list[k])
                
                # Rewrite the table
                temp_data = copy.deepcopy(self.sheet1_analy.data)
                for i in range(self.sheet1_analy.visible_rows[1]):
                    for j in range(self.sheet1_analy.visible_columns[1]):
                        self.sheet1_analy.set_cell_data(i,j,temp_data[index_list[i]][j])
                self.sheet1_analy.redraw()

            # Enable Bindings
            self.sheet1_analy.enable_bindings('single_select','cell_select', 'column_select', 'edit_cell',"arrowkeys", "right_click_popup_menu")
            self.sheet1_analy.popup_menu_add_command('Sort', lambda : sort_cols(self), table_menu = True, index_menu = True, header_menu = True)
            self.sheet1_analy.extra_bindings([("cell_select", lambda event: self.cell_select_anly(event, 'sheet1_analy'))])

            # Set Column Widths
            window.update_idletasks()
            total_width = self.sheet1_analy.winfo_width()
            self.sheet1_analy.column_width(column = 0, width = int(total_width*self.Placement['Analysis']['Sheet1'][4]), redraw = True)
            self.sheet1_analy.column_width(column = 1, width = int(total_width*self.Placement['Analysis']['Sheet1'][5]), redraw = True)
            self.sheet1_analy.column_width(column = 2, width = int(total_width*self.Placement['Analysis']['Sheet1'][6]), redraw = True)
            self.sheet1_analy.table_align(align = 'c',redraw=True)

            # Set unit dictionary
            Units = {'Stress':['GPa','MPa','kPa','Pa','msi','ksi','psi'],
                    'Time':['s'],
                    'Time-1':['1/s']
                    }

            # Set Rows
            for i in range(len(self.Params_VE)):
                self.sheet1_analy.set_cell_data(i,0, self.Params_VE[i])
                if self.Params_VE_Units[i] != None:
                    for key in list(Units.keys()):
                        if self.Params_VE_Units[i] in Units[key]:
                            units_list = Units[key]
                else:
                    units_list = []
                def_val = self.Params_VE_Units[i]
                    
                self.sheet1_analy.create_dropdown(r=i, c = 1,values=units_list)
                if def_val != None:
                    self.sheet1_analy.set_cell_data(i,1, def_val)

            # Add Existing Data
            for i in range(len(self.sheet1_analy_data)):
                try:
                    # Find the corresponding index
                    rown = None
                    for j in range(len(self.sheet1_analy.data)):
                        if self.sheet1_analy.data[j][0] == self.sheet1_analy_data[i][0]:
                            rown = j

                    if rown != None:
                        for j in range(1,len(Cols)):
                            try:
                                try:
                                    self.sheet1_analy.set_cell_data(rown,j, '{:0.4e}'.format(self.sheet1_analy_data[i][j]))
                                except:
                                    self.sheet1_analy.set_cell_data(rown,j, self.sheet1_analy_data[i][j])
                            except:
                                pass
                except:
                    pass

            # Redraw the table
            self.sheet1_analy.redraw()

            # Update the Model Data
            UpdateModelData(None, self, 1, 'Analysis')

        def VE_param(values):
            #------------------------------------------------------------------
            #
            #   PURPOSE: Get List of of ViscoElastic Mechanisms.
            #
            #------------------------------------------------------------------

            # Get the value
            value = self.optmenu4_analy.get()

            # Initialize Parameters
            self.Params_VE = []
            self.Params_VE_Units = []

            # Add non-mechanism dependent parameters
            for i in range(len(self.Compare['Analysis']['Model Info']['Reversible Deformation Parameters'])):
                param = self.Compare['Analysis']['Model Info']['Reversible Deformation Parameters'][i]
                unit = self.Compare['Analysis']['Model Info']['Reversible Deformation Parameter Units'][i]
                if '_[M]' not in param:
                    self.Params_VE.append(param)
                    self.Params_VE_Units.append(unit)
                else:
                    for i in range(int(value)):
                        param_mech = param.replace("_[M]",str(i+1))
                        self.Params_VE.append(param_mech)
                        self.Params_VE_Units.append(unit)

            # Update the reversible mechanisms table
            update_reversible_table(self)

        def update_irreversible_table(self):
            #------------------------------------------------------------------
            #
            #   PURPOSE: Update the irreversible model parameters table.
            #
            #------------------------------------------------------------------

            # Delete table if it exists
            if hasattr(self,"sheet2_analy"):
                if hasattr(self,"res_flag2") == True:
                    if self.res_flag2 == 0:
                        # Store data
                        self.sheet2_analy_data = self.sheet2_analy.data
                else:
                    # Store data
                    self.sheet2_analy_data = self.sheet2_analy.data

                # Delete sheet
                self.sheet2_analy.destroy()
                del self.sheet2_analy
            
            # Check if previous data exists
            if 'Analysis' in list(self.Compare.keys()):
                # Set the reversible model type
                if 'VP_Param' in list(self.Compare['Analysis'].keys()):
                    if hasattr(self,"res_flag2") == True:
                        if self.res_flag2 == 0:
                            self.sheet2_analy_data = self.Compare['Analysis']['VP_Param']
                        else:
                            self.res_flag2 = 0
                    else:
                        self.sheet2_analy_data = self.Compare['Analysis']['VP_Param']

            if hasattr(self,"sheet2_analy_data") == False:
                self.sheet2_analy_data = []

            # Set the columns
            Cols = ['Parameter', 'Units','Value']

            # Create the table
            self.sheet2_analy = tksheet.Sheet(
                                            self.nb_tab_tab4, 
                                            total_rows = len(self.Params_VP), 
                                            total_columns = len(Cols), 
                                            headers = Cols,
                                            show_x_scrollbar = False, 
                                            show_y_scrollbar = True,
                                            font = ("Segoe UI",max([self.min_font, int(12*self.scale)]),"normal"),
                                            header_font = ("Segoe UI",max([self.min_font, int(12*self.scale)]),"bold"),
                                            )
            self.sheet2_analy.place(
                            anchor = 'n', 
                            relx = self.Placement['Analysis']['Sheet2'][0], 
                            rely = self.Placement['Analysis']['Sheet2'][1],
                            relwidth = self.Placement['Analysis']['Sheet2'][2], 
                            relheight = self.Placement['Analysis']['Sheet2'][3],
                            )
            
            if 'self.sheet2_analy' not in self.atts['Analysis']['Local']:
                self.atts['Analysis']['Local'].append('self.sheet1_analy')

            self.sheet2_analy.change_theme("blue")
            self.sheet2_analy.set_index_width(0)

            def sort_cols(self):
                #----------------------------------------------------------------------
                #
                #   PURPOSE: Custom soring function.
                #
                #----------------------------------------------------------------------

                # Get currently selected element
                currently_selected = self.sheet2_analy.get_currently_selected()
                
                # Get the list of values
                sort_list = []
                for i in range(self.sheet2_analy.visible_rows[1]):
                    sort_list.append(self.sheet2_analy.data[i][currently_selected.column])
                index_list = sorted(range(len(sort_list)), key=lambda k: sort_list[k])
                
                # Rewrite the table
                temp_data = copy.deepcopy(self.sheet2_analy.data)
                for i in range(self.sheet2_analy.visible_rows[1]):
                    for j in range(self.sheet2_analy.visible_columns[1]):
                        self.sheet2_analy.set_cell_data(i,j,temp_data[index_list[i]][j])
                self.sheet2_analy.redraw()

            # Enable Bindings
            self.sheet2_analy.enable_bindings('single_select','cell_select', 'column_select', 'edit_cell',"arrowkeys", "right_click_popup_menu")
            self.sheet2_analy.popup_menu_add_command('Sort', lambda : sort_cols(self), table_menu = True, index_menu = True, header_menu = True)
            self.sheet2_analy.extra_bindings([("cell_select", lambda event: self.cell_select_anly(event, 'sheet2_analy'))])

            # Set Column Widths
            window.update_idletasks()
            total_width = self.sheet2_analy.winfo_width()
            self.sheet2_analy.column_width(column = 0, width = int(total_width*self.Placement['Analysis']['Sheet2'][4]), redraw = True)
            self.sheet2_analy.column_width(column = 1, width = int(total_width*self.Placement['Analysis']['Sheet2'][5]), redraw = True)
            self.sheet2_analy.column_width(column = 2, width = int(total_width*self.Placement['Analysis']['Sheet2'][6]), redraw = True)
            self.sheet2_analy.table_align(align = 'c',redraw=True)

            # Set unit dictionary
            Units = {'Stress':['GPa','MPa','kPa','Pa','msi','ksi','psi'],
                     'Stress-Time':['GPa-s','MPa-s','kPa-s','Pa-s','msi-s','ksi-s','psi-s'],
                    'Time':['s'],
                    'Time-1':['1/s']
                    }

            # Set Rows
            for i in range(len(self.Params_VP)):
                self.sheet2_analy.set_cell_data(i,0, self.Params_VP[i])
                if self.Params_VP_Units[i] != None:
                    for key in list(Units.keys()):
                        if self.Params_VP_Units[i] in Units[key]:
                            units_list = Units[key]
                else:
                    units_list = []
                def_val = self.Params_VP_Units[i]
                    
                self.sheet2_analy.create_dropdown(r=i, c = 1,values=units_list)
                if def_val != None:
                    self.sheet2_analy.set_cell_data(i,1, def_val)

            # Add Existing Data
            for i in range(len(self.sheet2_analy_data)):
                try:
                    # Find the corresponding index
                    rown = None
                    for j in range(len(self.sheet2_analy.data)):
                        if self.sheet2_analy.data[j][0] == self.sheet2_analy_data[i][0]:
                            rown = j

                    if rown != None:
                        for j in range(1,len(Cols)):
                            try:
                                try:
                                    self.sheet2_analy.set_cell_data(rown,j, '{:0.4e}'.format(self.sheet2_analy_data[i][j]))
                                except:
                                    self.sheet2_analy.set_cell_data(rown,j, self.sheet2_analy_data[i][j])
                            except:
                                pass
                except:
                    pass

            # Redraw the table
            self.sheet2_analy.redraw()

            # Update the Model Data
            UpdateModelData(None, self, 2, 'Analysis')

        # Get Number of ViscoPlastic Mechanisms
        def VP_param(values):
            #------------------------------------------------------------------
            #
            #   PURPOSE: Get List of of ViscoPlastic Mechanisms.
            #
            #------------------------------------------------------------------

            # Get Value
            value = self.optmenu5_analy.get()

            # Initialize Parameters
            self.Params_VP = []
            self.Params_VP_Units = []

            # Add non-mechanism dependent parameters
            for i in range(len(self.Compare['Analysis']['Model Info']['Irreversible Deformation Parameters'])):
                param = self.Compare['Analysis']['Model Info']['Irreversible Deformation Parameters'][i]
                unit = self.Compare['Analysis']['Model Info']['Irreversible Deformation Parameter Units'][i]
                if '_[N]' not in param:
                    self.Params_VP.append(param)
                    self.Params_VP_Units.append(unit)
                else:
                    for i in range(int(value)):
                        param_mech = param.replace("_[N]",str(i+1))
                        self.Params_VP.append(param_mech)
                        self.Params_VP_Units.append(unit)

            # Update the reversible mechanisms table
            update_irreversible_table(self)

        # Get number of viscoelastic parameters
        self.VEMech = self.Compare['Analysis']['Model Info']['Reversible Mechanisms']
        if len(self.VEMech) > 0:
            # Create the label
            self.desc4_analy = ttk.Label(
                                self.nb_tab_tab4, 
                                text="Viscoelastic Mechanisms (M):", 
                                anchor=tk.CENTER,       
                                style = "Modern1.TLabel"                   
                                )
            self.desc4_analy.place(
                            anchor = 'n', 
                            relx = self.Placement['Analysis']['LabelVE'][0], 
                            rely = self.Placement['Analysis']['LabelVE'][1]
                            )
            
            if 'self.desc4_analy' not in self.atts['Analysis']['Local']:
                self.atts['Analysis']['Local'].append('self.desc4_analy')

            # Initialize number of viscoelastic mechanisms
            ve_opt = self.VEMech[0]

            # Check if previous data exists
            if 'Analysis' in list(self.Compare.keys()):
                # Set the reversible model type
                if 'M' in list(self.Compare['Analysis'].keys()):
                    if int(self.Compare['Analysis']['M']) in self.VEMech:
                        ve_opt = int(self.Compare['Analysis']['M'])

            # Create the drop down
            self.optmenu4_analy = ttk.Combobox(
                                        self.nb_tab_tab4,
                                        values=self.VEMech,
                                        style="Modern.TCombobox",
                                        state="readonly"
                                        )
            self.optmenu4_analy.configure(font = self.style_man['Combo'])
            self.optmenu4_analy.option_add('*TCombobox*Listbox.font', self.style_man['Combo'])
            self.optmenu4_analy.place(
                                anchor='n', 
                                relx = self.Placement['Analysis']['ComboVE'][0], 
                                rely = self.Placement['Analysis']['ComboVE'][1],
                                relwidth = self.Placement['Analysis']['ComboVE'][2],
                                relheight = self.Placement['Analysis']['ComboVE'][3],
                                )
            self.optmenu4_analy.set(ve_opt)
            self.optmenu4_analy.bind("<<ComboboxSelected>>",  VE_param)
            if 'self.optmenu4_analy' not in self.atts['Analysis']['Local']:
                self.atts['Analysis']['Local'].append('self.optmenu4_analy')

            # Get list of viscoelastic parameters
            VE_param(ve_opt)

        # Get Number of Viscoplastic Mechanisms
        self.VPMech = self.Compare['Analysis']['Model Info']['Irreversible Mechanisms']
        if len(self.VPMech) > 0:
            # Create the label
            self.desc5_analy = ttk.Label(self.nb_tab_tab4, 
                            text="Viscoplastic Mechanisms (N):", 
                            anchor=tk.CENTER,       
                            style = "Modern1.TLabel"                  
                            )
            self.desc5_analy.place(
                            anchor = 'n', 
                            relx = self.Placement['Analysis']['LabelVP'][0], 
                            rely = self.Placement['Analysis']['LabelVP'][1]
                            )
            if 'self.desc5_analy' not in self.atts['Analysis']['Local']:
                self.atts['Analysis']['Local'].append('self.desc5_analy')

            # Initialize Viscoplastic number of mechanisms
            vp_opt = self.VPMech[0]

            # Create the drop down menu
            self.optmenu5_analy = ttk.Combobox(
                                        self.nb_tab_tab4,
                                        values=self.VEMech,
                                        style="Modern.TCombobox",
                                        state="readonly"
                                        )
            self.optmenu5_analy.configure(font = self.style_man['Combo'])
            self.optmenu5_analy.option_add('*TCombobox*Listbox.font', self.style_man['Combo'])
            self.optmenu5_analy.place(
                                anchor='n', 
                                relx = self.Placement['Analysis']['ComboVP'][0], 
                                rely = self.Placement['Analysis']['ComboVP'][1], 
                                relwidth = self.Placement['Analysis']['ComboVP'][2], 
                                relheight = self.Placement['Analysis']['ComboVP'][3]
                                )
            self.optmenu5_analy.set(ve_opt)
            self.optmenu5_analy.bind("<<ComboboxSelected>>",  VP_param)
            if 'self.optmenu5_analy' not in self.atts['Analysis']['Local']:
                self.atts['Analysis']['Local'].append('self.optmenu5_analy')

            # Get list of viscoplastic parameters
            VP_param(vp_opt)

        # Create the Load from Database button
        self.btn_load_analy = ttk.Button(
                                    self.nb_tab_tab4, 
                                    text = "Load from Excel", 
                                    command = lambda:self.load_from_db('Analysis'), 
                                    style = "Modern3.TButton",
                                    )
        self.btn_load_analy.place(
                            anchor = 'w', 
                            relx = self.Placement['Analysis']['ButtonLoad'][0], 
                            rely = self.Placement['Analysis']['ButtonLoad'][1],
                            relwidth = self.Placement['Analysis']['ButtonLoad'][2], 
                            relheight = self.Placement['Analysis']['ButtonLoad'][3]
                            )
        if 'self.btn_load_analy' not in self.atts['Analysis']['Local']:
                self.atts['Analysis']['Local'].append('self.btn_load_analy')

        # Create button to view/delete models
        self.btn_modlib_analy = ttk.Button(
                                    self.nb_tab_tab4, 
                                    text = "Model Library", 
                                    command = lambda : self.Model_Library('Analysis'), 
                                    style = "Modern3.TButton",
                                    )
        self.btn_modlib_analy.place(
                            anchor = 'w', 
                            relx = self.Placement['Analysis']['ButtonModLib'][0], 
                            rely = self.Placement['Analysis']['ButtonModLib'][1],
                            relwidth = self.Placement['Analysis']['ButtonModLib'][2], 
                            relheight = self.Placement['Analysis']['ButtonModLib'][3]
                            )
        if 'self.btn_modlib_analy' not in self.atts['Analysis']['Local']:
                self.atts['Analysis']['Local'].append('self.btn_modlib_analy')

        # Create the Analysis button
        self.btn_anly = ttk.Button(
                                self.nb_tab_tab4, 
                                text = "Analyze", 
                                command = self.analyze, 
                                style = "Modern3.TButton",
                                )
        self.btn_anly.place(
                            anchor = 'w', 
                            relx = self.Placement['Analysis']['ButtonAnaly'][0], 
                            rely = self.Placement['Analysis']['ButtonAnaly'][1],
                            relwidth = self.Placement['Analysis']['ButtonAnaly'][2], 
                            relheight = self.Placement['Analysis']['ButtonAnaly'][3]
                            )
        if 'self.btn_anly' not in self.atts['Analysis']['Local']:
                self.atts['Analysis']['Local'].append('self.btn_anly')

        def save_model_local():
            #------------------------------------------------------------------
            #
            #   PURPOSE: Save models to the project library.
            #
            #------------------------------------------------------------------

            # Set the model type
            self.Compare['Analysis']['Compare Type'] = 'Analysis'

            # Get the save name
            save_flag = 0
            while save_flag == 0:
                user_input = simpledialog.askstring("Save Model", "Enter the model name:")

                if user_input in list(self.Compare['Model Library'].keys()):
                    askyn = messagebox.askyesno(title = 'Save Model', message = 'Do you want to overwrite ' + user_input + ' ?')
                    if askyn == True:
                        save_flag = 1
                else:
                    save_flag = 1
            
            # Save to binary in the model library
            json_string = json.dumps(self.Compare['Analysis'])
            binary_data = json_string.encode('utf-8')
            self.Compare['Model Library'][user_input] = binary_data

            # Set the model name
            self.Compare['Model ID'] = user_input

        # Create button to save a model
        self.btn_savemod_analy = ttk.Button(
                                    self.nb_tab_tab4, 
                                    text = "Save Model", 
                                    command = save_model_local, 
                                    style = "Modern3.TButton",
                                    )
        self.btn_savemod_analy.place(
                            anchor = 'w', 
                            relx = self.Placement['Analysis']['ButtonSaveMod'][0], 
                            rely = self.Placement['Analysis']['ButtonSaveMod'][1],
                            relwidth = self.Placement['Analysis']['ButtonSaveMod'][2], 
                            relheight = self.Placement['Analysis']['ButtonSaveMod'][3]
                            )
        if 'self.btn_savemod_analy' not in self.atts['Analysis']['Local']:
                self.atts['Analysis']['Local'].append('self.btn_savemod_analy')

        def add_note():
            #------------------------------------------------------------------
            #
            #   PURPOSE: Add a note to the model.
            #
            #------------------------------------------------------------------

            # Set the flag for the window
            if hasattr(self,'note_click') == False:
                self.note_click = 0

            # Open the window if it does not exist
            if self.note_click == 0:
                self.note_click = 1

                # Create the window
                root = tk.Toplevel(window)
                root.geometry(f"{int(600*self.scale)}x{int(400*self.scale)}")
                root.title("Enter Model Notes") 
                root.configure(bg='white')
                root.grab_set()
                
                # Create the label
                ttk.Label(
                        root, 
                        text="Enter Model Notes:", 
                        style = "Modern1.TLabel"
                        ).place(anchor='n', 
                                relx = self.Placement['Analysis']['NotesLabel'][0], 
                                rely = self.Placement['Analysis']['NotesLabel'][1],
                                ) 
                
                # Create the note area
                text_area = scrolledtext.ScrolledText(
                                                    root, 
                                                    wrap=tk.WORD, 
                                                    width=int(40*self.scale), 
                                                    height=int(8*self.scale), 
                                                    font=("Segoe UI", max([self.min_font, int(14*self.scale)]))) 
                text_area.place(anchor='c', 
                                relx = self.Placement['Optimization']['NotesArea'][0], 
                                rely = self.Placement['Optimization']['NotesArea'][1],  
                                relwidth = self.Placement['Optimization']['NotesArea'][2], 
                                relheight = self.Placement['Optimization']['NotesArea'][3], 
                                )

                # Display any existing notes
                if 'Note' in list(self.Compare['Analysis'].keys()):
                    text_area.insert("end", self.Compare['Analysis']['Note']) 
                
                # placing cursor in text area 
                text_area.focus()

                def on_closing_root(self):
                    #--------------------------------------------------------------
                    #
                    #   PURPOSE: Create exit protocol for the note window.
                    #
                    #--------------------------------------------------------------

                    # Save the note
                    try:
                        self.Compare['Analysis']['Note'] = text_area.get("1.0",'end-1c')
                    except:
                        pass

                    # Reset the window
                    self.note_click = 0
                    root.destroy()

                # Add the exit protocol to the root
                root.protocol("WM_DELETE_WINDOW", lambda:on_closing_root(self))


        # Create button to add a note
        self.btn_addnote_analy = ttk.Button(
                                    self.nb_tab_tab4, 
                                    text = "Model Notes", 
                                    command = add_note, 
                                    style = "Modern3.TButton",
                                    )
        self.btn_addnote_analy.place(
                            anchor = 'w', 
                            relx = self.Placement['Analysis']['ButtonNote'][0], 
                            rely = self.Placement['Analysis']['ButtonNote'][1],
                            relwidth = self.Placement['Analysis']['ButtonNote'][2], 
                            relheight = self.Placement['Analysis']['ButtonNote'][3]
                            )
        if 'self.btn_addnote_analy' not in self.atts['Analysis']['Local']:
                self.atts['Analysis']['Local'].append('self.btn_addnote_analy')

        # Function to view model history
        def view_history(self):
            #--------------------------------------------------------------
            #
            #   PURPOSE: View run history for this project.
            #
            #--------------------------------------------------------------

            def view_hist_data(self):
                #--------------------------------------------------------------
                #
                #   PURPOSE: Get parameters and test error for chosen run.
                #
                #--------------------------------------------------------------

                # Delete the old tables if they exist
                try:
                    self.hist_param_sheet.destroy()
                    self.hist_test_sheet.destroy()
                except:
                    pass

                # Get currently selected row
                currently_selected = self.run_hist_sheet.get_currently_selected()

                # Highlight Row
                for i in range(len(self.run_hist_sheet.data)):
                    self.run_hist_sheet.highlight_rows(i,'white','black')
                self.run_hist_sheet.highlight_rows(currently_selected.row,'lightblue1','black')

                # Get parameters
                params = self.run_history[self.run_hist_sheet.data[currently_selected.row][0]]['Parameters']
                param_keys = list(params.keys())

                # Create Parameter sheet
                Cols = ['Parameter', 'Value', 'Unit']
                self.hist_param_sheet = tksheet.Sheet(
                                                root, 
                                                total_rows = len(param_keys), 
                                                total_columns = len(Cols), 
                                                headers = Cols, 
                                                show_x_scrollbar = False, 
                                                show_y_scrollbar = True,
                                                font = ("Segoe UI",max([self.min_font, int(12*self.scale)]),"normal"),
                                                header_font = ("Segoe UI",max([self.min_font, int(12*self.scale)]),"bold"))
                self.hist_param_sheet.place(
                                    anchor = 'ne', 
                                    relx = self.Placement['Optimization']['HistSheetPar'][0], 
                                    rely = self.Placement['Optimization']['HistSheetPar'][1],
                                    relwidth = self.Placement['Optimization']['HistSheetPar'][2], 
                                    relheight = self.Placement['Optimization']['HistSheetPar'][3],
                                    )

                # Format the sheet
                self.hist_param_sheet.change_theme("blue")
                root.update_idletasks()
                total_width = self.hist_param_sheet.winfo_width()
                self.hist_param_sheet.column_width(column = 0, width = int(total_width*self.Placement['Optimization']['HistSheetPar'][4]), redraw = True)
                self.hist_param_sheet.column_width(column = 1, width = int(total_width*self.Placement['Optimization']['HistSheetPar'][5]), redraw = True)
                self.hist_param_sheet.column_width(column = 2, width = int(total_width*self.Placement['Optimization']['HistSheetPar'][6]), redraw = True)
                self.hist_param_sheet.table_align(align = 'c',redraw=True)
                self.hist_param_sheet.set_index_width(0)

                # Fill existing values values
                for i, key in enumerate(param_keys):
                    self.hist_param_sheet.set_cell_data(i,0,key)
                    self.hist_param_sheet.set_cell_data(i,1,params[key][0])
                    self.hist_param_sheet.set_cell_data(i,2,params[key][1]) 
                self.hist_param_sheet.redraw()

                # Get parameters
                tests = self.run_history[self.run_hist_sheet.data[currently_selected.row][0]]['Test Error']
                test_keys = list(tests.keys())

                # Create Parameter sheet
                Cols = ['Name', 'Type', 'Weight', 'Error']
                self.hist_test_sheet = tksheet.Sheet(
                                                root, 
                                                total_rows = len(test_keys), 
                                                total_columns = len(Cols), 
                                                headers = Cols,
                                                show_x_scrollbar = False, 
                                                show_y_scrollbar = True,
                                                font = ("Segoe UI",max([self.min_font, int(12*self.scale)]),"normal"),
                                                header_font = ("Segoe UI",max([self.min_font, int(12*self.scale)]),"bold"),
                                                )
                self.hist_test_sheet.place(
                                    anchor = 'ne', 
                                    relx = self.Placement['Optimization']['HistSheetTest'][0], 
                                    rely = self.Placement['Optimization']['HistSheetTest'][1], 
                                    relwidth = self.Placement['Optimization']['HistSheetTest'][2], 
                                    relheight = self.Placement['Optimization']['HistSheetTest'][3], 
                                    )
                
                # Format the sheet
                self.hist_test_sheet.change_theme("blue")
                root.update_idletasks()
                total_width = self.hist_test_sheet.winfo_width()
                self.hist_test_sheet.column_width(column = 0, width = int(total_width*self.Placement['Optimization']['HistSheetTest'][4]), redraw = True)
                self.hist_test_sheet.column_width(column = 1, width = int(total_width*self.Placement['Optimization']['HistSheetTest'][5]), redraw = True)
                self.hist_test_sheet.column_width(column = 2, width = int(total_width*self.Placement['Optimization']['HistSheetTest'][6]), redraw = True)
                self.hist_test_sheet.column_width(column = 3, width = int(total_width*self.Placement['Optimization']['HistSheetTest'][7]), redraw = True)
                self.hist_test_sheet.table_align(align = 'c',redraw=True)
                self.hist_test_sheet.set_index_width(0)

                # Fill existing values values
                for i, key in enumerate(test_keys):
                    self.hist_test_sheet.set_cell_data(i,0,key)
                    self.hist_test_sheet.set_cell_data(i,1,tests[key][0])
                    self.hist_test_sheet.set_cell_data(i,2,tests[key][1]) 
                    self.hist_test_sheet.set_cell_data(i,3,tests[key][2]) 
                self.hist_test_sheet.redraw()

                # Deselect
                self.run_hist_sheet.deselect("all", redraw=True)

            # Get the log file
            with open(self.log_file, "r", encoding="utf-8") as f:
                lines = f.readlines()
            lines = [line.strip() for line in lines]

            # Get indices of each test run
            test_start = []
            for i, line in enumerate(lines):
                if "-- NEW RUN --" in line:
                    test_start.append(i)
            test_start.append(len(lines))

            self.run_history = {}
            for i in range(len(test_start)-1):
                # Initialize Data
                self.run_history['Run #' + str(i+1)] = {}

                # Get Optimization Results
                self.run_history['Run #' + str(i+1)]['Parameters'] = {}
                self.run_history['Run #' + str(i+1)]['Global Error'] = None
                self.run_history['Run #' + str(i+1)]['Test Error'] = {}
                for j in range(test_start[i], test_start[i+1]):
                    keys = ['Model Name', 'Reversible Model Name','Irreversible Model','Viscoelastic Mechanisms','Viscoplastic Mechanisms']
                    for key in keys:
                        if ':' in lines[j] and key == lines[j].split(':')[0].strip():
                            self.run_history['Run #' + str(i+1)][key.split(':')[0]] = lines[j].split(':')[1].strip()

                    if "OPTIMIZATION RESULTS:" in lines[j]:
                        ct = 2
                        while lines[j+ct] != '':
                            line = lines[j+ct]
                            line = line.split(' ')
                            cline = [x for x in line if x != '']
                            try:
                                self.run_history['Run #' + str(i+1)]['Parameters'][cline[0]] = [float(cline[2]), cline[1]]
                            except:
                                self.run_history['Run #' + str(i+1)]['Parameters'][cline[0]] = [float(cline[1]), '']
                            ct = ct + 1
                            if j+ct == test_start[i+1]:
                                break

                    if "ERROR:" in lines[j]:
                        self.run_history['Run #' + str(i+1)]['Global Error'] = float(lines[j+1].split('=')[1])
                        ct = 5
                        while lines[j+ct] != '':
                            line = lines[j+ct]
                            line = line.split(' ')
                            cline = [x for x in line if x != '']
                            self.run_history['Run #' + str(i+1)]['Test Error'][cline[0]] = [cline[1], float(cline[2]), float(cline[3])]

                            ct = ct + 1
                            if j+ct == test_start[i+1]:
                                break


            def load_hist_data(self):
                #--------------------------------------------------------------
                #
                #   PURPOSE: Load a previous run back into PYCOMPARE
                #
                #--------------------------------------------------------------

                # Get currently selected row
                currently_selected = self.run_hist_sheet.get_currently_selected()

                # Get the model type
                mod_type = self.run_history[self.run_hist_sheet.data[currently_selected.row][0]]['Model Name']
                try:
                    self.optmenu1_analy.set(mod_type)
                    change_model(mod_type)
                except:
                    messagebox.showerror('Unable to load model.')

                # Get the reversible model
                try:
                    if 'Reversible Model Name' in self.run_history[self.run_hist_sheet.data[currently_selected.row][0]].keys():
                        rev_model = self.run_history[self.run_hist_sheet.data[currently_selected.row][0]]['Reversible Model Name']
                        if rev_model in self.optmenu2_analy['values']:
                            self.optmenu2_analy.set(rev_model)
                            UpdateModelData(None, self, 1, 'Analysis')
                except:
                    pass

                # Get the reversible mechanisms
                try:
                    if 'Viscoelastic Mechanisms' in self.run_history[self.run_hist_sheet.data[currently_selected.row][0]].keys():
                        rev_mech = self.run_history[self.run_hist_sheet.data[currently_selected.row][0]]['Viscoelastic Mechanisms']
                        if rev_mech in self.optmenu4_analy['values']:
                            self.optmenu4_analy.set(rev_mech)
                            VE_param(rev_mech)
                except:
                    pass

                # Get the irreversible model
                try:
                    if 'Irreversible Model Name' in self.run_history[self.run_hist_sheet.data[currently_selected.row][0]].keys():
                        irrev_model = self.run_history[self.run_hist_sheet.data[currently_selected.row][0]]['Irreversible Model Name']
                        if irrev_model in self.optmenu3_analy['values']:
                            self.optmenu3_analy.set(irrev_model)
                            UpdateModelData(None, self, 2, 'Analysis')
                except:
                    pass

                # Get the irreversible mechanisms
                try:
                    if 'Viscoplastic Mechanisms' in self.run_history[self.run_hist_sheet.data[currently_selected.row][0]].keys():
                        irrev_mech = self.run_history[self.run_hist_sheet.data[currently_selected.row][0]]['Viscoplastic Mechanisms']
                        if irrev_mech in self.optmenu5_analy['values']:
                            self.optmenu5_analy.set(irrev_mech)
                            VP_param(irrev_mech)
                except:
                    pass

                # Get parameters
                params = self.run_history[self.run_hist_sheet.data[currently_selected.row][0]]['Parameters']
                param_keys = list(params.keys())

                # Write Parameters
                for key in param_keys:
                    for i in range(len(self.sheet1_analy.data)):
                        if key == self.sheet1_analy.data[i][0]:
                            self.sheet1_analy.set_cell_data(i,1,params[key][1])
                            self.sheet1_analy.set_cell_data(i,2,'{:0.4e}'.format(params[key][0]))

                    for i in range(len(self.sheet2_analy.data)):
                        if key == self.sheet2_analy.data[i][0]:
                            self.sheet2_analy.set_cell_data(i,1,params[key][1])
                            self.sheet2_analy.set_cell_data(i,2,'{:0.4e}'.format(params[key][0]))
            
                # Recreate the Optimize Page
                self.analy_init = 1
                self.viz_init = 0
                root.destroy()
                CreateAnalysisTab(self,window)


            # Create the run history window
            root = tk.Toplevel(window)
            root.title("Run History")
            root.geometry(f"{int(900*self.scale)}x{int(600*self.scale)}")
            root.resizable(False, False)
            root.configure(bg='white')
            root.grab_set()

            # Create the label
            ttk.Label(
                        root, 
                        text="Run History", 
                        anchor=tk.CENTER,       
                        style = "Modern1.TLabel"                   
                        ).place(
                                anchor='n', 
                                relx = self.Placement['Optimization']['HistLabel'][0], 
                                rely = self.Placement['Optimization']['HistLabel'][1]
                                )
            
            # Create the sheet
            Cols = ['Run', 'No. of Tests', 'Global Error']
            self.run_hist_sheet = tksheet.Sheet(
                                            root, 
                                            total_rows = len(self.run_history.keys()), 
                                            total_columns = len(Cols), 
                                            headers = Cols,
                                            show_x_scrollbar = False, 
                                            show_y_scrollbar = True,
                                            font = ("Segoe UI",max([self.min_font, int(12*self.scale)]),"normal"),
                                            header_font = ("Segoe UI",max([self.min_font, int(12*self.scale)]),"bold"),
            )
            self.run_hist_sheet.place(
                                anchor = 'nw', 
                                relx = self.Placement['Optimization']['HistSheetRun'][0], 
                                rely = self.Placement['Optimization']['HistSheetRun'][1],
                                relwidth = self.Placement['Optimization']['HistSheetRun'][2], 
                                relheight = self.Placement['Optimization']['HistSheetRun'][3], 
                                )
            
            
            # Format the sheet
            self.run_hist_sheet.change_theme("blue")
            root.update_idletasks()
            total_width = self.run_hist_sheet.winfo_width()
            self.run_hist_sheet.column_width(column = 0, width = int(total_width*self.Placement['Optimization']['HistSheetRun'][4]), redraw = True)
            self.run_hist_sheet.column_width(column = 1, width = int(total_width*self.Placement['Optimization']['HistSheetRun'][5]), redraw = True)
            self.run_hist_sheet.column_width(column = 2, width = int(total_width*self.Placement['Optimization']['HistSheetRun'][6]), redraw = True)
            self.run_hist_sheet.table_align(align = 'c',redraw=True)
            self.run_hist_sheet.set_index_width(0)

            # Enable Bindings
            self.run_hist_sheet.enable_bindings('single_select','cell_select', 'column_select',"arrowkeys","rc_popup_menu")
            self.run_hist_sheet.popup_menu_add_command('View Data', lambda : view_hist_data(self), table_menu = True, index_menu = True, header_menu = True)
            self.run_hist_sheet.popup_menu_add_command('Load Model', lambda : load_hist_data(self), table_menu = True, index_menu = True, header_menu = True)

            
            # Fill existing values values
            key_list = list(self.run_history.keys())
            key_list.reverse()
            for i, key in enumerate(key_list):
                j = list(self.run_history.keys()).index(key)
                self.run_hist_sheet.set_cell_data(i,0,list(self.run_history.keys())[j])
                self.run_hist_sheet.set_cell_data(i,1,len(self.run_history[list(self.run_history.keys())[j]]['Test Error'].keys()))
                self.run_hist_sheet.set_cell_data(i,2,self.run_history[list(self.run_history.keys())[j]]['Global Error']) 
            self.run_hist_sheet.redraw()


        # Create button to view model history
        self.btn_view_hist_analy = ttk.Button(
                                    self.nb_tab_tab4, 
                                    text = "Run History", 
                                    command = lambda : view_history(self), 
                                    style = "Modern3.TButton",
                                    )
        self.btn_view_hist_analy.place(
                            anchor = 'w', 
                            relx = self.Placement['Analysis']['ButtonView'][0], 
                            rely = self.Placement['Analysis']['ButtonView'][1], 
                            relwidth = self.Placement['Analysis']['ButtonView'][2], 
                            relheight = self.Placement['Analysis']['ButtonView'][3]
                            )
        if 'self.btn_view_hist_analy' not in self.atts['Analysis']['Local']:
            self.atts['Analysis']['Local'].append('self.btn_view_hist_analy')

        # Update Model Data
        UpdateModelData(None, self, 3, 'Analysis')

        # Update Tables
        if len(self.Compare['Analysis']['Model Info']['Reversible Models']) > 0:
                update_reversible_table(self)
        if len(self.Compare['Analysis']['Model Info']['Irreversible Models']) > 0:
            update_irreversible_table(self)

    # Create the label for Model Type
    
    
    if self.analy_init == 1:

        self.desc1_analy = ttk.Label(
                        self.nb_tab_tab4, 
                        text="Select the Model:", 
                        anchor=tk.NW,       
                        style = "Modern1.TLabel"                   
                        )
                        
        self.desc1_analy.place(
                        anchor = 'nw', 
                        relx = self.Placement['Analysis']['LabelSelModel'][0], 
                        rely = self.Placement['Analysis']['LabelSelModel'][1],
                        relwidth = self.Placement['Analysis']['LabelSelModel'][2], 
                        relheight = self.Placement['Analysis']['LabelSelModel'][3]
                        )

        self.atts['Analysis']['Permanent'].append('self.desc1_analy') 

        # Initialize the model option
        mod_opt = self.Models[0]

        # Check if previous value exists
        if 'Analysis' in list(self.Compare.keys()):
            # Set the model name
            if 'Model Name' in list(self.Compare['Analysis'].keys()):
                if self.Compare['Analysis']['Model Name'] in self.Models:
                    mod_opt = self.Compare['Analysis']['Model Name']

        # Create Option Menu for Model Type
        self.optmenu1_analy = ttk.Combobox(
                                    self.nb_tab_tab4,
                                    values=self.Models,
                                    style="Modern.TCombobox",
                                    state="readonly"
                                    )
        self.optmenu1_analy.configure(font = self.style_man['Combo'])
        self.optmenu1_analy.option_add('*TCombobox*Listbox.font', self.style_man['Combo'])
        self.optmenu1_analy.place(
                            anchor='nw', 
                            relx = self.Placement['Analysis']['ComboSelModel'][0], 
                            rely = self.Placement['Analysis']['ComboSelModel'][1],
                            relwidth = self.Placement['Analysis']['ComboSelModel'][2], 
                            relheight = self.Placement['Analysis']['ComboSelModel'][3]
                            )
        self.optmenu1_analy.set(mod_opt)
        self.optmenu1_analy.bind("<<ComboboxSelected>>",  change_model)
        change_model(mod_opt)
        self.atts['Analysis']['Permanent'].append('self.optmenu1_analy')