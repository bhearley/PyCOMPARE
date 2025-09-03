#--------------------------------------------------------------------------------------------------------------------------------------------------------------------
#
# CreateExportTab.py
#
# PURPOSE: Create the Export tab. The Export tab allows users to export all project information to Excel.
#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------
def CreateExportTab(self,window):
    # Import Modules
    import json
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Border, Side
    import threading
    import tkinter as tk
    from tkinter import filedialog
    from tkinter import messagebox
    from tkinter import ttk

    # Check for export initialization
    if self.exp_init == 1:

        # Create Export Buttons
        def export():
            #--------------------------------------------------------------------------
            #
            #   PURPOSE: Begin the export process.
            #
            #--------------------------------------------------------------------------

            # Check that a model exists
            if  'Model ID' in self.Compare.keys():
                if self.Compare['Model ID'] != None:

                    def export_ex(callback, self):
                        #--------------------------------------------------------------------------
                        #
                        #   PURPOSE: Create the excel output file.
                        #
                        #--------------------------------------------------------------------------

                        # Get the model information
                        json_string = self.Compare['Model Library'][self.Compare['Model ID']].decode('utf-8')
                        data = json.loads(json_string)

                        # Get the characterization test names
                        char_tests = list(self.Compare['Characterization'].keys())
                        pred_tests = list(self.Compare['Prediction'].keys())

                        # Open the excel template
                        wb = load_workbook(self.Compare['Paths']['Export Template'], data_only=True)

                        def FormatArea(ws, x1, x2, y1, y2):
                            #----------------------------------------------------------
                            #
                            #   PURPOSE: Format an area in the excel output file.
                            #
                            #----------------------------------------------------------

                            # Make all cells white
                            for i in range(x1, x2+1):
                                for j in range(y1, y2+1):
                                    ws.cell(row=i, column=j).fill = PatternFill(fill_type='solid', start_color='FFFFFF', end_color='FFFFFF')

                            # Create the borders
                            border_thin = Side(style='thin')
                            border_thick = Side(style='thick')

                            # Format Border
                            for i in range(x1, x2+1):
                                ws.cell(row = i, column = y1).border = Border(top=border_thin, left=border_thick, right=border_thin, bottom=border_thin)
                                ws.cell(row = i, column = y2).border = Border(top=border_thin, left=border_thin, right=border_thick, bottom=border_thin)

                            for i in range(y1, y2+1):
                                ws.cell(row = x1, column = i).border = Border(top=border_thick, left=border_thin, right=border_thin, bottom=border_thin)
                                ws.cell(row = x2, column = i).border = Border(top=border_thin, left=border_thin, right=border_thin, bottom=border_thick)

                            # Format Interior
                            for i in range(x1+1,x2):
                                for j in range(y1+1,y2):
                                    ws.cell(row = i, column = j).border = Border(top=border_thin, left=border_thin, right=border_thin, bottom=border_thin)

                            # Format Corners
                            ws.cell(row = x1, column = y1).border = Border(top=border_thick, left=border_thick, right=border_thin, bottom=border_thin)
                            ws.cell(row = x2, column = y1).border = Border(top=border_thin, left=border_thick, right=border_thin, bottom=border_thick)
                            ws.cell(row = x1, column = y2).border = Border(top=border_thick, left=border_thin, right=border_thick, bottom=border_thin)
                            ws.cell(row = x2, column = y2).border = Border(top=border_thin, left=border_thin, right=border_thick, bottom=border_thick)
            
                        # Model Information
                        ws = wb['Model Information']
                        if self.var_exp_chk1.get() == 1:
                            
                            # -- General Information
                            ws.cell(row = 3, column = 3).value = self.Compare['Model ID']
                            ws.cell(row = 4, column = 3).value = data['Model Name']
                            ws.cell(row = 5, column = 3).value = self.Compare['Characterization'][char_tests[0]]['Temperature'][0]
                            ws.cell(row = 6, column = 3).value = data['Compare Type']
                            if 'Note' in data.keys():
                                ws.cell(row = 7, column = 3).value = data['Note']
                            FormatArea(ws, 2, 7, 2, 8)

                            # -- Reversible Model
                            if 'Reversible Model Name' in data.keys():
                                ws.cell(row = 10, column = 3).value = data['Reversible Model Name']
                            if 'M' in data.keys():
                                ws.cell(row = 11, column = 3).value = int(data['M'])
                            ct = 0
                            if 'VE_Param' in data.keys():
                                ct = len(data['VE_Param'])
                                for i in range(len(data['VE_Param'])):
                                    ws.cell(row=14+i,column=2).value = data['VE_Param'][i][0]
                                    ws.cell(row=14+i,column=3).value = data['VE_Param'][i][1]
                                    if data['Compare Type'] == 'Optimize':
                                        ws.cell(row=14+i,column=4).value = data['VE_Param'][i][6]
                                    else:
                                        ws.cell(row=14+i,column=4).value = data['VE_Param'][i][2]
                            FormatArea(ws, 9, 13+ ct, 2, 4)

                            # -- Irreversible Model
                            if 'Irreversible Model Name' in data.keys():
                                ws.cell(row = 10, column = 7).value = data['Irreversible Model Name']
                            if 'N' in data.keys():
                                ws.cell(row = 11, column = 7).value = int(data['N'])
                            ct = 0
                            if 'VP_Param' in data.keys():
                                ct = len(data['VP_Param'])
                                for i in range(len(data['VP_Param'])):
                                    ws.cell(row=14+i,column=6).value = data['VP_Param'][i][0]
                                    ws.cell(row=14+i,column=7).value = data['VP_Param'][i][1]
                                    if data['Compare Type'] == 'Optimize':
                                        ws.cell(row=14+i,column=8).value = data['VP_Param'][i][6]
                                    else:
                                        ws.cell(row=14+i,column=8).value = data['VP_Param'][i][2]
                            FormatArea(ws, 9, 13+ ct, 6, 8)
                        else:
                            ws.sheet_state = 'hidden'

                        # Test information
                        ws = wb['Test Information']
                        if self.var_exp_chk2.get() == 1:
                            
                            # -- Write Characaterization Test Data
                            for i in range(len(char_tests)):
                                ws.cell(row = 3+i, column = 2).value = char_tests[i]
                                ws.cell(row = 3+i, column = 3).value = self.Compare['Characterization'][char_tests[i]]['Test Type']
                                ws.cell(row = 3+i, column = 4).value = self.Compare['Characterization'][char_tests[i]]['Loading Direction'][0]
                                ws.cell(row = 3+i, column = 5).value = self.Compare['Characterization'][char_tests[i]]['Control'][0]
                                ws.cell(row = 3+i, column = 6).value = self.Compare['Characterization'][char_tests[i]]['Load Rate'][0][0]
                                ws.cell(row = 3+i, column = 7).value = self.Compare['Characterization'][char_tests[i]]['Load Rate'][0][1]
                                ws.cell(row = 3+i, column = 8).value = self.Compare['Characterization'][char_tests[i]]['Angle']
                                ws.cell(row = 3+i, column = 9).value = 'Characterization'
                                ws.cell(row = 3+i, column = 10).value = self.Compare['Characterization'][char_tests[i]]['RelWeight']
                                ws.cell(row = 3+i, column = 11).value = self.Compare['Prediction'][char_tests[i]]['Error']
                            ct = len(char_tests)

                            # -- Write Verification Test Data
                            for i in range(len(pred_tests)):
                                if self.Compare['Prediction'][pred_tests[i]] is not None:
                                    if pred_tests[i] not in char_tests:
                                        if self.Compare['Prediction'][pred_tests[i]]['Error'] is not None:
                                            ws.cell(row = 3+ct, column = 2).value = pred_tests[i]
                                            ws.cell(row = 3+ct, column = 3).value = self.Compare['Data'][pred_tests[i]]['Test Type']
                                            ws.cell(row = 3+ct, column = 4).value = self.Compare['Data'][pred_tests[i]]['Loading Direction'][0]
                                            ws.cell(row = 3+ct, column = 5).value = self.Compare['Data'][pred_tests[i]]['Control'][0]
                                            ws.cell(row = 3+ct, column = 6).value = self.Compare['Data'][pred_tests[i]]['Load Rate'][0][0]
                                            ws.cell(row = 3+ct, column = 7).value = self.Compare['Data'][pred_tests[i]]['Load Rate'][0][1]
                                            ws.cell(row = 3+ct, column = 8).value = self.Compare['Data'][pred_tests[i]]['Angle']
                                            ws.cell(row = 3+ct, column = 9).value = 'Verification'
                                            ws.cell(row = 3+ct, column = 11).value = self.Compare['Prediction'][pred_tests[i]]['Error']
                                            ct = ct + 1

                            FormatArea(ws, 2, 2+ct, 2, 11)
                        else:
                            ws.sheet_state = 'hidden'

                        # Experimental Curves
                        ws = wb['Experimental Curves']
                        if self.var_exp_chk3.get() == 1:
                            # Set Max x
                            max_x = 2

                            # Write response curves
                            start_col = 2
                            for i in range(len(pred_tests)):
                                start_col_test = start_col

                                if self.Compare['Prediction'][pred_tests[i]] is not None:
                                    # -- Write the test name
                                    ws.cell(row = 2, column=start_col).value = pred_tests[i]

                                    # -- Write Time
                                    if 'Time' in self.Compare['Prediction'][pred_tests[i]].keys():
                                        ws.cell(row = 3, column=start_col).value = 'Time (s)'
                                        for j in range(len(self.Compare['Data'][pred_tests[i]]['Time'][:self.Compare['Data'][pred_tests[i]]['Stage Index'][-1]])):
                                            ws.cell(row = 4+j, column=start_col).value = self.Compare['Data'][pred_tests[i]]['Time'][j]
                                        if j+4 > max_x:
                                            max_x = j+4
                                        start_col = start_col + 1

                                    # -- Write Strain
                                    if 'Strain' in self.Compare['Prediction'][pred_tests[i]].keys():
                                        keys = list(self.Compare['Prediction'][pred_tests[i]]['Strain'].keys())
                                        for key in keys:
                                            ws.cell(row = 3, column=start_col).value = 'Strain-' + str(key) + '(-)'
                                            for j in range(len(self.Compare['Data'][pred_tests[i]]['Strain'][key][:self.Compare['Data'][pred_tests[i]]['Stage Index'][-1]])):
                                                ws.cell(row = 4+j, column=start_col).value = self.Compare['Data'][pred_tests[i]]['Strain'][key][j]
                                            if j+4 > max_x:
                                                max_x = j+4
                                            start_col = start_col + 1

                                    # -- Write Stress
                                    if 'Stress' in self.Compare['Prediction'][pred_tests[i]].keys():
                                        keys = list(self.Compare['Prediction'][pred_tests[i]]['Stress'].keys())
                                        for key in keys:
                                            ws.cell(row = 3, column=start_col).value = 'Stress-' + str(key) + '(MPa)'
                                            for j in range(len(self.Compare['Data'][pred_tests[i]]['Stress'][key][:self.Compare['Data'][pred_tests[i]]['Stage Index'][-1]])):
                                                ws.cell(row = 4+j, column=start_col).value = self.Compare['Data'][pred_tests[i]]['Stress'][key][j]
                                            if j > max_x:
                                                max_x = j+4
                                            start_col = start_col + 1

                                    # -- Merge Test Cells
                                    ws.merge_cells(start_row=2, start_column=start_col_test, end_row=2, end_column=start_col-1)

                            FormatArea(ws, 2, max_x, 2, start_col-1)
                        else:
                            ws.sheet_state = 'hidden'

                        # Reduced Curves
                        ws = wb['Reduced Curves']
                        if self.var_exp_chk4.get() == 1:
                            # Set Max x
                            max_x = 4

                            # Write response curves
                            start_col = 2
                            for i in range(len(pred_tests)):
                                start_col_test = start_col

                                if self.Compare['Prediction'][pred_tests[i]] is not None:
                                    # -- Write the test name
                                    ws.cell(row = 2, column=start_col).value = pred_tests[i]

                                    # -- Write Time
                                    if 'Time' in self.Compare['Prediction'][pred_tests[i]].keys():
                                        ws.cell(row = 3, column=start_col).value = 'Time (s)'
                                        for j in range(len(self.Compare['Data'][pred_tests[i]]['Reduced Data']['Time'])):
                                            ws.cell(row = 4+j, column=start_col).value = self.Compare['Data'][pred_tests[i]]['Reduced Data']['Time'][j]
                                        if j+4 > max_x:
                                            max_x = j+4
                                        start_col = start_col + 1

                                    # -- Write Strain
                                    if 'Strain' in self.Compare['Prediction'][pred_tests[i]].keys():
                                        keys = list(self.Compare['Prediction'][pred_tests[i]]['Strain'].keys())
                                        for key in keys:
                                            ws.cell(row = 3, column=start_col).value = 'Strain-' + str(key) + '(-)'
                                            for j in range(len(self.Compare['Data'][pred_tests[i]]['Reduced Data']['Strain'][key])):
                                                ws.cell(row = 4+j, column=start_col).value = self.Compare['Data'][pred_tests[i]]['Reduced Data']['Strain'][key][j]
                                            if j+4 > max_x:
                                                max_x = j+4
                                            start_col = start_col + 1

                                    # -- Write Stress
                                    if 'Stress' in self.Compare['Prediction'][pred_tests[i]].keys():
                                        keys = list(self.Compare['Prediction'][pred_tests[i]]['Stress'].keys())
                                        for key in keys:
                                            ws.cell(row = 3, column=start_col).value = 'Stress-' + str(key) + '(MPa)'
                                            for j in range(len(self.Compare['Data'][pred_tests[i]]['Reduced Data']['Stress'][key])):
                                                ws.cell(row = 4+j, column=start_col).value = self.Compare['Data'][pred_tests[i]]['Reduced Data']['Stress'][key][j]
                                            if j+4 > max_x:
                                                max_x = j+4
                                            start_col = start_col + 1

                                    # -- Merge Test Cells
                                    ws.merge_cells(start_row=2, start_column=start_col_test, end_row=2, end_column=start_col-1)

                            FormatArea(ws, 2, max_x, 2, start_col-1)
                        else:
                            ws.sheet_state = 'hidden'
                        
                        # Predicted Curves
                        ws = wb['Predicted Curves']
                        if self.var_exp_chk5.get() == 1:
                            # Set Max x
                            max_x = 2

                            # Write response curves
                            start_col = 2
                            for i in range(len(pred_tests)):
                                start_col_test = start_col
                                if self.Compare['Prediction'][pred_tests[i]] is not None:
                                    # -- Write the test name
                                    ws.cell(row = 2, column=start_col).value = pred_tests[i]

                                    # -- Write Time
                                    if 'Time' in self.Compare['Prediction'][pred_tests[i]].keys():
                                        ws.cell(row = 3, column=start_col).value = 'Time (s)'
                                        for j in range(len(self.Compare['Prediction'][pred_tests[i]]['Time'])):
                                            ws.cell(row = 4+j, column=start_col).value = self.Compare['Prediction'][pred_tests[i]]['Time'][j]
                                        if j+4 > max_x:
                                            max_x = j+4
                                        start_col = start_col + 1

                                    # -- Write Strain
                                    if 'Strain' in self.Compare['Prediction'][pred_tests[i]].keys():
                                        keys = list(self.Compare['Prediction'][pred_tests[i]]['Strain'].keys())
                                        for key in keys:
                                            ws.cell(row = 3, column=start_col).value = 'Strain-' + str(key) + '(-)'
                                            for j in range(len(self.Compare['Prediction'][pred_tests[i]]['Strain'][key])):
                                                ws.cell(row = 4+j, column=start_col).value = self.Compare['Prediction'][pred_tests[i]]['Strain'][key][j]
                                            if j+4 > max_x:
                                                max_x = j+4
                                            start_col = start_col + 1

                                    # -- Write Stress
                                    if 'Stress' in self.Compare['Prediction'][pred_tests[i]].keys():
                                        keys = list(self.Compare['Prediction'][pred_tests[i]]['Stress'].keys())
                                        for key in keys:
                                            ws.cell(row = 3, column=start_col).value = 'Stress-' + str(key) + '(MPa)'
                                            for j in range(len(self.Compare['Prediction'][pred_tests[i]]['Stress'][key])):
                                                ws.cell(row = 4+j, column=start_col).value = self.Compare['Prediction'][pred_tests[i]]['Stress'][key][j]
                                            if j+4 > max_x:
                                                max_x = j+4
                                            start_col = start_col + 1

                                    # -- Merge Test Cells
                                    ws.merge_cells(start_row=2, start_column=start_col_test, end_row=2, end_column=start_col-1)

                            FormatArea(ws, 2, max_x, 2, start_col-1)

                        else:
                            ws.sheet_state = 'hidden'

                        # Ask for the save name from the user
                        file_path = filedialog.asksaveasfilename(
                                        defaultextension=".xlsx",
                                        filetypes=[("Excel Files", "*.xlsx")],
                                        title="Save As",
                                        confirmoverwrite=True
                                    )
                        
                        if file_path:
                            try:
                                wb.save(file_path)
                                messagebox.showinfo(message = 'Export File Saved!')
                            except:
                                messagebox.showerror(message = "An error occurred. Please check that the file is closed.")

                        # Notify when done
                        callback()
                    
                    # Function to display progress bar while running
                    def start_export(self):

                        #----------------------------------------------------------
                        #
                        #   PURPOSE: Show progress bar during export
                        #
                        #----------------------------------------------------------

                        # Create the window
                        loading = tk.Toplevel(window)
                        loading.title("Exporting to Excel")
                        loading.geometry("300x100")
                        loading.resizable(False, False)
                        loading.configure(bg='white')
                        loading.grab_set()  

                        # Function for progress bar Exit Protocol
                        def on_closing_saving(self):

                            # Don't allow exit while saving
                            return
                        
                        # Create the window exit protocal
                        loading.protocol("WM_DELETE_WINDOW", lambda:on_closing_saving(self))

                        # Create the loading label
                        ttk.Label(
                                loading, 
                                text="Writing Data to Excel - Please Wait", 
                                style = "Modern2.TLabel"
                                ).pack(pady=10)

                        # Create the progress bar
                        pb = ttk.Progressbar(
                                            loading, 
                                            mode='indeterminate',
                                            style = "Modern.Horizontal.TProgressbar"
                                            )
                        pb.pack(fill='x', padx=20, pady=10)
                        pb.start(10)

                        # Function to close window when task is completed
                        def on_task_done():

                            # Stop Progress bar
                            pb.stop()

                            # Destroy Window
                            loading.destroy()

                        # Begin save on background thread
                        threading.Thread(target=export_ex, args=(on_task_done,self), daemon=True).start()

                        # Wait until loading window is closed
                        window.wait_window(loading)

                    start_export(self)

                else:
                        messagebox.showinfo(message = 'No Model Selected.')
            else:
                messagebox.showinfo(message = 'No Model Selected.')


        # Create label
        self.exp_options = ttk.Label( 
                                self.nb_tab_tab6, 
                                text="Options:",
                                style = "Modern4.TLabel",
                                )
        self.exp_options.place(
                            anchor='nw', 
                            relx = self.Placement['Export']['LabelOpts'][0], 
                            rely = self.Placement['Export']['LabelOpts'][1]
                            )

        # Create Options
        self.var_exp_chk1 = tk.IntVar(value=1)
        self.exp_chk1 = ttk.Checkbutton(
                                        self.nb_tab_tab6, 
                                        text="Model Information", 
                                        variable=self.var_exp_chk1,
                                        style = "TCheckbutton" 
                                        )
        self.exp_chk1.place(
                            anchor='nw', 
                            relx = self.Placement['Export']['Check1'][0], 
                            rely = self.Placement['Export']['Check1'][1]
                            )
        
        self.var_exp_chk2 = tk.IntVar(value=1)
        self.exp_chk2 = ttk.Checkbutton(
                                        self.nb_tab_tab6, 
                                        text="Test Information", 
                                        variable=self.var_exp_chk2,
                                        style = "TCheckbutton" 
                                        )
        self.exp_chk2.place(
                            anchor='nw', 
                            relx = self.Placement['Export']['Check2'][0], 
                            rely = self.Placement['Export']['Check2'][1]
                            )
        
        self.var_exp_chk3 = tk.IntVar(value=1)
        self.exp_chk3 = ttk.Checkbutton(
                                        self.nb_tab_tab6, 
                                        text="Full Experimental Curves", 
                                        variable=self.var_exp_chk3,
                                        style = "TCheckbutton" 
                                        )
        self.exp_chk3.place(
                            anchor='nw', 
                            relx = self.Placement['Export']['Check3'][0], 
                            rely = self.Placement['Export']['Check3'][1]
                            )
        
        self.var_exp_chk4 = tk.IntVar(value=1)
        self.exp_chk4 = ttk.Checkbutton(
                                        self.nb_tab_tab6, 
                                        text="Reduced Experimental Curves", 
                                        variable=self.var_exp_chk4,
                                        style = "TCheckbutton" 
                                        )
        self.exp_chk4.place(
                            anchor='nw', 
                            relx = self.Placement['Export']['Check4'][0], 
                            rely = self.Placement['Export']['Check4'][1]
                            )
        
        self.var_exp_chk5 = tk.IntVar(value=1)
        self.exp_chk5 = ttk.Checkbutton(
                                        self.nb_tab_tab6, 
                                        text="Predicted Curves", 
                                        variable=self.var_exp_chk5,
                                        style = "TCheckbutton" 
                                        )
        self.exp_chk5.place(
                            anchor='nw', 
                            relx = self.Placement['Export']['Check5'][0], 
                            rely = self.Placement['Export']['Check5'][1]
                            )

        # Create Export Button
        self.exp_btn = ttk.Button(
                                    self.nb_tab_tab6, 
                                    text = "Export", 
                                    command = lambda:export(), 
                                    style = "Modern3.TButton",
                                    )
        self.exp_btn.place(
                                anchor = 'nw', 
                                relx = self.Placement['Export']['ButtonExp'][0],
                                rely = self.Placement['Export']['ButtonExp'][1], 
                                relwidth = self.Placement['Export']['ButtonExp'][2], 
                                relheight = self.Placement['Export']['ButtonExp'][3]
                                )

        # Set Flag
        self.exp_init = 0